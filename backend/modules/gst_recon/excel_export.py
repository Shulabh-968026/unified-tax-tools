"""GST Recon — Excel audit working-paper export.

Generates a multi-sheet XLSX from a run document with computed summary +
voucher-level invoice records. Sheets:

  1. Dashboard           — KPIs + status banners
  2. 12-Month Summary    — Outward + ITC reconciliation tables
  3. Outward Vouchers    — all monthly Books↔GSTR-1 match results, categorised
  4. Inward Vouchers     — all monthly Books↔GSTR-2B match results
  5. Pending Classification — unmapped ledger names from the mapping file
  6. Run Metadata        — FY, client, GSTIN, run id, uploads timeline

Designed to be the audit working-paper a CA can drop straight into their
file — every number in the dashboard cross-references the detailed sheets.
"""
from __future__ import annotations
import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ============================ Styles ==========================================
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL  = PatternFill("solid", fgColor="1F2937")  # gray-800
HEADER_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="F3F4F6")  # gray-100
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="111827")
TOTAL_FILL   = PatternFill("solid", fgColor="E5E7EB")  # gray-200
TOTAL_FONT   = Font(name="Calibri", size=10, bold=True)
WARN_FILL    = PatternFill("solid", fgColor="FEF3C7")  # amber-100
WARN_FONT    = Font(name="Calibri", size=10, color="92400E", bold=True)
OK_FILL      = PatternFill("solid", fgColor="D1FAE5")  # emerald-100
OK_FONT      = Font(name="Calibri", size=10, color="065F46", bold=True)
DANGER_FILL  = PatternFill("solid", fgColor="FEE2E2")  # red-100
DANGER_FONT  = Font(name="Calibri", size=10, color="991B1B", bold=True)
BODY_FONT    = Font(name="Calibri", size=10)
MONO_FONT    = Font(name="Consolas", size=10)

INR_FMT = '#,##0.00;(#,##0.00);"–"'
INT_FMT = "#,##0"


# ============================ Helpers =========================================
def _set_widths(ws, widths: List[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, row: int, headers: List[str], merge_first: bool = False):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 24


def _write_row(ws, row: int, values: List[Any], formats: List[str] = None,
               fill: PatternFill = None, font: Font = None):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = font or BODY_FONT
        cell.border = BORDER
        if formats and c <= len(formats) and formats[c - 1]:
            cell.number_format = formats[c - 1]
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if fill:
            cell.fill = fill


# ============================ Sheet builders ==================================
def _sheet_dashboard(ws, summary: Dict[str, Any], run: Dict[str, Any]):
    ws.title = "Dashboard"
    _set_widths(ws, [3, 32, 22, 22, 22, 22])

    # Title block
    ws["B2"] = "Reconciliation Health"
    ws["B2"].font = Font(name="Calibri", size=18, bold=True, color="111827")
    ws["B3"] = f"FY {run.get('fy', '')} · Client {run.get('client_id', '')} · Run {run.get('id', '')[:8]}"
    ws["B3"].font = Font(name="Calibri", size=10, color="6B7280")

    rows = summary.get("rows", [])
    totals = summary.get("totals", {})

    def _flagged(key: str) -> int:
        return sum(1 for r in rows if abs(r.get(key) or 0) >= 1)

    cards = [
        ("Books vs GSTR-1",     "Outward Turnover",
         totals.get("var_books_vs_r1_outward", 0),
         totals.get("r1_outward_taxable", 0),
         _flagged("var_books_vs_r1_outward")),
        ("GSTR-1 vs GSTR-3B",   "Outward Turnover",
         totals.get("var_r1_vs_r3b_outward", 0),
         totals.get("r3b_outward_taxable", 0),
         _flagged("var_r1_vs_r3b_outward")),
        ("Books vs GSTR-2B",    "Input Tax Credit",
         totals.get("var_books_vs_r2b_itc", 0),
         totals.get("r2b_itc_total", 0),
         _flagged("var_books_vs_r2b_itc")),
        ("GSTR-2B vs GSTR-3B",  "Input Tax Credit",
         totals.get("var_r2b_vs_r3b_itc", 0),
         totals.get("r3b_itc_total", 0),
         _flagged("var_r2b_vs_r3b_itc")),
    ]

    # Header row
    r = 5
    _write_header(ws, r, ["", "Comparison", "Variance (₹)", "Of Base (₹)", "% of Base", "Months Flagged"])
    r += 1
    for title, subtitle, value, base, flagged in cards:
        pct = (abs(value) / abs(base) * 100) if base else 0
        is_ok = flagged == 0
        is_danger = abs(value) > 100000
        fill = OK_FILL if is_ok else (DANGER_FILL if is_danger else WARN_FILL)
        font = OK_FONT if is_ok else (DANGER_FONT if is_danger else WARN_FONT)
        ws.cell(row=r, column=2, value=title).font = SECTION_FONT
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=3, value=value).number_format = INR_FMT
        ws.cell(row=r, column=4, value=base).number_format = INR_FMT
        ws.cell(row=r, column=5, value=pct / 100).number_format = "0.00%"
        ws.cell(row=r, column=6, value=f"{flagged} / 12")
        for c in range(3, 7):
            ws.cell(row=r, column=c).font = font
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
        ws.row_dimensions[r].height = 22
        r += 1

    # Status banner
    r += 1
    total_flagged = sum(c[4] for c in cards)
    if total_flagged == 0:
        ws.cell(row=r, column=2, value="ALL RECONCILED").fill = OK_FILL
        ws.cell(row=r, column=2).font = OK_FONT
    else:
        ws.cell(row=r, column=2, value=f"{total_flagged} MONTH-ISSUES FLAGGED — REVIEW DETAIL SHEETS").fill = WARN_FILL
        ws.cell(row=r, column=2).font = WARN_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28


def _sheet_summary(ws, summary: Dict[str, Any]):
    ws.title = "12-Month Summary"
    rows = summary.get("rows", [])
    totals = summary.get("totals", {})

    # Outward block
    ws.cell(row=1, column=1, value="OUTWARD TURNOVER (Books vs GSTR-1 vs GSTR-3B)").font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells("A1:F1")

    out_headers = ["Month", "Books", "GSTR-1", "GSTR-3B", "Books − R1", "R1 − R3B"]
    out_keys = ["books_outward_taxable", "r1_outward_taxable", "r3b_outward_taxable",
                "var_books_vs_r1_outward", "var_r1_vs_r3b_outward"]
    _write_header(ws, 2, out_headers)
    for i, row in enumerate(rows):
        r = 3 + i
        _write_row(ws, r, [row["month_label"]] + [row.get(k, 0) for k in out_keys],
                   formats=[None, INR_FMT, INR_FMT, INR_FMT, INR_FMT, INR_FMT],
                   fill=(SECTION_FILL if i % 2 else None))
    # Annual
    r_ann = 3 + len(rows)
    _write_row(ws, r_ann, ["Annual"] + [totals.get(k, 0) for k in out_keys],
               formats=[None, INR_FMT, INR_FMT, INR_FMT, INR_FMT, INR_FMT],
               fill=TOTAL_FILL, font=TOTAL_FONT)

    # ITC block (separated by blank row)
    r_itc_hdr = r_ann + 3
    ws.cell(row=r_itc_hdr, column=1,
            value="INPUT TAX CREDIT (Books vs GSTR-2B vs GSTR-3B Net)").font = SECTION_FONT
    ws.cell(row=r_itc_hdr, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r_itc_hdr, start_column=1, end_row=r_itc_hdr, end_column=6)

    itc_headers = ["Month", "Books", "GSTR-2B", "GSTR-3B (Net)", "Books − R2B", "R2B − R3B"]
    itc_keys = ["books_itc_total", "r2b_itc_total", "r3b_itc_total",
                "var_books_vs_r2b_itc", "var_r2b_vs_r3b_itc"]
    _write_header(ws, r_itc_hdr + 1, itc_headers)
    for i, row in enumerate(rows):
        r = r_itc_hdr + 2 + i
        _write_row(ws, r, [row["month_label"]] + [row.get(k, 0) for k in itc_keys],
                   formats=[None, INR_FMT, INR_FMT, INR_FMT, INR_FMT, INR_FMT],
                   fill=(SECTION_FILL if i % 2 else None))
    r_itc_ann = r_itc_hdr + 2 + len(rows)
    _write_row(ws, r_itc_ann, ["Annual"] + [totals.get(k, 0) for k in itc_keys],
               formats=[None, INR_FMT, INR_FMT, INR_FMT, INR_FMT, INR_FMT],
               fill=TOTAL_FILL, font=TOTAL_FONT)

    _set_widths(ws, [16, 18, 18, 18, 18, 18])


def _sheet_voucher_matches(ws, title: str, period_results: Dict[str, Dict[str, Any]],
                           portal_label: str):
    """Write Books↔Portal matches for every period this sheet's direction.
    period_results: {period_label: match_dict from match_invoices()}
    """
    ws.title = title
    headers = ["Month", "Category", "Party GSTIN", "Party Name",
               "Books #", f"{portal_label} #",
               "Books Total", f"{portal_label} Total", "Δ Total",
               "Books Date", f"{portal_label} Date", "Fuzzy", "Relaxed"]
    _write_header(ws, 1, headers)

    cat_fills = {
        "matched":          (None, None),
        "value_mismatch":   (WARN_FILL, WARN_FONT),
        "date_mismatch":    (PatternFill("solid", fgColor="DBEAFE"), Font(color="1E40AF", bold=True)),
        "missing_in_books": (DANGER_FILL, DANGER_FONT),
        "missing_in_portal":(DANGER_FILL, DANGER_FONT),
    }

    r = 2
    for period_label in sorted(period_results.keys()):
        m = period_results[period_label]
        for cat in ("matched", "value_mismatch", "date_mismatch",
                    "missing_in_portal", "missing_in_books"):
            for item in (m.get(cat) or []):
                fill, font = cat_fills[cat]
                if cat in ("matched", "value_mismatch", "date_mismatch"):
                    b = item.get("books") or {}
                    p = item.get("portal") or {}
                    row = [
                        period_label, cat,
                        b.get("party_gstin") or p.get("party_gstin") or "",
                        b.get("party_name") or p.get("party_name") or "",
                        b.get("voucher_no") or "",
                        p.get("invoice_no") or "",
                        b.get("total") or 0,
                        p.get("total") or 0,
                        item.get("value_diff") or 0,
                        item.get("books_date") or "",
                        item.get("portal_date") or "",
                        item.get("fuzzy_score") or "",
                        "Yes" if item.get("relaxed_match") else "",
                    ]
                else:
                    side = item if isinstance(item, dict) else {}
                    row = [
                        period_label, cat,
                        side.get("party_gstin", ""),
                        side.get("party_name", ""),
                        side.get("voucher_no", "") if cat == "missing_in_portal" else "",
                        side.get("invoice_no", "") if cat == "missing_in_books" else "",
                        side.get("total", 0) if cat == "missing_in_portal" else 0,
                        side.get("total", 0) if cat == "missing_in_books" else 0,
                        0,
                        side.get("date", "") if cat == "missing_in_portal" else "",
                        side.get("date", "") if cat == "missing_in_books" else "",
                        "", "",
                    ]
                _write_row(ws, r, row,
                           formats=[None, None, None, None, None, None,
                                    INR_FMT, INR_FMT, INR_FMT,
                                    None, None, None, None],
                           fill=fill, font=font)
                r += 1

    if r == 2:
        ws.cell(row=2, column=1, value="No vouchers in this section").font = Font(italic=True, color="6B7280")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    _set_widths(ws, [10, 20, 22, 30, 18, 22, 14, 14, 12, 12, 12, 8, 8])
    ws.freeze_panes = "A2"


def _sheet_partywise(ws, title: str, partywise: Dict[str, Any], portal_label: str,
                     direction: str = "inward"):
    """Annual Party-wise Summary — one row per GSTIN, sorted by largest variance first.
    Direction-aware columns:
      inward  → ITC (tax) values per party (Books vs GSTR-2B)
      outward → Taxable values per party  (Books vs GSTR-1)
    Columns: Party GSTIN | Party Name | Books | Portal | Books − Portal."""
    ws.title = title
    rows = partywise.get("rows", []) or []
    totals = partywise.get("totals", {}) or {}

    if direction == "inward":
        books_key, portal_key, diff_key = "books_tax", "portal_tax", "diff_tax"
        value_label = "ITC"
    else:
        books_key, portal_key, diff_key = "books_taxable", "portal_taxable", "diff_taxable"
        value_label = "Taxable Value"

    headers = ["Party GSTIN", "Party Name",
               f"Books ({value_label})", f"{portal_label} ({value_label})",
               f"Books − {portal_label}"]
    _write_header(ws, 1, headers)

    for i, r in enumerate(rows):
        _write_row(ws, 2 + i, [
            r["party_gstin"], r["party_name"],
            r.get(books_key, 0), r.get(portal_key, 0), r.get(diff_key, 0),
        ], formats=[None, None, INR_FMT, INR_FMT, INR_FMT],
           fill=(SECTION_FILL if i % 2 else None))

    if rows:
        annual_row = 2 + len(rows)
        _write_row(ws, annual_row, [
            "ANNUAL TOTAL", f"{len(rows)} parties",
            totals.get(books_key, 0),
            totals.get(portal_key, 0),
            totals.get(diff_key, 0),
        ], formats=[None, None, INR_FMT, INR_FMT, INR_FMT],
           fill=TOTAL_FILL, font=TOTAL_FONT)
    else:
        ws.cell(row=2, column=1, value="No party-wise records").font = Font(italic=True, color="6B7280")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    _set_widths(ws, [22, 36, 18, 20, 18])
    ws.freeze_panes = "A2"


def _sheet_unmapped(ws, run: Dict[str, Any]):
    ws.title = "Pending Classification"
    unmapped = run.get("mapping_unmapped_ledgers") or []
    ws.cell(row=1, column=1, value="LEDGERS NOT CLASSIFIED BY MAPPING FILE").font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells("A1:B1")
    ws.cell(row=2, column=1, value="These tax-related ledgers in the Books JSON do not have a Head/Group Parent in the Mapping XLSX. Books figures for transactions using these ledgers will NOT appear in the reconciliation until the Mapping is updated.").font = Font(italic=True, color="6B7280", size=9)
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 36

    _write_header(ws, 4, ["#", "Ledger Name"])
    for i, name in enumerate(unmapped):
        _write_row(ws, 5 + i, [i + 1, name])
    if not unmapped:
        ws.cell(row=5, column=1, value="✓ All ledgers classified").font = OK_FONT
        ws.cell(row=5, column=1).fill = OK_FILL
        ws.merge_cells("A5:B5")

    _set_widths(ws, [6, 50])


def _sheet_metadata(ws, run: Dict[str, Any]):
    ws.title = "Run Metadata"
    rows = [
        ("Run ID", run.get("id", "")),
        ("Name", run.get("name", "")),
        ("Client ID", run.get("client_id", "")),
        ("Financial Year", run.get("fy", "")),
        ("Status", run.get("status", "")),
        ("Created At (UTC)", run.get("created_at", "")),
        ("Has Books", "Yes" if run.get("has_books") else "No"),
        ("Has Mapping", "Yes" if run.get("has_mapping") else "No"),
        ("Mapping Filename", run.get("mapping_filename", "—")),
        ("Mapping Rows", run.get("mapping_row_count", "")),
        ("Total Files Uploaded", len(run.get("files", []))),
        ("Pending Classification (count)", len(run.get("mapping_unmapped_ledgers") or [])),
    ]
    ws.cell(row=1, column=1, value="RUN METADATA").font = SECTION_FONT
    ws.cell(row=1, column=1).fill = SECTION_FILL
    ws.merge_cells("A1:B1")
    for i, (k, v) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=1).fill = TOTAL_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2, value=v).font = MONO_FONT
        ws.cell(row=r, column=2).border = BORDER

    # Files table
    r = 3 + len(rows) + 2
    ws.cell(row=r, column=1, value="UPLOADED FILES").font = SECTION_FONT
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    _write_header(ws, r + 1, ["Bucket", "Filename"])
    for i, f in enumerate(run.get("files", [])):
        _write_row(ws, r + 2 + i, [f.get("bucket", ""), f.get("filename", "")])

    _set_widths(ws, [32, 60])


# ============================ Public entry ====================================
def build_workbook(run: Dict[str, Any], summary: Dict[str, Any],
                   outward_matches: Dict[str, Dict[str, Any]],
                   inward_matches: Dict[str, Dict[str, Any]],
                   partywise_outward: Optional[Dict[str, Any]] = None,
                   partywise_inward: Optional[Dict[str, Any]] = None) -> bytes:
    """Build the full audit workbook and return XLSX bytes.

    Sheet order:
      1. Dashboard
      2. Annual Party-wise (Outward) — main reconciliation working-paper for sales
      3. Annual Party-wise (Inward)  — main reconciliation working-paper for ITC
      4. 12-Month Summary           — month-on-month detail
      5. Outward Vouchers           — voucher-level Books↔GSTR-1
      6. Inward Vouchers            — voucher-level Books↔GSTR-2B
      7. Pending Classification
      8. Run Metadata
    """
    wb = Workbook()
    _sheet_dashboard(wb.active, summary, run)
    if partywise_outward:
        _sheet_partywise(wb.create_sheet(), "Annual Party-wise (Outward)",
                         partywise_outward, "GSTR-1", direction="outward")
    if partywise_inward:
        _sheet_partywise(wb.create_sheet(), "Annual Party-wise (Inward)",
                         partywise_inward, "GSTR-2B", direction="inward")
    _sheet_summary(wb.create_sheet(), summary)
    _sheet_voucher_matches(wb.create_sheet(), "Outward Vouchers", outward_matches, "GSTR-1")
    _sheet_voucher_matches(wb.create_sheet(), "Inward Vouchers", inward_matches, "GSTR-2B")
    _sheet_unmapped(wb.create_sheet(), run)
    _sheet_metadata(wb.create_sheet(), run)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
