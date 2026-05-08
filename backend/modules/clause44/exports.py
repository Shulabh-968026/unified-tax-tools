"""Excel export builder for single + consolidated runs.

Workbook layout (user-requested):
  1.  Clause 44 Summary   — Aggregate + per-ledger six-column pivot (the
                            "consolidated pivotable list").
  2.  Reconciliation      — Books → Schedule tie-out.
  3.  Col 3 · Exempt      — vouchers classified into Col 3.
  4.  Col 4 · Composition — vouchers classified into Col 4.
  5.  Col 5 · Other Reg.  — vouchers classified into Col 5 (ITC).
  6.  Col 7 · Unregistered — vouchers classified into Col 7.

Each cohort sheet can be pivoted in Excel (all columns present, no merged
cells inside the data region) so auditors can slice further without going
back to the app.
"""
import io
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi.responses import StreamingResponse

INDIAN_FMT = "[>=10000000]##\\,##\\,##\\,##0.00;[>=100000]##\\,##\\,##0.00;##,##0.00"

BOLD_WHITE = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="0F172A")
THIN = Side(style="thin", color="D4D4D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(bold=True, size=14)

BUCKET_META = [
    ("col3", "Col 3 · Exempt",        "Col 3 — Exempt"),
    ("col4", "Col 4 · Composition",   "Col 4 — Composition"),
    ("col5", "Col 5 · Other Reg ITC", "Col 5 — Other Registered (ITC)"),
    ("col7", "Col 7 · Unregistered",  "Col 7 — Unregistered"),
    ("col8", "Col 8 · Excluded",      "Col 8 — Excluded from Col 3-7 reporting"),
]
BUCKET_LABEL = {k: long for k, _, long in BUCKET_META}

# Sub-bucket labels within Col 8 — mirror the recon ICAI buckets so the
# working paper tells the reviewer why each line was excluded.
COL8_SUB_BUCKETS = [
    ("non_cash", "Non-cash charges"),
    ("sch3", "Schedule III items"),
    ("money", "Money / Securities"),
    ("capex_addback", "Capex add-back"),
    ("other", "Other exclusions"),
]


def _style_header_row(ws, row_idx: int, n_cols: int):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.font = BOLD_WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _apply_indian_fmt(ws, min_row: int, cols: List[int]):
    max_col = max(cols)
    min_col = min(cols)
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        for c in row:
            if c.column in cols and isinstance(c.value, (int, float)):
                c.number_format = INDIAN_FMT


def _write_summary_sheet(wb, run: Dict[str, Any]):
    summary = run.get("summary", {}) or {}
    by_ledger = run.get("by_ledger", {}) or {}

    ws = wb.active
    ws.title = "Clause 44 Summary"

    ws["A1"] = "Clause 44 of Form 3CD — Expenditure Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = f"Company: {run.get('company_name','')}"
    ws["A3"] = f"Generated: {run.get('generated_at','')}"

    headers = [
        "Particulars",
        "Col 2: Total Expenditure",
        "Col 3: Exempt Supply",
        "Col 4: Composition Dealer",
        "Col 5: Other Registered (ITC)",
        "Col 6: Total (3+4+5)",
        "Col 7: Unregistered",
        "Col 8: Excluded",
    ]

    # Aggregate row — Col 2 is now the gross total per books.
    ws.append([])
    ws.append(headers)
    _style_header_row(ws, ws.max_row, len(headers))
    ws.append([
        "Aggregate (All Expenditure)",
        summary.get("col2_total", 0),
        summary.get("col3", 0),
        summary.get("col4", 0),
        summary.get("col5", 0),
        summary.get("col6", 0),
        summary.get("col7", 0),
        summary.get("col8", 0),
    ])

    # Per-ledger pivot (7 data columns).
    ws.append([])
    ws.append(["Per-Ledger Breakdown (seven-column pivot)"])
    ws.cell(ws.max_row, 1).font = TITLE_FONT
    ws.append(headers)
    pivot_header_row = ws.max_row
    _style_header_row(ws, pivot_header_row, len(headers))

    for lname, row in sorted(by_ledger.items(), key=lambda kv: kv[1].get("total", 0), reverse=True):
        col3 = row.get("col3", 0) or 0
        col4 = row.get("col4", 0) or 0
        col5 = row.get("col5", 0) or 0
        col7 = row.get("col7", 0) or 0
        col8 = row.get("col8", 0) or 0
        total = row.get("total", 0) or (col3 + col4 + col5 + col7 + col8)
        ws.append([
            lname,
            total,
            col3, col4, col5,
            col3 + col4 + col5,
            col7, col8,
        ])

    ws.column_dimensions["A"].width = 36
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 22
    _apply_indian_fmt(ws, min_row=5, cols=[2, 3, 4, 5, 6, 7, 8])

    ws.freeze_panes = ws.cell(row=pivot_header_row + 1, column=2)


def _write_recon_sheet(wb, run: Dict[str, Any]):
    recon = run.get("recon") or {}
    ws = wb.create_sheet("Reconciliation")
    ws["A1"] = "Reconciliation — Books to Clause 44 (ICAI Para 79.4)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.append([])
    ws.append(["Particulars", "Amount"])
    _style_header_row(ws, ws.max_row, 2)

    pl_total = recon.get("pl_total")
    capex_total = recon.get("capex_total")
    reportable = recon.get("reportable_total")

    # ICAI 5-line format — render only when the new fields are present.
    if pl_total is not None and capex_total is not None:
        ws.append(["Total Expenditure as per Profit & Loss", float(pl_total or 0)])
        ws.append(["+ Capital expenditure additions (ICAI Para 79.18)", float(capex_total or 0)])

        buckets = [
            ("non_cash", "Less: Non-cash charges (depreciation, provisions, fair-value losses)"),
            ("sch3",     "Less: Schedule III items (salary, wages, PF/ESI, gratuity, dividend declared, sale of land/building)"),
            ("money",    "Less: Money / Securities transactions (interest, TDS, investments, share transactions)"),
            ("other",    "Less: Other auditor-elected exclusions"),
        ]
        for key, label in buckets:
            lines = recon.get(f"{key}_lines") or []
            total = recon.get(f"{key}_total") or 0
            ws.append([label, -float(total or 0)])
            header_row = ws.max_row
            ws.cell(row=header_row, column=1).font = Font(bold=True)
            for line in lines:
                ws.append([f"   • {line['name']}", -float(line.get('amount') or 0)])

        ws.append(["= Reportable Expenditure (Col 2 of Clause 44)", float(reportable or 0)])
        last_row = ws.max_row
        for col in (1, 2):
            c = ws.cell(row=last_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F3F4F1")
            c.border = BORDER
    else:
        # Legacy single-bucket recon (pre-Release-1 runs).
        ws.append(["Total Expenditure as per Books", recon.get("total_books", 0)])
        ws.append(["Less : Expenditures excluded from Clause 44 Report", None])
        for line in (recon.get("excluded_lines") or []):
            ws.append([f"   • {line['name']}", -float(line.get("amount") or 0)])
        ws.append(["Expenditure as per Clause 44 Report", recon.get("balance", 0)])
        last_row = ws.max_row
        for col in (1, 2):
            c = ws.cell(row=last_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F3F4F1")
            c.border = BORDER

    # Disclaimer block — appended regardless of recon shape.
    disclaimer = run.get("disclaimer_text") or ""
    if disclaimer:
        ws.append([])
        ws.append(["Disclaimer"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.append([disclaimer])
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ws.max_row].height = 90
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=2)

    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 22
    _apply_indian_fmt(ws, min_row=4, cols=[2])


def _write_cohort_sheet(wb, sheet_name: str, long_label: str, txns: List[Dict[str, Any]]):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{long_label} — Transaction Breakup"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    has_division = any(t.get("division_name") for t in txns)
    # ICAI Para 79.20 illustrative working-paper columns:
    #   Date | Voucher Type | Voucher No | [Division] | Ledger | Party |
    #   Party GSTIN | Party Reg | Country | RCM | Amount |
    #   Value Eligible for ITC | Reason for NIL GST / Notes | Auditor Remarks
    headers = ["Date", "Voucher Type", "Voucher No"]
    if has_division:
        headers.append("Division")
    headers += [
        "Ledger", "Party", "Party GSTIN", "Party Reg.", "Country",
        "RCM", "Amount", "Value Eligible for ITC",
        "Reason for NIL GST / Classification Notes", "Auditor Remarks",
    ]

    ws.append([])
    ws.append(headers)
    header_row = ws.max_row
    _style_header_row(ws, header_row, len(headers))

    # Dynamically compute column indices so division-injection doesn't break them.
    idx = {h: i + 1 for i, h in enumerate(headers)}
    amount_col = idx["Amount"]
    itc_col = idx["Value Eligible for ITC"]

    total = 0.0
    total_itc = 0.0
    for t in txns:
        row = [t.get("date", ""), t.get("voucher_type", ""), t.get("voucher_number", "")]
        if has_division:
            row.append(t.get("division_name", "—"))
        amt = float(t.get("amount") or 0)
        total += amt
        # Per Para 79.20 "Value eligible for ITC": amount only if a proper ITC-
        # ledger entry sat on the voucher AND the voucher isn't RCM AND the
        # line isn't an exempt-tagged Input A contribution.  The engine
        # already stores these flags on each line.
        itc_eligible = 0.0
        if t.get("has_itc_ledger") and not t.get("is_rcm") and t.get("col3_source") != "input_a":
            itc_eligible = amt
        total_itc += itc_eligible
        row += [
            t.get("ledger_name", ""),
            t.get("party_name", ""),
            t.get("party_gstin", ""),
            t.get("party_reg", ""),
            t.get("party_country", ""),
            "Yes" if t.get("is_rcm") else "",
            amt,
            itc_eligible,
            t.get("reason", ""),
            "",  # blank column for the auditor to note remarks
        ]
        ws.append(row)

    if txns:
        footer = ["Total"] + [""] * (len(headers) - 1)
        footer[amount_col - 1] = total
        footer[itc_col - 1] = total_itc
        ws.append(footer)
        last_row = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=last_row, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F3F4F1")
            c.border = BORDER
    else:
        ws.append(["— No vouchers classified into this cohort —"])

    # Column widths — keep the meta columns tight, the narratives wide.
    base_widths = [12, 14, 14]
    if has_division:
        base_widths.append(16)
    base_widths += [28, 28, 18, 14, 14, 8, 16, 18, 60, 28]
    for i, w in enumerate(base_widths[:len(headers)], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    _apply_indian_fmt(ws, min_row=header_row + 1, cols=[amount_col, itc_col])

    # Freeze header row + auto-filter (pivot-friendly out of the box).
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"{ws.cell(row=header_row, column=1).coordinate}:{ws.cell(row=header_row, column=len(headers)).coordinate}"


def _write_col8_sheet(wb, run: Dict[str, Any]):
    """Col 8 — Excluded sheet with ICAI sub-buckets (option ii).

    Vouchers grouped by Non-cash / Sch III / Money / Capex add-back /
    Other so the reviewer can see *why* each line was excluded.  Within
    each group, the standard Para 79.20 column set applies.
    """
    txns = [t for t in (run.get("transactions") or []) if t.get("bucket") == "col8"]
    recon = run.get("recon") or {}

    # Build lookup: excluded ledger → sub-bucket (via recon line detail)
    ledger_to_sub: Dict[str, str] = {}
    for sub_key, _ in COL8_SUB_BUCKETS:
        for line in recon.get(f"{sub_key}_lines") or []:
            ledger_to_sub[line["name"]] = sub_key
    # Anything not in the recon lookup → "other"
    sub_tx: Dict[str, List[Dict[str, Any]]] = {k: [] for k, _ in COL8_SUB_BUCKETS}
    for t in txns:
        sub = ledger_to_sub.get(t.get("ledger_name", ""), "other")
        sub_tx.setdefault(sub, []).append(t)

    ws = wb.create_sheet("Col 8 · Excluded")
    ws["A1"] = "Col 8 — Excluded from Col 3-7 reporting · ICAI recon sub-buckets"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    has_division = any(t.get("division_name") for t in txns)
    headers = ["Date", "Voucher Type", "Voucher No"]
    if has_division:
        headers.append("Division")
    headers += [
        "Ledger", "Party", "Party GSTIN", "Party Reg.", "Country",
        "RCM", "Amount", "Classification Notes", "Auditor Remarks",
    ]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    amount_col = idx["Amount"]

    ws.append([])

    overall_total = 0.0
    for sub_key, sub_label in COL8_SUB_BUCKETS:
        rows = sub_tx.get(sub_key) or []
        if not rows:
            continue
        # Sub-bucket header band
        ws.append([f"{sub_label}  ·  sub-bucket"])
        band = ws.cell(row=ws.max_row, column=1)
        band.font = Font(bold=True, color="FFFFFF")
        band.fill = PatternFill("solid", fgColor="0F172A")
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(headers))

        # Column headers
        ws.append(headers)
        hr = ws.max_row
        _style_header_row(ws, hr, len(headers))

        subtotal = 0.0
        for t in rows:
            row = [t.get("date", ""), t.get("voucher_type", ""), t.get("voucher_number", "")]
            if has_division:
                row.append(t.get("division_name", "—"))
            amt = float(t.get("amount") or 0)
            subtotal += amt
            row += [
                t.get("ledger_name", ""),
                t.get("party_name", ""),
                t.get("party_gstin", ""),
                t.get("party_reg", ""),
                t.get("party_country", ""),
                "Yes" if t.get("is_rcm") else "",
                amt,
                t.get("reason", ""),
                "",
            ]
            ws.append(row)

        # Sub-total row
        foot = ["Sub-total · " + sub_label] + [""] * (len(headers) - 1)
        foot[amount_col - 1] = subtotal
        ws.append(foot)
        last = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=last, column=col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="F3F4F1")
            c.border = BORDER
        overall_total += subtotal

        _apply_indian_fmt(ws, min_row=hr + 1, cols=[amount_col])
        ws.append([])  # blank row between groups

    # Overall total
    if overall_total:
        foot = ["Col 8 Total · Excluded expenditure"] + [""] * (len(headers) - 1)
        foot[amount_col - 1] = overall_total
        ws.append(foot)
        last = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=last, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F172A")
            c.border = BORDER

    if not txns:
        ws.append(["— No vouchers classified into Col 8 (no exclusions elected) —"])

    # Column widths
    base_widths = [12, 14, 14]
    if has_division:
        base_widths.append(16)
    base_widths += [28, 28, 18, 14, 14, 8, 16, 60, 28]
    for i, w in enumerate(base_widths[:len(headers)], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=2, column=1)


# ─────────────────────────────────────────────────────────────────────────
# Mapping Snapshot — pre-generate working paper that captures the engine's
# current auto-suggestions for the three pools (Exempt / ITC / Exclusions)
# alongside what the auditor has currently ticked.  Lets the auditor
# review / share the proposed selections before clicking *Generate*.
# ─────────────────────────────────────────────────────────────────────────
def _yn(v):
    return "Yes" if bool(v) else "No"


def _write_mapping_meta(ws, run: Dict[str, Any], title: str):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    meta = [
        ("Company", run.get("company_name") or ""),
        ("Client", run.get("client_name") or ""),
        ("Period", run.get("period") or ""),
        ("Division", run.get("division_name") or "—"),
        ("Run ID", run.get("run_id") or ""),
        ("Snapshot taken", run.get("snapshot_at") or ""),
    ]
    for label, value in meta:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])


def _write_mapping_exempt(wb, run, exempt_selection: set):
    ws = wb.create_sheet("Exempt Purchases")
    _write_mapping_meta(ws, run, "Pool 1 · Exempt Purchases — Mapping Snapshot")

    headers = [
        "Ledger Name", "Subhead", "Group Parent", "Head",
        "Closing Balance",
        "Vouchers", "ITC-Overlap Vouchers", "Demoted by ITC Cross-Check?",
        "Auto-Suggested?", "Currently Selected?",
    ]
    ws.append(headers)
    header_row = ws.max_row
    _style_header_row(ws, header_row, len(headers))

    rows = run.get("exempt_ledgers") or []
    rows = sorted(rows, key=lambda r: (not r.get("suggested"), (r.get("name") or "").lower()))
    for r in rows:
        ws.append([
            r.get("name") or "",
            r.get("subhead") or "",
            r.get("group_parent") or "",
            r.get("head") or "",
            r.get("closing_balance"),
            int(r.get("total_vouchers") or 0),
            int(r.get("itc_overlap_vouchers") or 0),
            _yn(r.get("itc_overlap_demoted")),
            _yn(r.get("suggested")),
            _yn((r.get("name") or "") in exempt_selection),
        ])

    if not rows:
        ws.append(["— No exempt-purchase candidates in this run —"])

    widths = [40, 26, 26, 26, 18, 12, 18, 22, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    _apply_indian_fmt(ws, min_row=header_row + 1, cols=[5])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_mapping_itc(wb, run, itc_selection: set):
    ws = wb.create_sheet("ITC Ledgers")
    _write_mapping_meta(ws, run, "Pool 2 · ITC Ledgers — Mapping Snapshot")

    headers = [
        "Ledger Name", "Subhead", "Group Parent", "Head",
        "Closing Balance", "Kind", "Kind Source",
        "Purchase Vouchers", "Sales Vouchers", "Usage Conflict?",
        "In Default View?", "Auto-Suggested?", "Currently Selected?",
    ]
    ws.append(headers)
    header_row = ws.max_row
    _style_header_row(ws, header_row, len(headers))

    # Use the full BS-side universe so the snapshot includes ledgers the
    # default view hides (e.g. ITC ledgers with non-standard subheads).
    rows = run.get("itc_ledgers_all_bs") or run.get("itc_ledgers") or []
    rows = sorted(
        rows,
        key=lambda r: (
            not r.get("suggested"),
            not r.get("in_default_view", True),
            (r.get("name") or "").lower(),
        ),
    )
    for r in rows:
        ws.append([
            r.get("name") or "",
            r.get("subhead") or "",
            r.get("group_parent") or "",
            r.get("head") or "",
            r.get("closing_balance"),
            (r.get("kind") or "other"),
            (r.get("kind_source") or "—"),
            int(r.get("n_purchase") or 0),
            int(r.get("n_sales") or 0),
            _yn(r.get("usage_conflict")),
            _yn(r.get("in_default_view", True)),
            _yn(r.get("suggested")),
            _yn((r.get("name") or "") in itc_selection),
        ])

    if not rows:
        ws.append(["— No ITC candidates in this run —"])

    widths = [40, 26, 26, 26, 18, 12, 14, 14, 14, 14, 16, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    _apply_indian_fmt(ws, min_row=header_row + 1, cols=[5])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _write_mapping_exclusions(wb, run, exclusion_selection: set, exclusion_categories: Dict[str, str]):
    ws = wb.create_sheet("Exclusions")
    _write_mapping_meta(ws, run, "Pool 3 · Exclusions — Mapping Snapshot")

    headers = [
        "Ledger Name", "Subhead", "Group Parent", "Head",
        "Closing Balance", "Recon Role", "Auto-Suggested?",
        "Currently Selected?", "Recon Bucket (override)",
    ]
    ws.append(headers)
    header_row = ws.max_row
    _style_header_row(ws, header_row, len(headers))

    rows = run.get("exclusion_ledgers") or []
    rows = sorted(rows, key=lambda r: (not r.get("suggested"), (r.get("name") or "").lower()))
    for r in rows:
        name = r.get("name") or ""
        ws.append([
            name,
            r.get("subhead") or "",
            r.get("group_parent") or "",
            r.get("head") or "",
            r.get("closing_balance"),
            r.get("recon_role") or "subtract",
            _yn(r.get("suggested")),
            _yn(name in exclusion_selection),
            exclusion_categories.get(name, "—"),
        ])

    if not rows:
        ws.append(["— No exclusion candidates in this run —"])

    widths = [40, 26, 26, 26, 18, 14, 16, 18, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    _apply_indian_fmt(ws, min_row=header_row + 1, cols=[5])
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_mapping_export_response(run: Dict[str, Any], fname_prefix: str):
    """Pre-generate Mapping Snapshot — 3 sheets (Exempt / ITC / Exclusions)
    capturing the engine's auto-suggestions plus the auditor's current
    ticks.  Available from the Mapping step onwards (no Generate required).
    """
    itc_selection = set(run.get("itc_selection") or [])
    exempt_selection = set(run.get("exempt_selection") or [])
    exclusion_selection = set(run.get("exclusion_selection") or [])
    exclusion_categories = run.get("exclusion_categories") or {}

    wb = openpyxl.Workbook()
    # Replace the auto-created default sheet with our first sheet.
    default = wb.active
    wb.remove(default)

    _write_mapping_exempt(wb, run, exempt_selection)
    _write_mapping_itc(wb, run, itc_selection)
    _write_mapping_exclusions(wb, run, exclusion_selection, exclusion_categories)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{fname_prefix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def build_export_response(run: Dict[str, Any], fname_prefix: str):
    txns = run.get("transactions", []) or []

    wb = openpyxl.Workbook()

    # Sheet 1 — Summary + per-ledger 7-col pivot
    _write_summary_sheet(wb, run)

    # Sheet 2 — Reconciliation (ICAI 5-line format)
    _write_recon_sheet(wb, run)

    # Sheets 3-6 — Col 3/4/5/7 cohort sheets with Para 79.20 columns
    for key, short, long_label in BUCKET_META:
        if key == "col8":
            continue  # handled separately with sub-bucket grouping
        cohort_txns = [t for t in txns if t.get("bucket") == key]
        _write_cohort_sheet(wb, short, long_label, cohort_txns)

    # Sheet 7 — Col 8 · Excluded (sub-bucketed per ICAI recon)
    _write_col8_sheet(wb, run)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{fname_prefix}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
