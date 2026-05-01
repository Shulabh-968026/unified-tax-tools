"""Excel round-trip for Opening WDV by Block.

3CD JSON only carries opening WDV at the **rate level**, but the IT
Depreciation working needs sub-block resolution (e.g. "15% Block – P&M"
vs "15% Block – Vehicles" both at 15%). This module ships an editable
workbook with one row per active block_label so the auditor can fill
sub-block opening values, and a parser that re-imports the same workbook.

The 3CD JSON is a separate, optional validation step — see
`validate_against_3cd` below.

Workflow:
    GET  /runs/{rid}/block-opening/export.xlsx
        → one sheet, one row per active block_label, pre-populated with
          current `fa_block_opening` values (0 when none yet).

    POST /runs/{rid}/block-opening/import.xlsx
        → parse, upsert each block's opening_wdv + description with
          source="manual_xlsx", return a JSON summary.

    POST /runs/{rid}/block-opening/validate-3cd
        → parse the optional prior-year 3CD JSON, sum current openings
          by rate, and return a per-rate diff so the auditor can sign off
          on the "sub-block totals tie back to 3CD" check.

Editable cells on import: opening_wdv, description.
Block label + Rate are LOCKED — the canonical block_label survives via
the hidden first column so cosmetic edits to the visible label cell
don't break the import.
"""
from __future__ import annotations
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---- Styling -----------------------------------------------------------
GREY = PatternFill("solid", fgColor="EDEDE7")
YELLOW = PatternFill("solid", fgColor="FFF8CC")
HEADER = PatternFill("solid", fgColor="1A1A1A")
THIN = Side(border_style="thin", color="B5B5B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HEADER
    c.font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c


def _locked_cell(ws, row, col, value, *, mono=False, num=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = GREY
    c.alignment = Alignment(
        horizontal=("right" if num else ("center" if mono else "left")),
        vertical="center",
    )
    c.font = Font(name=("JetBrains Mono" if mono or num else "Inter"), size=10)
    c.border = BORDER
    if num:
        c.number_format = '#,##0.00;(#,##0.00);"–"'
    return c


def _editable_cell(ws, row, col, value, *, num=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = YELLOW
    c.alignment = Alignment(horizontal=("right" if num else "left"), vertical="center")
    c.font = Font(name=("JetBrains Mono" if num else "Inter"), size=10)
    c.border = BORDER
    if num:
        c.number_format = '#,##0.00;(#,##0.00);"–"'
    return c


# ---- Export ------------------------------------------------------------
def build_workbook(run: Dict[str, Any], rows: List[Dict[str, Any]]) -> bytes:
    """`rows` shape: [{block_label, rate, opening_wdv, description, source}]
    sorted by descending rate. Returns the .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Opening WDV"

    # Hidden column A — canonical block_label primary key for round-trip
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["A"].width = 32

    # Visible columns
    widths = {"B": 42, "C": 8, "D": 20, "E": 50}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    # Title strip (rows 1-2)
    ws.cell(row=1, column=2, value=f"Opening WDV — {run.get('client_name') or 'Run'} · FY {run.get('fy_label') or run.get('fy_end','')}")
    ws.cell(row=1, column=2).font = Font(name="Inter", size=13, bold=True)
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=5)
    ws.cell(row=2, column=2, value="Edit yellow cells only · Block + Rate are locked · Save & re-import via Compute tab.")
    ws.cell(row=2, column=2).font = Font(name="Inter", size=10, italic=True, color="52524E")
    ws.merge_cells(start_row=2, end_row=2, start_column=2, end_column=5)

    # Totals strip (row 3) — driven by SUM formulas so the auditor sees
    # their edits reflected live in Excel.
    _hdr_cell(ws, 3, 1, "addition_id")    # hidden — kept to mirror the additions workbook ergonomics
    _hdr_cell(ws, 3, 2, "TOTAL")
    _hdr_cell(ws, 3, 3, "")
    end_row = len(rows) + 4
    total = ws.cell(row=3, column=4, value=f"=SUM(D5:D{end_row})")
    total.fill = GREY
    total.font = Font(name="JetBrains Mono", size=11, bold=True)
    total.alignment = Alignment(horizontal="right", vertical="center")
    total.number_format = '#,##0.00;(#,##0.00);"–"'
    total.border = BORDER
    ws.cell(row=3, column=5, value="").fill = GREY

    # Header row (row 4)
    _hdr_cell(ws, 4, 1, "block_label_key")  # hidden
    _hdr_cell(ws, 4, 2, "Block")
    _hdr_cell(ws, 4, 3, "Rate %")
    _hdr_cell(ws, 4, 4, "Opening WDV (₹)")
    _hdr_cell(ws, 4, 5, "Note (optional)")
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "B5"

    # Data rows (row 5+)
    for i, r in enumerate(rows):
        rr = 5 + i
        # hidden canonical key
        kc = ws.cell(row=rr, column=1, value=r["block_label"])
        kc.font = Font(name="JetBrains Mono", size=9, color="999999")
        # locked block label + rate
        _locked_cell(ws, rr, 2, r["block_label"])
        _locked_cell(ws, rr, 3, float(r.get("rate") or 0), mono=True).number_format = "0.0%/100"  # display 15%
        # 0.0%/100 not standard — fall back to plain numeric:
        ws.cell(row=rr, column=3).number_format = "0.0\\%"
        # editable Opening WDV + Note
        _editable_cell(ws, rr, 4, float(r.get("opening_wdv") or 0), num=True)
        _editable_cell(ws, rr, 5, r.get("description") or "")

    # Footer note
    foot = end_row + 2
    ws.cell(row=foot, column=2, value="Source priority on import: manual_xlsx > prior_3cd > prior_run > manual.")
    ws.cell(row=foot, column=2).font = Font(name="Inter", size=9, italic=True, color="6B6B66")
    ws.merge_cells(start_row=foot, end_row=foot, start_column=2, end_column=5)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


# ---- Import ------------------------------------------------------------
def parse_workbook(file_bytes: bytes, valid_blocks: List[str]) -> Dict[str, Any]:
    """Parse a previously-exported workbook. Returns:
        {
          "rows": [{block_label, opening_wdv, description}],
          "unknown_blocks": ["raw label", ...],   # rows we couldn't map
          "errors": ["...", ...],
        }
    Block matching uses the hidden canonical key (column A); falls back
    to the visible Block label (column B) when A is empty.
    """
    valid = set(valid_blocks)
    out_rows: List[Dict[str, Any]] = []
    unknown: List[str] = []
    errors: List[str] = []

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:  # noqa: BLE001
        return {"rows": [], "unknown_blocks": [], "errors": [f"Could not open workbook: {e}"]}

    if "Opening WDV" not in wb.sheetnames:
        # Fall back to the active sheet — be lenient about renames
        ws = wb.active
    else:
        ws = wb["Opening WDV"]

    # Headers expected on row 4. Walk from row 5.
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 4:
            continue
        canonical = (row[0] or "").strip() if isinstance(row[0], str) else ""
        visible = (row[1] or "").strip() if isinstance(row[1], str) else ""
        # Footer / merged informational rows have no canonical key — skip.
        if not canonical:
            continue
        block_label = canonical or visible
        if block_label not in valid:
            unknown.append(block_label)
            continue
        try:
            opening = float(row[3]) if row[3] is not None else 0.0
        except (TypeError, ValueError):
            errors.append(f"{block_label}: non-numeric opening WDV {row[3]!r}")
            continue
        desc = (row[4] or "").strip() if len(row) > 4 and isinstance(row[4], str) else ""
        out_rows.append({
            "block_label": block_label,
            "opening_wdv": round(opening, 2),
            "description": desc,
        })

    return {"rows": out_rows, "unknown_blocks": unknown, "errors": errors}


# ---- 3CD validation ----------------------------------------------------
def validate_against_3cd(
    current_openings: List[Dict[str, Any]],
    cd_rate_rows: List[Dict[str, Any]],
    *,
    tolerance: float = 1.0,
) -> Dict[str, Any]:
    """Compare the auditor's current per-block opening WDV (summed by rate)
    against the prior-year 3CD's per-rate closing WDV.

    Returns:
      {
        "ok": bool,                # True iff every rate matches within ±₹1
        "tolerance": float,
        "rows": [{
          rate, opening_excel (sum of fa_block_opening at rate),
          opening_3cd (3CD prior closing WDV at rate),
          diff (excel − 3cd),
          status: "match" | "mismatch" | "missing_in_3cd" | "missing_in_excel",
          blocks: [block_label, ...]    # which sub-blocks contributed
        }],
        "totals": {opening_excel, opening_3cd, diff},
      }
    """
    by_rate_excel: Dict[float, Dict[str, Any]] = defaultdict(lambda: {"sum": 0.0, "blocks": []})
    for r in current_openings:
        rate = round(float(r.get("rate") or 0), 4)
        if rate <= 0:
            continue
        v = float(r.get("opening_wdv") or 0)
        by_rate_excel[rate]["sum"] += v
        if v != 0:
            by_rate_excel[rate]["blocks"].append(r["block_label"])

    by_rate_3cd: Dict[float, float] = {}
    for r in cd_rate_rows:
        rate = round(float(r.get("rate") or 0), 4)
        if rate <= 0:
            continue
        # Prior-year CLOSING WDV becomes current OPENING WDV for validation
        by_rate_3cd[rate] = float(r.get("closing_wdv") or 0)

    rates = sorted(set(by_rate_excel.keys()) | set(by_rate_3cd.keys()), reverse=True)
    rows: List[Dict[str, Any]] = []
    tot_e = 0.0
    tot_c = 0.0
    ok = True
    for rate in rates:
        e_entry = by_rate_excel.get(rate, {"sum": 0.0, "blocks": []})
        excel_sum = round(e_entry["sum"], 2)
        cd_val = round(by_rate_3cd.get(rate, 0.0), 2)
        diff = round(excel_sum - cd_val, 2)
        if rate not in by_rate_3cd:
            # Both sides zero ⇒ silent match (a block with no opening on
            # either side is not a true mismatch — the auditor doesn't need
            # a noisy row for it).
            if excel_sum == 0:
                continue
            status = "missing_in_3cd"
        elif rate not in by_rate_excel or excel_sum == 0:
            if cd_val == 0:
                continue
            status = "missing_in_excel"
        elif abs(diff) <= tolerance:
            status = "match"
        else:
            status = "mismatch"
        if status not in ("match",):
            ok = False
        rows.append({
            "rate":           rate,
            "opening_excel":  excel_sum,
            "opening_3cd":    cd_val,
            "diff":           diff,
            "status":         status,
            "blocks":         e_entry["blocks"],
        })
        tot_e += excel_sum
        tot_c += cd_val
    return {
        "ok":        ok,
        "tolerance": tolerance,
        "rows":      rows,
        "totals": {
            "opening_excel": round(tot_e, 2),
            "opening_3cd":   round(tot_c, 2),
            "diff":          round(tot_e - tot_c, 2),
        },
    }


__all__ = ["build_workbook", "parse_workbook", "validate_against_3cd"]
