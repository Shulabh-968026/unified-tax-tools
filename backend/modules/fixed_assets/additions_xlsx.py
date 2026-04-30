"""Per-block Excel round-trip for the Additions Register.

Workflow:
    GET  /runs/{rid}/additions/export.xlsx
        → one sheet per active block_label, each with a totals strip,
          a Ledger column for traceability, and hidden id columns so
          `parent_addition_id` linkage survives the round-trip.

    POST /runs/{rid}/additions/import.xlsx?dry_run=true
        → parse, diff against the current DB, compare totals (drift check),
          return a preview JSON; nothing is written.

    POST /runs/{rid}/additions/import.xlsx?dry_run=false
        → apply the diff, persist a drift_warning on fa_runs when totals
          mismatch (auditor must explicitly clear it before computing).

Editable cells on import (everything else is read-only join data):
    description, party_name, voucher_no, invoice_no,
    invoice_date, put_to_use_date,
    other_expenses, itc_reversed, interest_capitalized, forex_fluctuations,
    discount_credits.
`invoice_cost` is intentionally locked end-to-end — it is the books figure
and must never drift from Tally without a re-ingest.
"""
from __future__ import annotations
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- Cell layout shared between export + import ------------------------
EDITABLE_NUM = ("other_expenses", "itc_reversed", "interest_capitalized",
                "forex_fluctuations", "discount_credits")
EDITABLE_TXT = ("description", "party_name", "voucher_no", "invoice_no")
EDITABLE_DATE = ("invoice_date", "put_to_use_date")

# Column ORDER drives both export and import. (key, header, width, kind)
COLUMNS: List[Tuple[str, str, int, str]] = [
    ("addition_id",          "addition_id",         36, "id"),
    ("parent_addition_id",   "parent_addition_id",  36, "id"),
    ("ledger_name",          "Ledger",              28, "ro"),
    ("accounting_date",      "Acc Date",            12, "ro_date"),
    ("put_to_use_date",      "PTU Date",            12, "edit_date"),
    ("description",          "Description",         42, "edit_txt"),
    ("invoice_cost",         "Invoice Cost",        14, "ro_num"),
    ("other_expenses",       "Other Exp",           12, "edit_num"),
    ("itc_reversed",         "ITC Reversed",        12, "edit_num"),
    ("interest_capitalized", "Interest Cap",        12, "edit_num"),
    ("forex_fluctuations",   "Forex",               10, "edit_num"),
    ("discount_credits",     "Discounts",           12, "edit_num"),
    ("total_capitalised",    "Total Capitalised",   16, "ro_num"),
    ("party_name",           "Supplier",            22, "edit_txt"),
    ("voucher_no",           "Voucher No",          14, "edit_txt"),
    ("invoice_no",           "Invoice No",          14, "edit_txt"),
    ("invoice_date",         "Inv Date",            12, "edit_date"),
    ("source",               "Source",              10, "ro"),
]
COL_KEYS    = [c[0] for c in COLUMNS]
HEADER_ROW  = 4   # row 1=title, 2=totals labels, 3=totals values, 4=headers


def _round2(v: Any) -> float:
    try:
        return round(float(v or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _capitalised(a: Dict[str, Any]) -> float:
    return _round2(
        _round2(a.get("invoice_cost"))
        + _round2(a.get("other_expenses"))
        - _round2(a.get("itc_reversed"))
        + _round2(a.get("interest_capitalized"))
        + _round2(a.get("forex_fluctuations"))
        - _round2(a.get("discount_credits"))
    )


def _block_totals(rows: List[Dict[str, Any]]) -> Dict[str, float]:
    out = {k: 0.0 for k in (
        "invoice_cost", "other_expenses", "itc_reversed", "interest_capitalized",
        "forex_fluctuations", "discount_credits", "total_capitalised")}
    for r in rows:
        # Merged child rows: their invoice_cost has already been rolled into
        # the parent's adjustment columns at link time, so they no longer
        # contribute to the block total.
        if r.get("parent_addition_id"):
            continue
        for k in ("invoice_cost", "other_expenses", "itc_reversed",
                  "interest_capitalized", "forex_fluctuations", "discount_credits"):
            out[k] += _round2(r.get(k))
        out["total_capitalised"] += _capitalised(r)
    return {k: _round2(v) for k, v in out.items()}


# ---- Export ------------------------------------------------------------
def build_additions_workbook(
    *, client_name: str, fy: str, rows_by_block: Dict[str, List[Dict[str, Any]]],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    title_font   = Font(bold=True, size=12, color="FFFFFF")
    title_fill   = PatternFill("solid", fgColor="1F2937")
    head_font    = Font(bold=True, size=10, color="FFFFFF")
    head_fill    = PatternFill("solid", fgColor="334155")
    total_lbl    = Font(bold=True, size=9, color="475569")
    total_val    = Font(bold=True, size=10, color="0F172A")
    locked_fill  = PatternFill("solid", fgColor="F1F5F9")  # read-only cells
    edit_fill    = PatternFill("solid", fgColor="FEFCE8")  # editable cells (subtle yellow)
    discount_fill = PatternFill("solid", fgColor="FEE2E2")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if not rows_by_block:
        rows_by_block = {"(no data)": []}

    for block_label, rows in rows_by_block.items():
        # Sheet titles in xlsx have a 31-char limit and forbid `:[]/\?*`
        safe = "".join(ch if ch not in ':[]/\\?*' else "-" for ch in block_label)[:31] or "Block"
        ws = wb.create_sheet(title=safe)

        # Title row
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(COLUMNS))
        ws.cell(row=1, column=1,
                value=f"{client_name} · FY {fy} · {block_label}").font = title_font
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 22

        totals = _block_totals(rows)
        # Totals strip (rows 2 + 3) — labels then frozen totals values
        for ci, (key, hdr, _w, _k) in enumerate(COLUMNS, start=1):
            if key in totals:
                ws.cell(row=2, column=ci, value=hdr.upper()).font = total_lbl
                ws.cell(row=3, column=ci, value=totals[key]).font = total_val
                ws.cell(row=3, column=ci).number_format = "#,##0.00;(#,##0.00)"
        ws.cell(row=2, column=1, value="TOTALS →").font = total_lbl

        # Header row
        for ci, (key, hdr, w, _k) in enumerate(COLUMNS, start=1):
            c = ws.cell(row=HEADER_ROW, column=ci, value=hdr)
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Hide the two id columns (still readable but tucked away)
        ws.column_dimensions["A"].hidden = True
        ws.column_dimensions["B"].hidden = True

        # Data rows
        for ri, r in enumerate(rows, start=HEADER_ROW + 1):
            is_disc = r.get("source") == "discount_credit"
            is_merged = bool(r.get("parent_addition_id"))
            for ci, (key, _hdr, _w, kind) in enumerate(COLUMNS, start=1):
                if key == "total_capitalised":
                    val = _capitalised(r)
                else:
                    val = r.get(key)
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = border
                if kind in ("edit_num", "ro_num"):
                    cell.number_format = "#,##0.00;(#,##0.00)"
                    cell.alignment = Alignment(horizontal="right")
                elif kind in ("edit_date", "ro_date"):
                    cell.alignment = Alignment(horizontal="center")
                if kind in ("ro", "ro_num", "ro_date", "id") or is_disc or is_merged:
                    cell.fill = discount_fill if is_disc else locked_fill
                else:
                    cell.fill = edit_fill

        # Freeze the totals + header rows so the auditor always sees them
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=4)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---- Import ------------------------------------------------------------
def _coerce_num(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return _round2(v)
    except (TypeError, ValueError):
        return 0.0


def _coerce_date(v: Any) -> str:
    """Returns YYYY-MM-DD or empty string."""
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if hasattr(v, "isoformat") and not isinstance(v, str):
        try:
            return v.isoformat()[:10]
        except Exception:  # noqa: BLE001
            return ""
    s = str(v).strip()
    if not s:
        return ""
    # Tolerate dd-mm-yyyy / dd/mm/yyyy / yyyy-mm-dd
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return s[:10] if len(s) >= 10 else ""


def parse_additions_workbook(content: bytes) -> Dict[str, Any]:
    """Parse the uploaded XLSX and emit `{rows: [{addition_id, fields...}],
    sheets: [{name, totals}], errors: []}`. The caller diffs the rows
    against the live DB."""
    wb = load_workbook(BytesIO(content), data_only=True)
    out_rows: List[Dict[str, Any]] = []
    sheets: List[Dict[str, Any]] = []
    errors: List[str] = []

    key_to_idx = {k: i for i, k in enumerate(COL_KEYS)}

    for ws in wb.worksheets:
        # Header row position is fixed in our exporter — rely on it. If the
        # auditor inserted/deleted columns we'll detect it via a header check.
        header_cells = [ws.cell(row=HEADER_ROW, column=ci + 1).value
                        for ci in range(len(COLUMNS))]
        expected = [h for _, h, _, _ in COLUMNS]
        if header_cells[:3] != expected[:3]:
            errors.append(
                f"Sheet '{ws.title}': header row at row {HEADER_ROW} doesn't "
                f"match the original export — skipped (insert/delete columns "
                f"isn't supported, edit values only).")
            continue

        sheet_totals = {k: 0.0 for k in (
            "invoice_cost", "other_expenses", "itc_reversed", "interest_capitalized",
            "forex_fluctuations", "discount_credits", "total_capitalised")}

        ri = HEADER_ROW + 1
        rows_in_sheet = 0
        while True:
            aid = ws.cell(row=ri, column=1).value  # addition_id
            if aid is None:
                # Stop on first fully blank row
                if all(ws.cell(row=ri, column=ci + 1).value in (None, "")
                       for ci in range(len(COLUMNS))):
                    break
                ri += 1
                continue
            row = {"addition_id": str(aid).strip()}
            for ci, (key, _hdr, _w, kind) in enumerate(COLUMNS, start=1):
                v = ws.cell(row=ri, column=ci).value
                if key == "addition_id":
                    continue
                if key == "parent_addition_id":
                    row[key] = (str(v).strip() if v not in (None, "") else "")
                elif kind in ("edit_num", "ro_num"):
                    row[key] = _coerce_num(v)
                elif kind in ("edit_date", "ro_date"):
                    row[key] = _coerce_date(v)
                else:
                    row[key] = ("" if v is None else str(v).strip())

            if (row.get("source") or "") != "discount_credit" \
                    and not row.get("parent_addition_id"):
                for k in ("invoice_cost", "other_expenses", "itc_reversed",
                          "interest_capitalized", "forex_fluctuations",
                          "discount_credits"):
                    sheet_totals[k] += _round2(row.get(k))
                sheet_totals["total_capitalised"] += _round2(row.get("total_capitalised"))

            out_rows.append(row)
            rows_in_sheet += 1
            ri += 1
            if ri > 50000:  # belt-and-braces stop
                errors.append(f"Sheet '{ws.title}': stopped reading at 50000 rows.")
                break

        sheets.append({
            "name":   ws.title,
            "rows":   rows_in_sheet,
            "totals": {k: _round2(v) for k, v in sheet_totals.items()},
        })

    _ = key_to_idx  # placeholder for future column-shape checks
    return {"rows": out_rows, "sheets": sheets, "errors": errors}


def diff_additions(*, db_rows: List[Dict[str, Any]], xl_rows: List[Dict[str, Any]],
                   tolerance: float = 0.01) -> Dict[str, Any]:
    """Build a per-row diff between Excel and DB. Only the editable fields
    are diffed — everything else is read-only and silently ignored."""
    db_by_id = {r["addition_id"]: r for r in db_rows}
    changes: List[Dict[str, Any]] = []
    unknown: List[str] = []
    fields = list(EDITABLE_NUM) + list(EDITABLE_TXT) + list(EDITABLE_DATE)

    for x in xl_rows:
        aid = x.get("addition_id") or ""
        if not aid or aid.startswith("discount-"):
            continue
        if aid not in db_by_id:
            unknown.append(aid)
            continue
        d = db_by_id[aid]
        if (d.get("source") or "") == "discount_credit":
            continue  # locked end-to-end
        if d.get("parent_addition_id"):
            continue  # merged children — invoice_cost lives in the parent
        per_row: Dict[str, Dict[str, Any]] = {}
        for f in fields:
            old = d.get(f)
            new = x.get(f)
            if f in EDITABLE_NUM:
                old_n = _round2(old)
                new_n = _round2(new)
                if abs(old_n - new_n) > tolerance:
                    per_row[f] = {"old": old_n, "new": new_n}
            else:
                old_s = (old or "").strip() if isinstance(old, str) else (old or "")
                new_s = (new or "").strip() if isinstance(new, str) else (new or "")
                if old_s != new_s:
                    per_row[f] = {"old": old_s, "new": new_s}
        if per_row:
            changes.append({
                "addition_id": aid,
                "ledger_name": d.get("ledger_name", ""),
                "description": d.get("description") or d.get("particulars", ""),
                "block_label": d.get("block_label", ""),
                "changes":     per_row,
            })

    return {"changes": changes, "unknown_ids": unknown}


def block_drift(*, db_rows: List[Dict[str, Any]], xl_changes: List[Dict[str, Any]],
                tolerance: float = 1.0) -> Dict[str, Any]:
    """Apply the proposed Excel changes virtually, recompute block totals,
    and compare to the current DB block totals. Anything beyond `tolerance`
    is flagged as drift (and a persistent banner is shown until cleared)."""
    db_by_id = {r["addition_id"]: dict(r) for r in db_rows}
    for ch in xl_changes:
        row = db_by_id.get(ch["addition_id"])
        if not row:
            continue
        for f, v in ch["changes"].items():
            row[f] = v["new"]

    by_block_db: Dict[str, List[Dict[str, Any]]] = {}
    by_block_xl: Dict[str, List[Dict[str, Any]]] = {}
    for r in db_rows:
        by_block_db.setdefault(r.get("block_label") or "(unblocked)", []).append(r)
    for r in db_by_id.values():
        by_block_xl.setdefault(r.get("block_label") or "(unblocked)", []).append(r)

    blocks_out = []
    drifted = False
    for bl in sorted(set(by_block_db) | set(by_block_xl)):
        td = _block_totals(by_block_db.get(bl, []))
        tx = _block_totals(by_block_xl.get(bl, []))
        diff = _round2(tx["total_capitalised"] - td["total_capitalised"])
        if abs(diff) > tolerance:
            drifted = True
        blocks_out.append({
            "block_label":      bl,
            "db_total":         td["total_capitalised"],
            "excel_total":      tx["total_capitalised"],
            "diff":             diff,
        })
    return {"drifted": drifted, "blocks": blocks_out}
