"""Excel export — IT Depreciation Schedule.

Single workbook, 4 sheets:
  1. Block Summary  — mirrors the user's Sample IT Depreciation Schedule
  2. Additions Register
  3. Deletions Register
  4. Workings — formula notes for audit trail
"""
from __future__ import annotations
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_THIN = Side(style="thin", color="888888")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HDR_FILL = PatternFill("solid", fgColor="F1F5F9")
TOT_FILL = PatternFill("solid", fgColor="E2E8F0")
TITLE_FONT = Font(bold=True, size=14)
HDR_FONT = Font(bold=True, size=11)
NUM_FMT = "#,##,##0.00;(#,##,##0.00);-"


def _set(cell, value, *, bold=False, fill=None, num=False, align=None):
    cell.value = value
    if bold:
        cell.font = HDR_FONT
    if fill:
        cell.fill = fill
    if num:
        cell.number_format = NUM_FMT
    if align:
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER


def _autosize(ws, widths: Dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _format_inr_indian(v: float) -> str:
    """Format a number in Indian grouping for *measurement only* — matches
    what Excel will display under the NUM_FMT format string."""
    n = float(v or 0)
    if n == 0:
        return "-"
    s = f"{abs(n):,.2f}"
    intpart, _, dec = s.partition(".")
    intpart = intpart.replace(",", "")
    if len(intpart) > 3:
        last3 = intpart[-3:]
        rest = intpart[:-3]
        groups = [rest[max(0, i - 2):i] for i in range(len(rest), 0, -2)][::-1]
        intpart = ",".join(groups) + "," + last3
    formatted = f"{intpart}.{dec}"
    return f"({formatted})" if n < 0 else formatted


def _fit_column_widths(ws, *, header_row: int, last_row: int,
                       num_cols: int, num_col_indexes: set,
                       text_cap: int = 50, num_cap: int = 22, padding: float = 1.4):
    """Measure each column's widest content (header + every data cell)
    and override the explicit column widths so numbers never wrap and
    text columns shrink to fit. `num_col_indexes` contains 1-based indices
    of columns formatted as numbers — those are measured via the Indian-
    grouping string representation. Other columns measure raw `str(value)`.

    `text_cap` limits a runaway text column (e.g. 200-char particulars)
    so the workbook stays printable; `num_cap` keeps numeric columns
    sane even if a single cell happens to be very long."""
    widths: Dict[int, float] = {}
    for col in range(1, num_cols + 1):
        max_len = 0
        for r in range(header_row, last_row + 1):
            cell = ws.cell(r, col)
            v = cell.value
            if v is None or v == "":
                continue
            if col in num_col_indexes and isinstance(v, (int, float)):
                txt = _format_inr_indian(float(v))
            else:
                txt = str(v)
            # Use the longest line if the cell wraps via \n
            for line in txt.split("\n"):
                if len(line) > max_len:
                    max_len = len(line)
        cap = num_cap if col in num_col_indexes else text_cap
        # +padding for cell inset (Excel uses ~1 char visual padding)
        widths[col] = min(max(8.0, max_len + padding), cap)
    _autosize(ws, widths)


def write_block_summary(ws, *, client_name: str, fy_start: str, fy_end: str,
                        rows: List[Dict[str, Any]], totals: Dict[str, float]):
    ws.title = "Block Summary"
    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(1, 1).value = client_name or ""
    ws.cell(1, 1).font = TITLE_FONT
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.cell(2, 1).value = f"IT Depreciation Schedule for the period {fy_start} to {fy_end}"
    ws.cell(2, 1).font = Font(bold=True, size=11)
    ws.cell(2, 1).alignment = Alignment(horizontal="center")

    # Header row (row 4)
    headers = [
        "PARTICULARS", "Rate", "Opening WDV", "Adds ≥ 180 days",
        "Adds < 180 days", "Deletions (Sales)", "Total before Depn",
        "Depreciation", "STCG u/s 50", "Closing WDV",
    ]
    for i, h in enumerate(headers, start=1):
        _set(ws.cell(4, i), h, bold=True, fill=HDR_FILL, align="center")

    # Data rows
    r = 5
    for row in rows:
        _set(ws.cell(r, 1), row["block_label"], align="left")
        _set(ws.cell(r, 2), f"{row['rate']:.0f}%" if row.get("rate") else "", align="center")
        _set(ws.cell(r, 3), row["opening_wdv"], num=True, align="right")
        _set(ws.cell(r, 4), row["adds_full"], num=True, align="right")
        _set(ws.cell(r, 5), row["adds_half"], num=True, align="right")
        _set(ws.cell(r, 6), row["deletions"], num=True, align="right")
        _set(ws.cell(r, 7), row["total_block"], num=True, align="right")
        _set(ws.cell(r, 8), row["depreciation"], num=True, align="right")
        _set(ws.cell(r, 9), row.get("stcg_sec50", 0), num=True, align="right")
        _set(ws.cell(r, 10), row["closing_wdv"], num=True, align="right")
        r += 1

    # Totals
    _set(ws.cell(r, 1), "TOTAL", bold=True, fill=TOT_FILL, align="left")
    _set(ws.cell(r, 2), "", fill=TOT_FILL)
    _set(ws.cell(r, 3), totals["opening_wdv"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 4), totals["adds_full"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 5), totals["adds_half"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 6), totals["deletions"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 7), totals["total_block"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 8), totals["depreciation"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 9), totals["stcg_sec50"], bold=True, fill=TOT_FILL, num=True, align="right")
    _set(ws.cell(r, 10), totals["closing_wdv"], bold=True, fill=TOT_FILL, num=True, align="right")

    # Auto-fit column widths to widest content (header + body + total).
    _fit_column_widths(
        ws, header_row=4, last_row=r,
        num_cols=10, num_col_indexes={3, 4, 5, 6, 7, 8, 9, 10},
        # PARTICULARS has long block names — give it more room
        text_cap=42,
    )


def write_additions(ws, additions: List[Dict[str, Any]]):
    ws.title = "Additions Register"
    headers = [
        "Block", "Voucher No", "Voucher Type", "Acc Date", "Inv Date",
        "PTU Date", "Half Rate?", "Party", "Particulars",
        "Invoice Cost", "Discount/Credits", "Other Expenses",
        "ITC Reversed", "Interest Cap", "Forex",
        "Capitalised Cost", "Ledger",
    ]
    for i, h in enumerate(headers, start=1):
        _set(ws.cell(1, i), h, bold=True, fill=HDR_FILL, align="center")

    from modules.fixed_assets.compute import adjusted_cost
    r = 2
    for a in additions:
        cap = adjusted_cost(a)
        _set(ws.cell(r, 1),  a.get("block_label", ""))
        _set(ws.cell(r, 2),  a.get("voucher_no", ""))
        _set(ws.cell(r, 3),  a.get("voucher_type", ""))
        _set(ws.cell(r, 4),  a.get("accounting_date", ""))
        _set(ws.cell(r, 5),  a.get("invoice_date", ""))
        _set(ws.cell(r, 6),  a.get("put_to_use_date", ""))
        _set(ws.cell(r, 7),  "Yes" if not a.get("is_more_than_180", True) else "")
        _set(ws.cell(r, 8),  a.get("party_name", ""))
        _set(ws.cell(r, 9),  a.get("particulars", "")[:200])
        _set(ws.cell(r, 10), float(a.get("invoice_cost") or 0), num=True, align="right")
        _set(ws.cell(r, 11), float(a.get("discount_credits") or 0), num=True, align="right")
        _set(ws.cell(r, 12), float(a.get("other_expenses") or 0), num=True, align="right")
        _set(ws.cell(r, 13), float(a.get("itc_reversed") or 0), num=True, align="right")
        _set(ws.cell(r, 14), float(a.get("interest_capitalized") or 0), num=True, align="right")
        _set(ws.cell(r, 15), float(a.get("forex_fluctuations") or 0), num=True, align="right")
        _set(ws.cell(r, 16), cap, bold=True, num=True, align="right")
        _set(ws.cell(r, 17), a.get("ledger_name", ""))
        r += 1
    _fit_column_widths(
        ws, header_row=1, last_row=max(r - 1, 1),
        num_cols=17, num_col_indexes={10, 11, 12, 13, 14, 15, 16},
        # Particulars is the runaway column; cap a bit higher than default
        text_cap=50,
    )


def write_deletions(ws, deletions: List[Dict[str, Any]]):
    ws.title = "Deletions Register"
    headers = [
        "Block", "Voucher No", "Sale Date", "Buyer",
        "Sale Value", "Particulars", "Ledger",
    ]
    for i, h in enumerate(headers, start=1):
        _set(ws.cell(1, i), h, bold=True, fill=HDR_FILL, align="center")
    r = 2
    for d in [d for d in deletions if (d.get("classification") or "") == "sale"]:
        _set(ws.cell(r, 1), d.get("block_label", ""))
        _set(ws.cell(r, 2), d.get("voucher_no", ""))
        _set(ws.cell(r, 3), d.get("sale_date", ""))
        _set(ws.cell(r, 4), d.get("buyer_name", ""))
        _set(ws.cell(r, 5), float(d.get("sale_value") or 0), num=True, align="right")
        _set(ws.cell(r, 6), d.get("particulars", "")[:200])
        _set(ws.cell(r, 7), d.get("ledger_name", ""))
        r += 1
    _fit_column_widths(
        ws, header_row=1, last_row=max(r - 1, 1),
        num_cols=7, num_col_indexes={5},
        text_cap=50,
    )


def write_workings(ws):
    ws.title = "Workings"
    ws.column_dimensions["A"].width = 110
    notes = [
        "Computation method — Section 32, Income-tax Act, 1961 — WDV (Block of Assets):",
        "",
        "1. Capitalised Cost per addition  =  Invoice Cost  −  Discount / Credits  +  Other Expenses",
        "                                       −  ITC Reversed  +  Interest Capitalised  +  Forex Fluctuations.",
        "",
        "2. Block WDV before depreciation  =  Opening WDV  +  Total Additions (full + half)  −  Sales (deletions).",
        "",
        "3. 180-day rule:",
        "      Adds ≥ 180 days (PTU on/before Oct 3 for FY ending 31 March)  ⇒  full statutory rate.",
        "      Adds < 180 days  ⇒  half the statutory rate (rate ÷ 2).",
        "",
        "4. Sale allocation: Sales first reduce the full-rate pool (Opening + Adds≥180);",
        "   any excess reduces the half-rate pool (Adds<180).",
        "",
        "5. Depreciation =  (Eligible_at_full_rate × rate)  +  (Eligible_at_half_rate × rate ÷ 2).",
        "",
        "6. Sec 50 STCG: When Block WDV before depreciation < 0, depreciation = 0;",
        "   block extinguishes; the negative amount is reported as Short-Term Capital Gain.",
        "",
        "7. Closing WDV = Block WDV before depreciation − Depreciation.",
    ]
    for i, line in enumerate(notes, start=1):
        ws.cell(i, 1).value = line
        if i == 1:
            ws.cell(i, 1).font = HDR_FONT


def build_workbook(*, client_name: str, fy_start: str, fy_end: str,
                   rows: List[Dict[str, Any]], totals: Dict[str, float],
                   additions: List[Dict[str, Any]],
                   deletions: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    write_block_summary(wb.active, client_name=client_name, fy_start=fy_start,
                        fy_end=fy_end, rows=rows, totals=totals)
    write_additions(wb.create_sheet(), additions)
    write_deletions(wb.create_sheet(), deletions)
    write_workings(wb.create_sheet())
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
