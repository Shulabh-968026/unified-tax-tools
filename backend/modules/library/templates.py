"""Template generators for Library file-types.

Some file-types — notably Party Master — can be **pre-populated** from
data we already hold in the Library.  Auditor downloads the template,
fills the gaps offline (emails, MSME status, etc.), and re-uploads.

Each generator is a callable:
    generator(*, firm_id, client_id, period, division) -> (xlsx_bytes, filename)

Generators are registered by file_type below.  A file_type with no
registered generator returns 404 from the controller — the UI hides
the "Download Template" button accordingly via the catalog's
`has_template` flag.
"""
from __future__ import annotations

import io
import json
from typing import Awaitable, Callable, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from modules.library import service as lib_svc
from modules.library.controller import DEFAULT_FIRM_ID  # for callers


# ---------------------------------------------------------------------------
# Styling helpers — keep templates visually consistent.
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
PREFILLED_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")  # pale emerald
TODO_FILL = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")        # pale amber
SECTION_FILL = PatternFill(start_color="F3F4F1", end_color="F3F4F1", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=10, bold=True, color="52524E")


def _autosize(ws, headers: list[str], rows: list[list]):
    for ci, h in enumerate(headers, start=1):
        col = get_column_letter(ci)
        max_len = max([len(str(h or ""))] + [len(str(r[ci - 1]) if ci - 1 < len(r) else "") for r in rows])
        ws.column_dimensions[col].width = min(max(max_len + 2, 12), 48)


def _write_header(ws, headers: list[str], note_row: list[str] | None = None):
    """Write the first row as styled header; optional second row of notes."""
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    if note_row:
        for ci, n in enumerate(note_row, start=1):
            c = ws.cell(row=2, column=ci, value=n)
            c.fill = SECTION_FILL
            c.font = SECTION_FONT
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 24


def _write_rows(ws, rows: list[list], prefilled_cols: tuple, start_row: int = 3):
    """Write data rows starting at `start_row`, painting prefilled vs todo cells."""
    for ri, row in enumerate(rows, start=start_row):
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = PREFILLED_FILL if ci in prefilled_cols else TODO_FILL
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    if not rows:
        ws.cell(row=start_row, column=1, value="(no parties in this bucket)").font = Font(italic=True, color="8A8A83")


def _add_dropdown(ws, options: list[str], col_letter: str, last_row: int, *, allow_blank=True):
    """Attach a dropdown to `<col_letter>3:<col_letter>{last_row}`."""
    if last_row < 3:
        last_row = 3
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}3:{col_letter}{last_row}")


# ---------------------------------------------------------------------------
# Bucket detection — categorises Tally ledgers into the four
# confirmation universes plus Others.  Order matters — first match wins.
# ---------------------------------------------------------------------------
BUCKETS = [
    {
        "name": "Trade Payables",
        "key":  "payables",
        "hints": ["sundry creditor", "trade payable", "creditors", "creditor"],
    },
    {
        "name": "Trade Receivables",
        "key":  "receivables",
        "hints": ["sundry debtor", "trade receivable", "debtors", "debtor"],
    },
    {
        "name": "Unsecured Loans",
        "key":  "loans",
        "hints": ["unsecured loan", "secured loan", "loans (liabilit", "loans & advance", "loans and advance", "loan a/c"],
    },
    {
        "name": "Bank Accounts",
        "key":  "bank",
        "hints": ["bank account", "bank od", "bank o/d", "bank balance", "cash & cash equivalents", "bank a/c", "bank ac"],
    },
    {
        "name": "Others",
        "key":  "others",
        "hints": [],  # catch-all
    },
]


def _bucket_key_for(rec: dict, ledger_name: str) -> str:
    """Decide which bucket a ledger belongs to.

    Inspects (in order): subhead, groupParent, head, ledger_name itself.
    Bank-named ledgers without a clear subhead still surface in the
    Bank bucket via the name match.
    """
    s = " ".join([
        rec.get("subhead") or "",
        rec.get("groupParent") or "",
        rec.get("head") or "",
        ledger_name or "",
    ]).lower()
    for b in BUCKETS:
        if not b["hints"]:
            continue
        if any(h in s for h in b["hints"]):
            return b["key"]
    return "others"


# ---------------------------------------------------------------------------
# Column sets — different sheet, different schema.  Keys must match the
# slot count in `_row_for_bucket`.
# ---------------------------------------------------------------------------
COMMON_COLS = ["Party Name", "Subhead / Group", "Closing Balance (Cr / Dr)"]
ADDRESS_COLS = ["Address Line 1", "Address Line 2", "City", "Pincode"]

COLUMN_SETS = {
    "payables": {
        "headers": COMMON_COLS + [
            "GSTIN", "GST Registration Type", "Country",
            "Email ID", "Alternate Email", "Phone",
        ] + ADDRESS_COLS + [
            "PAN", "Notes",
        ],
        "notes": [
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled where Tally has it",
            "regular / composition / consumer / unregistered / overseas",
            "Pre-filled if non-India",
            "Required for Balance Confirmation",
            "Optional · CC on confirmation emails",
            "Optional",
            "Optional · pre-filled where Tally has it",
            "Optional",
            "Optional",
            "6-digit postal code",
            "Optional",
            "Free text",
        ],
        "prefilled_cols": (1, 2, 3, 4, 5, 6),
    },
    "receivables": {
        "headers": COMMON_COLS + [
            "GSTIN", "GST Registration Type", "Country",
            "Email ID", "Alternate Email", "Phone",
        ] + ADDRESS_COLS + [
            "PAN", "Notes",
        ],
        "notes": [
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled where Tally has it",
            "regular / composition / consumer / unregistered / overseas",
            "Pre-filled if non-India",
            "Required for Balance Confirmation",
            "Optional · CC on confirmation emails",
            "Optional",
            "Optional · pre-filled where Tally has it",
            "Optional",
            "Optional",
            "6-digit postal code",
            "Optional",
            "Free text",
        ],
        "prefilled_cols": (1, 2, 3, 4, 5, 6),
    },
    "loans": {
        "headers": COMMON_COLS + [
            "GSTIN", "Country",
            "Email ID", "Phone",
        ] + ADDRESS_COLS + [
            "PAN", "Loan Agreement Date", "Interest Rate (% p.a.)", "Notes",
        ],
        "notes": [
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled where Tally has it",
            "Pre-filled if non-India",
            "Required for Balance Confirmation",
            "Optional",
            "Optional",
            "Optional",
            "Optional",
            "6-digit postal code",
            "Optional",
            "DD-MM-YYYY",
            "Optional",
            "Free text",
        ],
        "prefilled_cols": (1, 2, 3, 4, 5),
    },
    "bank": {
        "headers": COMMON_COLS + [
            "Bank Name", "Branch", "Account Number", "IFSC",
            "Email ID", "Notes",
        ],
        "notes": [
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Required",
            "Optional",
            "Required for Bank Confirmation",
            "Optional",
            "Required for Bank Confirmation",
            "Free text",
        ],
        "prefilled_cols": (1, 2, 3),
    },
    "others": {
        "headers": COMMON_COLS + ["Notes"],
        "notes": [
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Pre-filled · do not edit",
            "Free text",
        ],
        "prefilled_cols": (1, 2, 3),
    },
}


# ---------------------------------------------------------------------------
# Per-bucket row builder.
# ---------------------------------------------------------------------------
def _split_party_address(json_party: dict) -> tuple:
    """Extract (line1, line2, city, pincode) from a Tally JSON party object."""
    def _s(v):
        return v.strip() if isinstance(v, str) else ""
    line1 = _s(json_party.get("addressLine1"))
    line2 = _s(json_party.get("addressLine2"))
    line3 = _s(json_party.get("addressLine3"))
    city = _s(json_party.get("city"))
    pincode = _s(json_party.get("pinCode") or json_party.get("pincode"))
    if line3:
        line2 = (line2 + ", " + line3).strip(", ") if line2 else line3
    # Legacy shape — sometimes addresses are nested in `address` list.
    if not (line1 or line2 or city or pincode):
        addr = json_party.get("address")
        parts: list = []
        if isinstance(addr, list):
            parts = [str(a).strip() for a in addr if a]
        elif isinstance(addr, str):
            parts = [p.strip() for p in addr.split(",") if p.strip()]
        if parts:
            if parts[-1].replace(" ", "").isdigit() and len(parts[-1].replace(" ", "")) == 6:
                pincode = parts.pop()
            if parts:
                city = parts.pop()
            line1 = parts[0] if parts else ""
            line2 = ", ".join(parts[1:]) if len(parts) > 1 else ""
    return line1, line2, city, pincode


def _row_for_bucket(bucket_key: str, name: str, rec: dict, json_party: dict | None) -> list:
    closing = rec.get("closingBalance")
    closing_str = ""
    if closing is not None:
        sign = "Cr" if closing < 0 else "Dr"
        closing_str = f"{abs(closing):,.2f} {sign}"
    json_party = json_party or {}
    country = (json_party.get("country") or "").strip()
    country_str = "" if country.lower() in ("india", "") else country
    subhead = rec.get("subhead") or rec.get("groupParent") or ""
    line1, line2, city, pincode = _split_party_address(json_party)

    if bucket_key in ("payables", "receivables"):
        return [
            name, subhead, closing_str,
            json_party.get("partyGSTIN") or "",
            (json_party.get("gstRegistrationType") or "").lower(),
            country_str,
            "",       # email
            "",       # alt email
            "",       # phone
            line1,    # Address Line 1
            line2,    # Address Line 2
            city,     # City
            pincode,  # Pincode
            "",       # PAN
            "",       # notes
        ]
    if bucket_key == "loans":
        return [
            name, subhead, closing_str,
            json_party.get("partyGSTIN") or "",
            country_str,
            "",       # email
            "",       # phone
            line1,    # Address Line 1
            line2,    # Address Line 2
            city,     # City
            pincode,  # Pincode
            "",       # PAN
            "",       # loan date
            "",       # interest rate
            "",       # notes
        ]
    if bucket_key == "bank":
        return [
            name, subhead, closing_str,
            "",  # bank name
            "",  # branch
            "",  # acct no
            "",  # IFSC
            "",  # email
            "",  # notes
        ]
    # others
    return [name, subhead, closing_str, ""]


# ---------------------------------------------------------------------------
# MSME Details sheet — vendor-only, with the same drop-downs the
# 43B(h) module already uses (single source of truth for dropdown
# options is `modules.msme43bh.schemas`).
# ---------------------------------------------------------------------------
MSME_HEADERS = [
    "Party Name",                        # pre-filled (Trade Payables only)
    "Subhead / Group",                   # pre-filled
    "MSME Registration Number",          # auditor — Udyam / EM-II
    "Sector",                            # dropdown · Manufacturing / Services / Trading
    "MSME Type",                         # dropdown · Micro / Small / Medium
    "Capital Goods / Fund Creditor",     # dropdown · Yes / No
    "Date of MSME Registration",         # auditor (DD-MM-YYYY)
    "Notes",                             # auditor
]
MSME_HEADER_NOTES = [
    "Pre-filled · vendors only",
    "Pre-filled · do not edit",
    "From Udyam certificate",
    "Choose from dropdown",
    "Choose from dropdown · leave blank if Not MSME",
    "Choose from dropdown · drives 43B(h) carve-out",
    "DD-MM-YYYY",
    "Free text",
]


# ---------------------------------------------------------------------------
# Party Master template (multi-bucket).
# ---------------------------------------------------------------------------
async def generate_party_master_template(
    *, firm_id: str, client_id: str, period: str, division: str | None,
) -> Tuple[bytes, str]:
    """Build a Party Master template with one sheet per confirmation
    universe + a vendor-only MSME Details sheet.

    Sheets:
      • README
      • Trade Payables       — vendors (drives Confirmations + 43B(h))
      • Trade Receivables    — customers (drives Confirmations)
      • Unsecured Loans      — loan parties (drives Confirmations)
      • Bank Accounts        — bank ledgers (drives Bank Confirmations)
      • Others               — anything else with a closing balance
      • MSME Details         — Trade Payables names + the 4 MSME columns
                                with drop-downs

    Pre-fill sources:
      • Ledger Mapping XLSX (current version) — name, group, closing balance.
      • Books JSON         (current version) — GSTIN, regType, country.
    """
    # Lazy import — avoids circular dep at module-load time.
    from modules.msme43bh.schemas import SECTOR_OPTIONS, MSME_TYPE_OPTIONS

    # ── 1. Read latest Library files we depend on.
    ledger_map = await lib_svc.get_current_file(
        firm_id=firm_id, client_id=client_id, period=period,
        division=division, file_type="ledger_mapping_xlsx",
    )
    books = await lib_svc.get_current_file(
        firm_id=firm_id, client_id=client_id, period=period,
        division=division, file_type="books_json",
    )
    if not ledger_map:
        raise FileNotFoundError("Ledger Mapping XLSX must be uploaded first.")

    from modules.clause44.service import parse_ledger_xlsx
    ledger_xlsx_bytes = await lib_svc.read_file_bytes(ledger_map["file_id"])
    ledger_records = parse_ledger_xlsx(ledger_xlsx_bytes)

    parties_by_name: dict[str, dict] = {}
    if books:
        try:
            data = json.loads(await lib_svc.read_file_bytes(books["file_id"]))
            for p in data.get("parties", []) or []:
                if p.get("name"):
                    parties_by_name[p["name"].strip()] = p
            for lg in data.get("ledgers", []) or []:
                nm = (lg.get("name") or "").strip()
                if nm and nm not in parties_by_name and lg.get("partyGSTIN"):
                    parties_by_name[nm] = lg
        except Exception:
            pass

    # ── 2. Bucketize.
    buckets: dict[str, list[Tuple[str, dict]]] = {b["key"]: [] for b in BUCKETS}
    for name, rec in ledger_records.items():
        if rec.get("bsOrPl") == "P":
            continue
        bkey = _bucket_key_for(rec, name)
        # In "others", only include rows with a non-zero closing balance
        # — keeps the sheet focused on confirmation candidates.
        if bkey == "others" and not rec.get("closingBalance"):
            continue
        buckets[bkey].append((name, rec))
    for k, lst in buckets.items():
        lst.sort(key=lambda t: t[0].lower())

    # ── 3. Build workbook.
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    _build_readme(readme, period=period)

    # ── One sheet per bucket.
    for b in BUCKETS:
        rows_data = buckets[b["key"]]
        rows = [
            _row_for_bucket(b["key"], name, rec, parties_by_name.get(name))
            for name, rec in rows_data
        ]
        spec = COLUMN_SETS[b["key"]]
        ws = wb.create_sheet(title=b["name"])
        _write_header(ws, spec["headers"], note_row=spec["notes"])
        _write_rows(ws, rows, prefilled_cols=spec["prefilled_cols"])
        _autosize(ws, spec["headers"], rows)
        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = False

    # ── MSME Details sheet — Trade Payables only.
    msme_rows = []
    for name, rec in buckets["payables"]:
        subhead = rec.get("subhead") or rec.get("groupParent") or ""
        msme_rows.append([name, subhead, "", "", "", "", "", ""])

    ws_msme = wb.create_sheet(title="MSME Details")
    _write_header(ws_msme, MSME_HEADERS, note_row=MSME_HEADER_NOTES)
    _write_rows(ws_msme, msme_rows, prefilled_cols=(1, 2))
    _autosize(ws_msme, MSME_HEADERS, msme_rows)
    ws_msme.freeze_panes = "A3"
    ws_msme.sheet_view.showGridLines = False

    last_row = max(2 + len(msme_rows), 3)
    _add_dropdown(ws_msme, SECTOR_OPTIONS,    "D", last_row)  # Sector
    _add_dropdown(ws_msme, MSME_TYPE_OPTIONS, "E", last_row)  # MSME Type
    _add_dropdown(ws_msme, ["Yes", "No"],     "F", last_row)  # Capital Goods

    # ── 4. Stream out.
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f"party_master_template__{client_id}__{period}.xlsx"


def _build_readme(ws, *, period: str):
    ws["A1"] = "Party Master Template — AssureAI Audit Utilities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = f"Generated for FY {period}"
    ws["A3"].font = Font(bold=False, size=10, italic=True, color="52524E")
    ws["A5"] = "Workbook structure (7 sheets in total)"
    ws["A5"].font = Font(bold=True, size=11)
    instructions = [
        "  README           — this sheet (orientation, legend, sheet index, instructions).",
        "  Trade Payables   — vendors / suppliers (drives Confirmations + 43B(h)).",
        "  Trade Receivables — customers (drives Confirmations).",
        "  Unsecured Loans  — loan / advance counter-parties (drives Confirmations).",
        "  Bank Accounts    — bank ledgers (drives Bank Confirmations).",
        "  Others           — anything else with a closing balance.",
        "  MSME Details     — Trade Payables vendors only · Sector / MSME Type / Capital Goods drop-downs · drives 43B(h).",
        "",
        "Legend",
        "  • Pre-filled (read-only) cells are shown with a pale emerald background.",
        "  • Auditor-fill cells are shown with a pale amber background.",
        "",
        "How to use",
        "1. Party Name, Subhead / Group, Closing Balance, GSTIN, GST Reg Type and Country are pre-filled from your Tally Books JSON and Ledger Mapping XLSX. Do NOT edit them.",
        "2. On Trade Payables / Trade Receivables / Unsecured Loans — fill the Email column for every party that needs a balance confirmation.  Add an Alternate Email if you'd like a CC.",
        "3. On Bank Accounts — fill Bank Name, Account Number, IFSC and Email for every bank ledger.  These power the Bank Confirmation flow.",
        "4. On MSME Details — Trade Payables vendors only.  Set Sector / MSME Type / Capital Goods using the dropdowns.  Drives 43B(h) Disallowance.  Leave the row blank if the vendor is not MSME-registered.",
        "5. The Others sheet is a catch-all bucket — review it for any party that needs a confirmation request that didn't fall into the four named buckets.",
        "6. Save and re-upload via the Library panel — the engine will pick up the enrichment automatically and downstream modules will pin to the new version.",
        "",
        "Notes",
        "  • Closing Balance is shown with a Cr / Dr suffix and lakh-style grouping for human readability.",
        "  • The Subhead / Group column is the Tally Schedule III tag we use to bucket each ledger; it is informational only.",
        "  • Re-uploads are versioned — three live versions are retained per file-type, older versions are soft-deleted (30-day grace) and never deleted while a run is pinned to them.",
    ]
    for i, line in enumerate(instructions, start=6):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110


# ---------------------------------------------------------------------------
# Fixed Assets Register template — placeholder.
# A formal design is pending; today we ship a starter workbook with
# header columns matching the user's stated FA Register fields plus a
# README that clearly states "design TBD".  Auditor can fill the rows
# offline and re-upload.
# ---------------------------------------------------------------------------
async def generate_fa_register_template(
    *, firm_id: str, client_id: str, period: str, division: str | None,
) -> Tuple[bytes, str]:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme["A1"] = "Fixed Assets Register — AssureAI Audit Utilities"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = f"Generated for FY {period}"
    readme["A3"].font = Font(size=10, italic=True, color="52524E")
    notes = [
        "",
        "Status",
        "  • Final layout TBD — this is a starter template.",
        "  • Once finalised, this same file_type will be auto-populated from prior-year ITR JSON / 3CD JSON.",
        "",
        "Sheets",
        "  • Asset Register — one row per asset, fill all yellow columns.",
        "  • Disposals    — one row per asset disposed during the year.",
        "",
        "Save and re-upload via the Library panel.",
    ]
    for i, line in enumerate(notes, start=5):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 110

    # Asset Register sheet — header-only starter.
    ws = wb.create_sheet(title="Asset Register")
    headers = [
        "Asset Code", "Asset Description", "Asset Class / Block",
        "Date of Acquisition", "Vendor / Supplier", "Invoice / Voucher No",
        "Cost of Acquisition (Gross ₹)", "GST / Input Tax Adjustment (₹)",
        "Net Capitalised Cost (₹)", "Opening WDV — Books (₹)",
        "Opening WDV — IT (₹)", "Useful Life (yrs)", "Salvage Value (₹)",
        "Depreciation Method", "Companies Act Rate (%)", "IT Rate (%)",
        "Location / Cost Centre", "Notes",
    ]
    _write_header(ws, headers)
    _autosize(ws, headers, [])
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    # Disposals sheet — header-only starter.
    ws_d = wb.create_sheet(title="Disposals")
    d_headers = [
        "Asset Code", "Asset Description", "Date of Disposal",
        "Mode of Disposal", "Sale Consideration (₹)",
        "Accumulated Depreciation (₹)", "WDV on Disposal Date (₹)",
        "Profit / (Loss) on Disposal (₹)", "Buyer", "Notes",
    ]
    _write_header(ws_d, d_headers)
    _autosize(ws_d, d_headers, [])
    ws_d.freeze_panes = "A2"
    ws_d.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f"fa_register_template__{client_id}__{period}.xlsx"


# ---------------------------------------------------------------------------
# IT Depreciation — Opening WDV template.
# Reuses the existing per-run round-trip workbook builder in
# `modules/fixed_assets/block_opening_xlsx.py` so the Library template
# matches the format the FA module already imports back.
# ---------------------------------------------------------------------------
async def generate_it_depreciation_opening_wdv_template(
    *, firm_id: str, client_id: str, period: str, division: str | None,
) -> Tuple[bytes, str]:
    from modules.fixed_assets.block_opening_xlsx import build_workbook
    from modules.fixed_assets.legal_master import get_block_labels_active
    from core.db import db

    blocks = await get_block_labels_active()
    if not blocks:
        # First-time setup — seed and retry.
        from modules.fixed_assets.legal_master import seed_legal_master
        await seed_legal_master(force=False)
        blocks = await get_block_labels_active()

    cli = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    rows = [
        {
            "block_label": b["block_label"],
            "rate": float(b.get("rate") or 0),
            "opening_wdv": 0.0,
            "description": "",
            "source": "manual",
        }
        for b in sorted(blocks, key=lambda x: (-float(x.get("rate") or 0), x["block_label"]))
    ]
    run_ctx = {
        "client_name": (cli or {}).get("name") or "",
        "fy_label":    period,
        "fy_end":      period,
    }
    blob = build_workbook(run_ctx, rows)
    return blob, f"it_depreciation_opening_wdv__{client_id}__{period}.xlsx"


# ---------------------------------------------------------------------------
# Registry — file_type → generator.  Add new entries as we extend support.
# ---------------------------------------------------------------------------
TemplateGenerator = Callable[..., Awaitable[Tuple[bytes, str]]]
TEMPLATE_GENERATORS: dict[str, TemplateGenerator] = {
    "party_master_xlsx": generate_party_master_template,
    "fa_register_xlsx": generate_fa_register_template,
    "it_depreciation_opening_wdv_xlsx": generate_it_depreciation_opening_wdv_template,
}


def has_template(file_type: str) -> bool:
    return file_type in TEMPLATE_GENERATORS


async def generate_template(
    *, file_type: str, firm_id: str, client_id: str, period: str, division: str | None,
) -> Tuple[bytes, str]:
    gen = TEMPLATE_GENERATORS.get(file_type)
    if not gen:
        raise FileNotFoundError(f"No template generator registered for '{file_type}'")
    return await gen(
        firm_id=firm_id, client_id=client_id, period=period, division=division,
    )
