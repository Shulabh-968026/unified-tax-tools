"""Phase 6 — Side-by-side reconciliation for disputed responses.

When a recipient's "Not Confirmed" submission carries an attached statement,
we offer the auditor a 2-pane comparator:

  • LEFT (our books)   — extracted from Tally JSON via letter_pdf.find_ledger_vouchers
  • RIGHT (their books) — parsed from the recipient's uploaded XLSX / CSV

This module focuses on the parsing + auto-matching half. It tries to handle the
two formats CAs see most often in practice:

  XLSX with header row containing some combination of: Date, Voucher Type,
  Voucher #, Particulars / Narration, Debit, Credit, Balance.

  CSV with the same columns in any order.

We DO NOT attempt to parse arbitrary PDF statements in v1 — too brittle.
The auditor can always download the attachment and review it manually.

Auto-matching is amount-only (within ±₹1) — date matching is fragile when
recipients post on different dates than we book. Manual link/unlink in the UI
covers the gap.
"""
from __future__ import annotations
import csv
import io
import re
from typing import Any, Dict, List, Optional, Tuple

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False


# ============================ Column heuristics ============================
DATE_HINTS    = ("date", "txn date", "transaction date", "voucher date", "posting date")
VTYPE_HINTS   = ("voucher type", "type", "particulars", "vch type")
VNO_HINTS     = ("voucher", "voucher no", "voucher #", "vch no", "ref", "doc", "invoice no", "bill no")
NARR_HINTS    = ("narration", "particulars", "description", "remarks", "details")
DEBIT_HINTS   = ("debit", "dr", "dr amount", "withdrawal")
CREDIT_HINTS  = ("credit", "cr", "cr amount", "deposit")
AMOUNT_HINTS  = ("amount", "amt", "value")
BALANCE_HINTS = ("balance", "running balance", "closing")


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _detect_columns(headers: List[str]) -> Dict[str, int]:
    """Map our canonical column names → 0-based index into the row.

    Each header is matched against multiple hint sets; first non-conflicting
    match wins. We tolerate Particulars being either narration or voucher
    type — context handled by the consumer.
    """
    norm_headers = [_norm(h) for h in headers]
    out: Dict[str, int] = {}

    def _hit(idx: int, want: str):
        if want not in out:
            out[want] = idx

    for i, h in enumerate(norm_headers):
        if any(k in h for k in DATE_HINTS) and "date" not in out:
            _hit(i, "date")
        elif any(h == k or h.endswith(" " + k) or h.startswith(k) for k in DEBIT_HINTS):
            _hit(i, "debit")
        elif any(h == k or h.endswith(" " + k) or h.startswith(k) for k in CREDIT_HINTS):
            _hit(i, "credit")
        elif any(k in h for k in BALANCE_HINTS) and "balance" not in out:
            _hit(i, "balance")
        elif any(k in h for k in VNO_HINTS) and "vno" not in out:
            _hit(i, "vno")
        elif any(k in h for k in VTYPE_HINTS) and "vtype" not in out:
            _hit(i, "vtype")
        elif any(k in h for k in NARR_HINTS) and "narration" not in out:
            _hit(i, "narration")
        elif any(k in h for k in AMOUNT_HINTS) and "amount" not in out:
            _hit(i, "amount")
    return out


def _parse_amount(v: Any) -> float:
    if v in (None, "", "-", "–"):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    # Handle parentheses-as-negative: "(1,000.00)" → -1000.00
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return 0.0
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return 0.0


def _parse_date(v: Any) -> str:
    """Best-effort ISO-format conversion. If parsing fails, return raw string."""
    if v in (None, ""):
        return ""
    s = str(v).strip()
    # Already ISO?
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # Common Indian formats: dd-mm-yyyy / dd/mm/yyyy
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = f"20{y}"
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    return s[:10]


def _build_record(row: List[Any], col_map: Dict[str, int]) -> Dict[str, Any]:
    def _g(key: str, default: Any = "") -> Any:
        i = col_map.get(key)
        if i is None or i >= len(row):
            return default
        return row[i]

    debit  = _parse_amount(_g("debit"))
    credit = _parse_amount(_g("credit"))
    amount = _parse_amount(_g("amount"))
    if debit == 0 and credit == 0 and amount != 0:
        # CSV variant with single signed Amount column → treat positive as credit, negative as debit
        debit, credit = (0.0, amount) if amount > 0 else (-amount, 0.0)
    return {
        "date":      _parse_date(_g("date")),
        "vtype":     str(_g("vtype") or ""),
        "vno":       str(_g("vno") or ""),
        "narration": str(_g("narration") or "")[:200],
        "debit":     round(debit, 2),
        "credit":    round(credit, 2),
        "balance":   round(_parse_amount(_g("balance")), 2),
    }


# ============================ XLSX / CSV parsing ===========================
def parse_xlsx(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, int], List[str]]:
    if not HAS_XLSX:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    # Pick the first non-empty sheet
    ws = next((s for s in wb.worksheets if s.max_row and s.max_row > 1), wb.active)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], {}, []
    # Find the first row that looks like headers (>=3 string cells, contains 'date'/'amount'/'debit'/'credit')
    header_idx = 0
    for i, r in enumerate(rows[:25]):
        if not r:
            continue
        norm = [_norm(x) for x in r]
        if sum(1 for n in norm if n) < 3:
            continue
        joined = " | ".join(norm)
        if any(k in joined for k in ("date", "debit", "credit", "amount", "narration", "particulars")):
            header_idx = i
            break
    headers = [str(c or "") for c in rows[header_idx]]
    col_map = _detect_columns(headers)
    body = rows[header_idx + 1:]
    records: List[Dict[str, Any]] = []
    for r in body:
        if not r or all(c in (None, "", " ") for c in r):
            continue
        rec = _build_record(list(r), col_map)
        # Skip rows with no date AND no amount — they're sub-headers / footers
        if not rec["date"] and rec["debit"] == 0 and rec["credit"] == 0:
            continue
        records.append(rec)
    return records, col_map, headers


def parse_csv(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, int], List[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Sniff delimiter
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = list(csv.reader(io.StringIO(text), dialect))
    if not reader:
        return [], {}, []
    # First row that has >=3 non-empty cells is the header
    header_idx = next(
        (i for i, r in enumerate(reader[:25])
         if sum(1 for c in r if (c or "").strip()) >= 3),
        0,
    )
    headers = reader[header_idx]
    col_map = _detect_columns(headers)
    body = reader[header_idx + 1:]
    records: List[Dict[str, Any]] = []
    for r in body:
        if not r or all((c or "").strip() == "" for c in r):
            continue
        rec = _build_record(r, col_map)
        if not rec["date"] and rec["debit"] == 0 and rec["credit"] == 0:
            continue
        records.append(rec)
    return records, col_map, headers


def parse_recipient_statement(filename: str, content: bytes) -> Dict[str, Any]:
    """Detect by extension, return {records, col_map, headers, format, supported}."""
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext in ("xlsx", "xlsm"):
        recs, cm, hd = parse_xlsx(content)
        return {"records": recs, "col_map": cm, "headers": hd,
                "format": "xlsx", "supported": True}
    if ext == "csv":
        recs, cm, hd = parse_csv(content)
        return {"records": recs, "col_map": cm, "headers": hd,
                "format": "csv", "supported": True}
    return {"records": [], "col_map": {}, "headers": [],
            "format": ext or "unknown", "supported": False,
            "message": "Only XLSX and CSV statements are auto-parsed in v1. "
                       "PDF statements should be reviewed manually via the "
                       "attachment download."}


# ============================ Auto-match =====================================
def auto_match(ours: List[Dict[str, Any]],
               theirs: List[Dict[str, Any]],
               tolerance: float = 1.0) -> List[Dict[str, Any]]:
    """Greedy amount-based match. For each "our" row, find the closest "their"
    row whose net (credit - debit) magnitude is within `tolerance` rupees.
    Returns a list of pair rows for the side-by-side UI:

        {our_idx, their_idx, status: "match" | "ours_only" | "theirs_only",
         our: <our row or None>, theirs: <their row or None>, diff: float}
    """
    # Normalise our amount: we stored {amount: signed} via find_ledger_vouchers,
    # but the caller may pass pre-shaped {debit, credit}. Handle both.
    def _our_net(r: Dict[str, Any]) -> float:
        if "amount" in r:
            return _parse_amount(r["amount"])
        return _parse_amount(r.get("credit", 0)) - _parse_amount(r.get("debit", 0))

    def _their_net(r: Dict[str, Any]) -> float:
        return _parse_amount(r.get("credit", 0)) - _parse_amount(r.get("debit", 0))

    used_their = set()
    pairs: List[Dict[str, Any]] = []

    for oi, our in enumerate(ours):
        oa = _our_net(our)
        if oa == 0:
            continue
        # Find best candidate
        best_j: Optional[int] = None
        best_diff = float("inf")
        for tj, their in enumerate(theirs):
            if tj in used_their:
                continue
            ta = _their_net(their)
            # Recipient ledger uses opposite sign convention to us when they're
            # a counterparty: a credit on our books = debit on theirs. We
            # match on |amount| within tolerance to avoid sign confusion.
            d = abs(abs(oa) - abs(ta))
            if d < best_diff:
                best_diff = d
                best_j = tj
        if best_j is not None and best_diff <= tolerance:
            used_their.add(best_j)
            pairs.append({
                "our_idx": oi, "their_idx": best_j,
                "status": "match",
                "our": our, "theirs": theirs[best_j],
                "diff": round(best_diff, 2),
            })
        else:
            pairs.append({
                "our_idx": oi, "their_idx": None,
                "status": "ours_only",
                "our": our, "theirs": None,
                "diff": None,
            })

    # Append any remaining 'theirs_only' rows
    for tj, their in enumerate(theirs):
        if tj in used_their:
            continue
        pairs.append({
            "our_idx": None, "their_idx": tj,
            "status": "theirs_only",
            "our": None, "theirs": their,
            "diff": None,
        })
    return pairs
