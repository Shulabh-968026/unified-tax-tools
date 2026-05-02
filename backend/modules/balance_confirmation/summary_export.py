"""Phase 5 — Confirmation Summary Report builder.

Two outputs from the same source data:

  • build_summary_xlsx() — multi-sheet workbook for the auditor's working file
        Sheet 1  Cover            (run header + KPI table + status banner)
        Sheet 2  Sent Tracker     (one row per ledger-with-email, every status timestamp)
        Sheet 3  Status Timeline  (every event from bc_send_log, chrono per party)
        Sheet 4  Variances        (disputed responses with our vs their figures + reason)
        Sheet 5  Confirmed        (clean list of confirmed parties for sign-off)
        Sheet 6  Notes            (free-text working notes — auditor writes here)

  • build_summary_pdf() — single signature-ready PDF for the audit file
        Page 1  Cover + KPI cards + status banner
        Page 2  Variances table (top-N with reasons)
        Page 3  Confirmed parties summary
        Page 4  Sign-off

KPI rules
---------
status_chip ∈ {not_sent, queued, sent, delivered, opened, clicked, confirmed,
                disputed, bounced, failed}. We bucket as:
  • Confirmed   — terminal, audit-ready
  • Disputed    — terminal, needs reconciliation
  • In-flight   — sent / delivered / opened / clicked / queued
  • Failed      — bounced / failed
  • No Action   — not_sent (no email or never queued)
"""
from __future__ import annotations
import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.platypus import (
    PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)


# ============================ Theme ===========================================
GREEN     = colors.HexColor("#047857")
INK       = colors.HexColor("#111827")
INK_SOFT  = colors.HexColor("#374151")
MUTED     = colors.HexColor("#6B7280")
BORDER    = colors.HexColor("#E5E7EB")
ROW_ALT   = colors.HexColor("#F9FAFB")
SECTION   = colors.HexColor("#F3F4F6")
OK_BG     = colors.HexColor("#D1FAE5")
OK_FG     = colors.HexColor("#065F46")
WARN_BG   = colors.HexColor("#FEF3C7")
WARN_FG   = colors.HexColor("#92400E")
DANGER_BG = colors.HexColor("#FEE2E2")
DANGER_FG = colors.HexColor("#991B1B")
BLUE_BG   = colors.HexColor("#DBEAFE")
BLUE_FG   = colors.HexColor("#1E40AF")


# ============================ Helpers =========================================
def _f(v) -> float:
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def _inr(n) -> str:
    n = _f(n)
    if n == 0:
        return "–"
    s = f"{abs(n):,.2f}"
    return f"({s})" if n < 0 else s


def _inr_short(n) -> str:
    """Compact Indian currency — '12.4 Cr', '45.1 L', '87.5 K' — for tight cells."""
    n = _f(n)
    a = abs(n)
    if a == 0:
        return "–"
    if a >= 1e7:
        return f"{a / 1e7:.2f} Cr"
    if a >= 1e5:
        return f"{a / 1e5:.2f} L"
    if a >= 1e3:
        return f"{a / 1e3:.1f} K"
    return f"{a:,.0f}"


def _slim(iso: Optional[str]) -> str:
    if not iso:
        return ""
    return str(iso)[:19].replace("T", " ")


def kpi_buckets(ledgers: List[Dict[str, Any]]) -> Dict[str, int]:
    """Group every ledger into one of 6 KPI buckets."""
    out = {"confirmed": 0, "disputed": 0, "in_flight": 0,
           "failed": 0, "no_action": 0, "no_email": 0}
    for L in ledgers:
        st = L.get("confirmation_status") or "not_sent"
        if st == "confirmed":
            out["confirmed"] += 1
        elif st == "disputed":
            out["disputed"] += 1
        elif st in ("sent", "delivered", "opened", "clicked", "queued"):
            out["in_flight"] += 1
        elif st in ("bounced", "failed"):
            out["failed"] += 1
        elif not L.get("email"):
            out["no_email"] += 1
        else:
            out["no_action"] += 1
    return out


# ============================ Excel ==========================================
THIN = Side(style="thin", color="D4D4D8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor="111827")
HEAD_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="111827")
MUTED_FONT = Font(name="Calibri", size=9, color="6B7280")
KPI_FILL = PatternFill("solid", fgColor="ECFDF5")
ROW_ALT_FILL = PatternFill("solid", fgColor="F9FAFB")
DISPUTED_FILL = PatternFill("solid", fgColor="FEF3C7")
CONFIRMED_FILL = PatternFill("solid", fgColor="D1FAE5")
FAILED_FILL = PatternFill("solid", fgColor="FEE2E2")


def _xl_set_widths(ws, widths: List[Tuple[int, int]]):
    for col, w in widths:
        ws.column_dimensions[get_column_letter(col)].width = w


def _xl_header_row(ws, row: int, headers: List[str]):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BOX


def _xl_cover(ws, *, run, client, ledgers, kpi):
    ws.title = "Cover"
    _xl_set_widths(ws, [(1, 32), (2, 28), (3, 22), (4, 22)])
    ws["A1"] = "Balance Confirmation — Summary Report"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="047857")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Client:"
    ws["A3"].font = MUTED_FONT
    ws["B3"] = client.get("name", "")
    ws["B3"].font = Font(bold=True, size=12)
    ws["A4"] = "GSTIN:"
    ws["A4"].font = MUTED_FONT
    ws["B4"] = client.get("gstin") or "—"
    ws["B4"].font = BODY_FONT
    ws["A5"] = "Financial Year:"
    ws["A5"].font = MUTED_FONT
    ws["B5"] = run.get("fy", "")
    ws["B5"].font = BODY_FONT
    ws["A6"] = "As at date:"
    ws["A6"].font = MUTED_FONT
    ws["B6"] = run.get("as_at_date", "")
    ws["B6"].font = BODY_FONT
    ws["A7"] = "Auditor:"
    ws["A7"].font = MUTED_FONT
    ws["B7"] = run.get("created_by_name") or ""
    ws["B7"].font = BODY_FONT
    ws["A8"] = "Generated:"
    ws["A8"].font = MUTED_FONT
    ws["B8"] = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC")
    ws["B8"].font = BODY_FONT

    # KPI block
    ws["A10"] = "Status Summary"
    ws["A10"].font = Font(name="Calibri", size=12, bold=True, color="111827")
    rows = [
        ("Confirmed",  kpi["confirmed"],  CONFIRMED_FILL),
        ("Disputed",   kpi["disputed"],   DISPUTED_FILL),
        ("In flight",  kpi["in_flight"],  PatternFill("solid", fgColor="DBEAFE")),
        ("Failed / bounced", kpi["failed"], FAILED_FILL),
        ("No action / not sent", kpi["no_action"], None),
        ("Without email",  kpi["no_email"], None),
        ("Total ledgers",  sum(kpi.values()), PatternFill("solid", fgColor="F3F4F6")),
    ]
    r = 12
    for label, count, fill in rows:
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        c = ws.cell(row=r, column=2, value=count)
        c.font = Font(bold=True, size=12, color="111827")
        c.alignment = Alignment(horizontal="right")
        if fill is not None:
            ws.cell(row=r, column=1).fill = fill
            c.fill = fill
        ws.cell(row=r, column=1).border = BOX
        ws.cell(row=r, column=2).border = BOX
        r += 1
    ws.row_dimensions[1].height = 28


def _xl_sent_tracker(ws, ledgers):
    ws.title = "Sent Tracker"
    headers = ["Party", "Group", "Category", "Email", "Closing", "Dr/Cr", "Status",
               "Queued At", "Sent At", "Delivered At", "Opened At",
               "Clicked At", "Bounced At", "Last Reminder", "Attempts"]
    _xl_header_row(ws, 1, headers)
    _xl_set_widths(ws, [(1, 30), (2, 22), (3, 16), (4, 28), (5, 14), (6, 6), (7, 12),
                        (8, 18), (9, 18), (10, 18), (11, 18), (12, 18),
                        (13, 18), (14, 18), (15, 9)])
    r = 2
    for L in sorted(ledgers, key=lambda x: x.get("name", "")):
        ws.cell(row=r, column=1, value=L.get("name", ""))
        ws.cell(row=r, column=2, value=L.get("parent_group", ""))
        ws.cell(row=r, column=3, value=L.get("category", ""))
        ws.cell(row=r, column=4, value=L.get("email", ""))
        ws.cell(row=r, column=5, value=round(abs(_f(L.get("closing_balance"))), 2))
        ws.cell(row=r, column=6, value=(L.get("dr_cr") or "").upper())
        ws.cell(row=r, column=7, value=L.get("confirmation_status", ""))
        ws.cell(row=r, column=8,  value=_slim(L.get("queued_at")))
        ws.cell(row=r, column=9,  value=_slim(L.get("sent_at")))
        ws.cell(row=r, column=10, value=_slim(L.get("delivered_at")))
        ws.cell(row=r, column=11, value=_slim(L.get("opened_at")))
        ws.cell(row=r, column=12, value=_slim(L.get("clicked_at")))
        ws.cell(row=r, column=13, value=_slim(L.get("bounced_at")))
        ws.cell(row=r, column=14, value=_slim(L.get("last_reminded_at")))
        ws.cell(row=r, column=15, value=int(L.get("send_attempts") or 0))
        # row tinting by status
        st = L.get("confirmation_status", "")
        fill = {"confirmed": CONFIRMED_FILL, "disputed": DISPUTED_FILL,
                "bounced": FAILED_FILL, "failed": FAILED_FILL}.get(st)
        if fill is None and r % 2 == 0:
            fill = ROW_ALT_FILL
        if fill is not None:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).border = BOX
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        r += 1
    ws.freeze_panes = "B2"


def _xl_status_timeline(ws, send_log: List[Dict[str, Any]],
                        ledger_name_by_id: Dict[str, str]):
    ws.title = "Status Timeline"
    headers = ["Timestamp", "Party", "Kind", "Status", "Resend ID", "To Email",
               "Subject / Note", "Actor"]
    _xl_header_row(ws, 1, headers)
    _xl_set_widths(ws, [(1, 20), (2, 30), (3, 12), (4, 12), (5, 28),
                        (6, 28), (7, 36), (8, 24)])
    r = 2
    for ev in sorted(send_log, key=lambda x: x.get("ts", "")):
        ws.cell(row=r, column=1, value=_slim(ev.get("ts")))
        ws.cell(row=r, column=2, value=ledger_name_by_id.get(ev.get("ledger_id", ""), ""))
        ws.cell(row=r, column=3, value=ev.get("kind", ""))
        ws.cell(row=r, column=4, value=ev.get("status", ""))
        ws.cell(row=r, column=5, value=ev.get("resend_id", ""))
        ws.cell(row=r, column=6, value=ev.get("to_email", ""))
        ws.cell(row=r, column=7, value=ev.get("subject", "") or ev.get("error", "") or "")
        ws.cell(row=r, column=8, value=ev.get("actor_email", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).border = BOX
        if r % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = ROW_ALT_FILL
        r += 1
    ws.freeze_panes = "A2"


def _xl_variances(ws, responses: List[Dict[str, Any]],
                  ledger_by_id: Dict[str, Dict[str, Any]]):
    ws.title = "Variances"
    headers = ["Party", "Our Books", "Our Dr/Cr", "Their Books", "Their Dr/Cr",
               "Difference", "Reason", "Responder", "Email", "Submitted",
               "Attachment"]
    _xl_header_row(ws, 1, headers)
    _xl_set_widths(ws, [(1, 30), (2, 16), (3, 8), (4, 16), (5, 8),
                        (6, 16), (7, 50), (8, 24), (9, 28), (10, 18), (11, 24)])
    r = 2
    disputed = [x for x in responses if x.get("decision") == "disputed"]
    for resp in sorted(disputed, key=lambda x: x.get("submitted_at", "")):
        L = ledger_by_id.get(resp.get("ledger_id", "")) or {}
        our_bal = abs(_f(L.get("closing_balance")))
        their_bal = _f(resp.get("their_balance"))
        diff = round(their_bal - our_bal, 2)
        ws.cell(row=r, column=1, value=L.get("name", ""))
        ws.cell(row=r, column=2, value=our_bal)
        ws.cell(row=r, column=3, value=(L.get("dr_cr") or "").upper())
        ws.cell(row=r, column=4, value=their_bal if resp.get("their_balance") is not None else "")
        ws.cell(row=r, column=5, value=resp.get("their_dr_cr", ""))
        ws.cell(row=r, column=6, value=diff)
        ws.cell(row=r, column=7, value=resp.get("reason", ""))
        ws.cell(row=r, column=8, value=resp.get("responder_name", ""))
        ws.cell(row=r, column=9, value=resp.get("responder_email", ""))
        ws.cell(row=r, column=10, value=_slim(resp.get("submitted_at")))
        ws.cell(row=r, column=11, value=resp.get("uploaded_filename", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        # tint by significance
        if abs(diff) > 1:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = DISPUTED_FILL
        r += 1
    if r == 2:  # no disputes
        ws.cell(row=2, column=1, value="No disputed responses on file.")
        ws.cell(row=2, column=1).font = MUTED_FONT
    ws.freeze_panes = "B2"


def _xl_confirmed(ws, responses: List[Dict[str, Any]],
                  ledger_by_id: Dict[str, Dict[str, Any]]):
    ws.title = "Confirmed"
    headers = ["Party", "Email", "Our Books", "Dr/Cr", "Confirmed By",
               "Confirmed Email", "Submitted", "Note"]
    _xl_header_row(ws, 1, headers)
    _xl_set_widths(ws, [(1, 30), (2, 28), (3, 16), (4, 8), (5, 24),
                        (6, 28), (7, 18), (8, 36)])
    r = 2
    confirmed = [x for x in responses if x.get("decision") == "confirmed"]
    for resp in sorted(confirmed, key=lambda x: x.get("submitted_at", "")):
        L = ledger_by_id.get(resp.get("ledger_id", "")) or {}
        ws.cell(row=r, column=1, value=L.get("name", ""))
        ws.cell(row=r, column=2, value=L.get("email", ""))
        ws.cell(row=r, column=3, value=abs(_f(L.get("closing_balance"))))
        ws.cell(row=r, column=4, value=(L.get("dr_cr") or "").upper())
        ws.cell(row=r, column=5, value=resp.get("responder_name", ""))
        ws.cell(row=r, column=6, value=resp.get("responder_email", ""))
        ws.cell(row=r, column=7, value=_slim(resp.get("submitted_at")))
        ws.cell(row=r, column=8, value=resp.get("note", ""))
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).font = BODY_FONT
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).fill = CONFIRMED_FILL
        ws.cell(row=r, column=3).number_format = "#,##0.00"
        r += 1
    if r == 2:
        ws.cell(row=2, column=1, value="No confirmed responses on file.")
        ws.cell(row=2, column=1).font = MUTED_FONT
    ws.freeze_panes = "B2"


def _xl_notes(ws):
    ws.title = "Notes"
    ws["A1"] = "Auditor working notes"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A3"] = ("Use this sheet for auditor-side observations: parties pending "
                "follow-up, sample selection rationale, materiality decisions, "
                "or sign-off remarks. This sheet is left blank intentionally.")
    ws["A3"].font = MUTED_FONT
    ws.column_dimensions["A"].width = 100
    ws.merge_cells("A3:F3")


def build_summary_xlsx(*, run: Dict[str, Any],
                       client: Dict[str, Any],
                       ledgers: List[Dict[str, Any]],
                       responses: List[Dict[str, Any]],
                       send_log: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    cover = wb.active

    kpi = kpi_buckets(ledgers)
    ledger_by_id = {L["ledger_id"]: L for L in ledgers if L.get("ledger_id")}
    ledger_name_by_id = {k: v.get("name", "") for k, v in ledger_by_id.items()}

    _xl_cover(cover, run=run, client=client, ledgers=ledgers, kpi=kpi)
    _xl_sent_tracker(wb.create_sheet(), ledgers)
    _xl_status_timeline(wb.create_sheet(), send_log, ledger_name_by_id)
    _xl_variances(wb.create_sheet(), responses, ledger_by_id)
    _xl_confirmed(wb.create_sheet(), responses, ledger_by_id)
    _xl_notes(wb.create_sheet())

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================ PDF ============================================
def _styles():
    base = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("h1", parent=base["Heading1"],
                             fontName="Helvetica-Bold", fontSize=16, leading=20,
                             textColor=INK, spaceAfter=4),
        "h2": ParagraphStyle("h2", parent=base["Heading2"],
                             fontName="Helvetica-Bold", fontSize=12, leading=15,
                             textColor=GREEN, spaceBefore=10, spaceAfter=6),
        "label": ParagraphStyle("label", parent=base["Normal"],
                                fontName="Helvetica", fontSize=7.5, leading=9,
                                textColor=MUTED),
        "labelMono": ParagraphStyle("labelMono", parent=base["Normal"],
                                    fontName="Courier", fontSize=8, leading=10,
                                    textColor=MUTED),
        "body": ParagraphStyle("body", parent=base["Normal"],
                               fontName="Helvetica", fontSize=9, leading=12,
                               textColor=INK_SOFT),
        "kpiNum": ParagraphStyle("kpiNum", parent=base["Normal"],
                                 fontName="Helvetica-Bold", fontSize=22, leading=24,
                                 textColor=INK),
        "kpiLabel": ParagraphStyle("kpiLabel", parent=base["Normal"],
                                   fontName="Helvetica", fontSize=8, leading=10,
                                   textColor=MUTED),
        "small": ParagraphStyle("small", parent=base["Normal"],
                                fontName="Helvetica", fontSize=7.5, leading=10,
                                textColor=MUTED),
    }


def _kpi_card(num: int, label: str, bg, fg, S):
    # Convert reportlab Color to '#RRGGBB' for inline HTML font tag
    fg_hex = "#{:02X}{:02X}{:02X}".format(int(fg.red * 255), int(fg.green * 255), int(fg.blue * 255))
    inner = Table([
        [Paragraph(f"<b><font color='{fg_hex}'>{num}</font></b>", S["kpiNum"])],
        [Paragraph(label, S["kpiLabel"])],
    ], colWidths=[55 * mm])
    inner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    return inner


def _kpi_grid(kpi: Dict[str, int], S):
    cells = [
        _kpi_card(kpi["confirmed"], "CONFIRMED",       OK_BG,    OK_FG,    S),
        _kpi_card(kpi["disputed"],  "DISPUTED",        WARN_BG,  WARN_FG,  S),
        _kpi_card(kpi["in_flight"], "IN FLIGHT",       BLUE_BG,  BLUE_FG,  S),
        _kpi_card(kpi["failed"],    "FAILED · BOUNCED", DANGER_BG, DANGER_FG, S),
    ]
    grid = Table([[cells[0], cells[1], cells[2], cells[3]]],
                 colWidths=[45 * mm] * 4, hAlign="LEFT")
    grid.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return grid


def _status_banner(kpi: Dict[str, int], total: int, S):
    pct = ((kpi["confirmed"] + kpi["disputed"]) / total * 100.0) if total else 0.0
    if kpi["disputed"] > 0:
        bg, fg = WARN_BG, WARN_FG
        label = (f"{kpi['confirmed']} confirmed · {kpi['disputed']} need reconciliation "
                 f"({pct:.0f}% response rate)")
    elif kpi["confirmed"] == total and total > 0:
        bg, fg = OK_BG, OK_FG
        label = f"All {total} parties confirmed — audit-ready"
    else:
        bg, fg = BLUE_BG, BLUE_FG
        label = (f"{kpi['confirmed']} confirmed · {kpi['in_flight']} awaiting response "
                 f"({pct:.0f}% response rate)")
    t = Table([[Paragraph(f"<b>{label}</b>", S["body"])]], colWidths=[180 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg),
        ("TEXTCOLOR", (0, 0), (-1, -1), fg),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
    ]))
    return t


def _client_block(run, client, S):
    rows = [
        [Paragraph("CLIENT", S["label"]),
         Paragraph(f"<b>{client.get('name','—')}</b>", S["body"]),
         Paragraph("FY", S["label"]),
         Paragraph(run.get("fy", "—"), S["body"])],
        [Paragraph("GSTIN", S["label"]),
         Paragraph(client.get("gstin") or "—", S["labelMono"]),
         Paragraph("AS AT", S["label"]),
         Paragraph(run.get("as_at_date", "—"), S["body"])],
        [Paragraph("AUDITOR", S["label"]),
         Paragraph(run.get("created_by_name") or "—", S["body"]),
         Paragraph("GENERATED", S["label"]),
         Paragraph(datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC"),
                   S["labelMono"])],
    ]
    t = Table(rows, colWidths=[26 * mm, 70 * mm, 22 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]))
    return t


def _variance_table(disputed: List[Dict[str, Any]],
                    ledger_by_id: Dict[str, Dict[str, Any]], S, top_n: int = 12):
    rows: List[List[Any]] = [["Party", "Our Books", "Their Books", "Diff", "Reason"]]
    sorted_d = sorted(
        disputed,
        key=lambda r: abs(_f(r.get("their_balance")) - abs(_f(
            (ledger_by_id.get(r.get("ledger_id"), {})).get("closing_balance")))),
        reverse=True,
    )
    extra = max(0, len(sorted_d) - top_n)
    for resp in sorted_d[:top_n]:
        L = ledger_by_id.get(resp.get("ledger_id", "")) or {}
        our = abs(_f(L.get("closing_balance")))
        their = _f(resp.get("their_balance"))
        diff = round(their - our, 2)
        reason = (resp.get("reason") or "")[:120]
        if len(resp.get("reason") or "") > 120:
            reason += "…"
        rows.append([
            (L.get("name") or "")[:40],
            f"{_inr(our)} {(L.get('dr_cr') or '').upper()}",
            f"{_inr(their)} {resp.get('their_dr_cr','')}" if resp.get("their_balance") is not None else "—",
            _inr(diff),
            Paragraph(reason, S["small"]),
        ])
    if extra:
        rows.append([f"+ {extra} more disputes (see Excel for full list)", "", "", "", ""])

    t = Table(rows, colWidths=[42 * mm, 28 * mm, 28 * mm, 22 * mm, 60 * mm], repeatRows=1)
    style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (3, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, INK),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    if extra:
        more_row = len(rows) - 1
        style += [
            ("SPAN", (0, more_row), (-1, more_row)),
            ("FONTSIZE", (0, more_row), (-1, more_row), 7),
            ("TEXTCOLOR", (0, more_row), (-1, more_row), MUTED),
            ("ALIGN", (0, more_row), (-1, more_row), "LEFT"),
            ("BACKGROUND", (0, more_row), (-1, more_row), ROW_ALT),
        ]
    t.setStyle(TableStyle(style))
    return t


def _confirmed_table(confirmed: List[Dict[str, Any]],
                     ledger_by_id: Dict[str, Dict[str, Any]], S):
    rows: List[List[Any]] = [["Party", "Email", "Books", "Confirmed By", "On"]]
    for resp in sorted(confirmed, key=lambda r: r.get("submitted_at", "")):
        L = ledger_by_id.get(resp.get("ledger_id", "")) or {}
        rows.append([
            (L.get("name") or "")[:40],
            (L.get("email") or "")[:32],
            f"{_inr(abs(_f(L.get('closing_balance'))))} {(L.get('dr_cr') or '').upper()}",
            (resp.get("responder_name") or "")[:24],
            _slim(resp.get("submitted_at"))[:10],
        ])
    t = Table(rows, colWidths=[44 * mm, 44 * mm, 30 * mm, 36 * mm, 26 * mm], repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), GREEN),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, OK_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _signoff(S):
    cells = [
        [Paragraph("PREPARED BY", S["label"]),
         Paragraph("REVIEWED BY", S["label"]),
         Paragraph("DATE", S["label"])],
        [Paragraph("&nbsp;<br/>&nbsp;<br/>______________________", S["body"]),
         Paragraph("&nbsp;<br/>&nbsp;<br/>______________________", S["body"]),
         Paragraph("&nbsp;<br/>&nbsp;<br/>______________________", S["body"])],
        [Paragraph("Name & Signature", S["small"]),
         Paragraph("Name & Signature", S["small"]),
         Paragraph("DD-MM-YYYY", S["small"])],
    ]
    t = Table(cells, colWidths=[60 * mm, 60 * mm, 60 * mm])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("LINEBETWEEN", (0, 0), (-1, -1), 0.3, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    return t


def _make_footer_drawer(S, run):
    rid = run.get("id", "")[:18]
    fy = run.get("fy", "")
    def _draw(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(MUTED)
        text = f"Balance Confirmation Summary · FY {fy} · Run {rid} · Page {doc.page}"
        canvas.drawString(15 * mm, 10 * mm, text)
        canvas.drawRightString(A4[0] - 15 * mm, 10 * mm,
                               "MSS × Assure · Audit Utilities")
        canvas.restoreState()
    return _draw


# ============================ Dashboard helpers (analytics-driven) ==========
# Status colour map shared with the frontend dashboard.
_STATUS_COLOR = {
    "confirmed":  (colors.HexColor("#D1FAE5"), colors.HexColor("#065F46")),
    "reconciled": (colors.HexColor("#CFFAFE"), colors.HexColor("#0E7490")),
    "disputed":   (colors.HexColor("#FEF3C7"), colors.HexColor("#92400E")),
    "in_flight":  (colors.HexColor("#DBEAFE"), colors.HexColor("#1E40AF")),
    "failed":     (colors.HexColor("#FEE2E2"), colors.HexColor("#991B1B")),
    "not_sent":   (colors.HexColor("#F3F4F6"), colors.HexColor("#4B5563")),
}
_STATUS_LABEL = {
    "confirmed":  "Confirmed",
    "reconciled": "Reconciled",
    "disputed":   "Disputed",
    "in_flight":  "In flight",
    "failed":     "Failed",
    "not_sent":   "Not sent",
}


def _hero_strip(overall: Dict[str, Any], S):
    """Total parties · Total ₹ · Coverage (count) · Coverage (₹)."""
    cov = overall["coverage"]
    cells = [
        _kpi_card(overall["count"],            "TOTAL PARTIES",    colors.HexColor("#F3F4F6"), INK, S),
        _kpi_card(_inr_short(overall['amount']), "TOTAL EXPOSURE", colors.HexColor("#ECFDF5"), GREEN, S),
        _kpi_card(f"{cov['audit_count_pct']}%", "COVERAGE · BY COUNT", colors.HexColor("#E0F2FE"), colors.HexColor("#075985"), S),
        _kpi_card(f"{cov['audit_amount_pct']}%", "COVERAGE · BY ₹",    colors.HexColor("#D1FAE5"), OK_FG, S),
    ]
    grid = Table([cells], colWidths=[45 * mm] * 4, hAlign="LEFT")
    grid.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return grid


_STATUS_TEXT_HEX = {
    "confirmed":  "065F46",
    "reconciled": "0E7490",
    "disputed":   "92400E",
    "in_flight":  "1E40AF",
    "failed":     "991B1B",
    "not_sent":   "4B5563",
}

# Solid-colour palette for the stacked bar & legend swatches (one tone richer
# than the tint used for text — gives the bar actual visual punch on paper).
_STATUS_SOLID = {
    "confirmed":  colors.HexColor("#10B981"),
    "reconciled": colors.HexColor("#06B6D4"),
    "disputed":   colors.HexColor("#F59E0B"),
    "in_flight":  colors.HexColor("#3B82F6"),
    "failed":     colors.HexColor("#EF4444"),
    "not_sent":   colors.HexColor("#9CA3AF"),
}


def _status_bar(bucket: Dict[str, Any], S, width_mm: float = 82.0,
                height_mm: float = 6.0):
    """₹-weighted stacked bar drawn as a single Drawing with filled Rect's.
    Using reportlab.graphics guarantees the bar renders as a solid coloured
    strip regardless of nested-table padding quirks."""
    W = width_mm * mm
    H = height_mm * mm
    d = Drawing(W, H)

    total = bucket["amount"] or 0.0
    # Outline
    d.add(Rect(0, 0, W, H, strokeColor=BORDER, strokeWidth=0.4,
               fillColor=colors.HexColor("#F3F4F6")))

    if total <= 0.0001:
        d.add(String(W / 2, H / 2 - 2.4, "No exposure",
                     fontName="Helvetica", fontSize=6.5,
                     fillColor=colors.HexColor("#9CA3AF"),
                     textAnchor="middle"))
        return d

    x = 0.0
    for st in ("confirmed", "reconciled", "disputed", "in_flight", "failed", "not_sent"):
        a = bucket["by_status"][st]["amount"]
        if a <= 0:
            continue
        w = (a / total) * W
        d.add(Rect(x, 0, w, H, strokeColor=None, strokeWidth=0,
                   fillColor=_STATUS_SOLID[st]))
        pct = a / total * 100.0
        if pct >= 14 and w >= 8 * mm:
            d.add(String(x + w / 2, H / 2 - 2.2, f"{pct:.0f}%",
                         fontName="Helvetica-Bold", fontSize=6.5,
                         fillColor=colors.white, textAnchor="middle"))
        x += w

    # Redraw outer stroke on top for crisp edge
    d.add(Rect(0, 0, W, H, strokeColor=BORDER, strokeWidth=0.4,
               fillColor=None))
    return d


def _legend_swatch(st: str, count: int, S):
    """Small filled colour square + label + count as a single Drawing — avoids
    the Helvetica ■ tofu issue we had with Paragraph-based legends."""
    w = 38 * mm
    h = 4.5 * mm
    d = Drawing(w, h)
    d.add(Rect(0, (h - 2.6 * mm) / 2, 2.6 * mm, 2.6 * mm,
               strokeColor=None, fillColor=_STATUS_SOLID[st]))
    d.add(String(4 * mm, h / 2 - 2.3, _STATUS_LABEL[st],
                 fontName="Helvetica", fontSize=7.5,
                 fillColor=INK_SOFT))
    # Count at right end
    d.add(String(w - 1, h / 2 - 2.3, str(count),
                 fontName="Helvetica-Bold", fontSize=7.5,
                 fillColor=INK, textAnchor="end"))
    return d


def _category_card(cat: Dict[str, Any], S):
    cov = cat["coverage"]
    # Header row: label (left) + parties & ₹ (right)
    head = Table([[
        Paragraph(f"<b>{cat['label']}</b>", S["body"]),
        Paragraph(f"<font color='#6B7280'>{cat['count']} parties · "
                  f"{_inr_short(cat['amount'])}</font>", S["small"]),
    ]], colWidths=[34 * mm, 48 * mm])
    head.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))

    bar = _status_bar(cat, S, width_mm=82)

    # 2×3 grid of legend swatches (drawings, not paragraphs)
    sw = [_legend_swatch(st, cat["by_status"][st]["count"], S)
          for st in ("confirmed", "reconciled", "disputed",
                     "in_flight", "failed", "not_sent")]
    legend = Table([sw[:3], sw[3:]], colWidths=[27 * mm, 27 * mm, 27 * mm],
                   rowHeights=[5 * mm, 5 * mm])
    legend.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    cov_line = Paragraph(
        f"<font color='#065F46' size='8.5'><b>Audit coverage "
        f"{cov['audit_amount_pct']}%</b></font> "
        f"<font color='#6B7280' size='7.5'>by ₹ · "
        f"{cov['audit_count_pct']}% by count · "
        f"Response {cov['response_amount_pct']}% by ₹</font>",
        S["small"])

    inner = Table([
        [head],
        [bar],
        [legend],
        [cov_line],
    ], colWidths=[82 * mm])
    inner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("BOX",       (0, 0), (-1, -1), 0.5, BORDER),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING",   (0, 0), (0, 0), 4),
        ("BOTTOMPADDING", (0, 0), (0, 0), 3),
        ("TOPPADDING",   (0, 1), (0, 1), 0),
        ("BOTTOMPADDING", (0, 1), (0, 1), 4),
        ("TOPPADDING",   (0, 2), (0, 2), 0),
        ("BOTTOMPADDING", (0, 2), (0, 2), 3),
        ("TOPPADDING",   (0, 3), (0, 3), 2),
        ("BOTTOMPADDING", (0, 3), (0, 3), 4),
    ]))
    return inner


def _category_grid(categories: List[Dict[str, Any]], S):
    """2×N grid of category cards — only emit cards for non-empty categories
    but keep the 4 auditor-facing ones in order."""
    visible = [c for c in categories if c["key"] != "other" or c["count"] > 0]
    # Render 2 columns
    rows = []
    pair = []
    for c in visible:
        pair.append(_category_card(c, S))
        if len(pair) == 2:
            rows.append(pair)
            pair = []
    if pair:
        pair.append("")  # pad
        rows.append(pair)

    t = Table(rows, colWidths=[90 * mm, 90 * mm], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    return t


def _funnel_table(funnel: List[Dict[str, Any]], S):
    rows = [["Stage", "Parties", "% of total", "Value (₹)", "% of ₹"]]
    for s in funnel:
        rows.append([
            s["label"],
            str(s["count"]),
            f"{s['count_pct']:.1f}%",
            _inr(s["amount"]),
            f"{s['amount_pct']:.1f}%",
        ])
    t = Table(rows, colWidths=[40 * mm, 24 * mm, 28 * mm, 36 * mm, 28 * mm],
              repeatRows=1)
    style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, INK),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    # Accent the final "Responded" row
    last = len(rows) - 1
    style += [
        ("FONTNAME", (0, last), (-1, last), "Helvetica-Bold"),
        ("BACKGROUND", (0, last), (-1, last), OK_BG),
        ("TEXTCOLOR", (0, last), (-1, last), OK_FG),
    ]
    t.setStyle(TableStyle(style))
    return t


def _top_disputed_table(rows: List[Dict[str, Any]], S):
    if not rows:
        return Paragraph("<font color='#9CA3AF'>No disputed confirmations on file.</font>",
                         S["body"])
    data = [["Party", "Cat", "Our ₹", "Their ₹", "Diff", "Reason"]]
    cat_abbr = {"trade_receivable": "Rec", "trade_payable": "Pay",
                "bank": "Bank", "unsecured_loans": "Loan", "other": "Oth"}
    for r in rows[:10]:
        data.append([
            (r["party"] or "")[:28],
            cat_abbr.get(r["category"], "—"),
            f"{_inr(r['our_amount'])} {r['our_dr_cr']}",
            ("—" if r["their_amount"] is None
             else f"{_inr(r['their_amount'])} {r['their_dr_cr']}"),
            ("—" if r["diff"] is None else _inr(r["diff"])),
            Paragraph((r["reason"] or "")[:100], S["small"]),
        ])
    t = Table(data, colWidths=[36 * mm, 10 * mm, 22 * mm, 22 * mm, 20 * mm, 70 * mm],
              repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), WARN_FG),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (2, 0), (4, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, WARN_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _top_unresponsive_table(rows: List[Dict[str, Any]], S):
    if not rows:
        return Paragraph("<font color='#9CA3AF'>No stale in-flight confirmations.</font>",
                         S["body"])
    data = [["Party", "Cat", "₹ Exposure", "Days pending", "Status", "Email"]]
    cat_abbr = {"trade_receivable": "Rec", "trade_payable": "Pay",
                "bank": "Bank", "unsecured_loans": "Loan", "other": "Oth"}
    for r in rows[:10]:
        data.append([
            (r["party"] or "")[:28],
            cat_abbr.get(r["category"], "—"),
            f"{_inr(r['amount'])} {r['dr_cr']}",
            f"{r['days_pending']}d",
            (r["status"] or "").title(),
            (r["email"] or "")[:28],
        ])
    t = Table(data, colWidths=[36 * mm, 10 * mm, 28 * mm, 20 * mm, 20 * mm, 46 * mm],
              repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (2, 0), (4, -1), "RIGHT"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BLUE_BG]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _subhead_heatmap_table(rows: List[Dict[str, Any]], S):
    if not rows:
        return Paragraph("<font color='#9CA3AF'>No subhead data.</font>", S["body"])
    data = [["Subhead", "Parties", "Exposure (₹)", "Coverage ₹ %", "Response ₹ %"]]
    for r in rows[:12]:
        data.append([
            (r["parent_group"] or "—")[:42],
            str(r["count"]),
            _inr(r["amount"]),
            f"{r['audit_amount_pct']:.1f}%",
            f"{r['response_amount_pct']:.1f}%",
        ])
    t = Table(data, colWidths=[60 * mm, 18 * mm, 32 * mm, 28 * mm, 28 * mm],
              repeatRows=1)
    style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    # Colour the coverage column by intensity
    for i, r in enumerate(rows[:12], start=1):
        pct = r["audit_amount_pct"]
        if pct >= 80:
            bg = colors.HexColor("#A7F3D0")
        elif pct >= 50:
            bg = colors.HexColor("#D1FAE5")
        elif pct >= 25:
            bg = colors.HexColor("#FEF3C7")
        elif pct > 0:
            bg = colors.HexColor("#FEE2E2")
        else:
            bg = colors.HexColor("#F3F4F6")
        style.append(("BACKGROUND", (3, i), (3, i), bg))
    t.setStyle(TableStyle(style))
    return t


def build_summary_pdf(*, run: Dict[str, Any],
                      client: Dict[str, Any],
                      ledgers: List[Dict[str, Any]],
                      responses: List[Dict[str, Any]],
                      analytics: Optional[Dict[str, Any]] = None) -> bytes:
    """Dashboard-style audit summary.

    Page 1  Cover + Hero KPIs + Category Matrix
    Page 2  Confirmation Funnel + Top Disputed
    Page 3  Top Unresponsive + Subhead Heatmap
    Page 4  Variances detail  (if any disputes)
    Page 5  Confirmed parties (if any confirmed)
    Page 6  Sign-off
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title="Balance Confirmation — Summary Report",
        author="MSS × Assure",
    )
    S = _styles()
    story: List[Any] = []
    kpi = kpi_buckets(ledgers)
    total = len(ledgers)
    ledger_by_id = {L["ledger_id"]: L for L in ledgers if L.get("ledger_id")}

    # ---------- Page 1 · Cover + Hero + Category Matrix ---------------------
    story.append(Paragraph("Balance Confirmation — Summary Report", S["h1"]))
    story.append(Paragraph("Audit Working-Paper · Section 44AB", S["body"]))
    story.append(Spacer(1, 4))
    story.append(_client_block(run, client, S))
    story.append(Spacer(1, 10))

    if analytics:
        story.append(Paragraph("Audit Coverage at a Glance", S["h2"]))
        story.append(_hero_strip(analytics["overall"], S))
        story.append(Spacer(1, 8))
        story.append(_status_banner(kpi, total, S))
        story.append(Spacer(1, 10))
        story.append(Paragraph("Category Matrix", S["h2"]))
        story.append(_category_grid(analytics["categories"], S))

        # ---------- Page 2 · Funnel + Top Disputed --------------------------
        story.append(PageBreak())
        story.append(Paragraph("Confirmation Funnel", S["h2"]))
        story.append(Paragraph(
            "<font color='#6B7280'>Drop-off from total ledgers identified to "
            "recipients who responded — amount-weighted against total exposure.</font>",
            S["body"]))
        story.append(Spacer(1, 4))
        story.append(_funnel_table(analytics["funnel"], S))

        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Top Disputed Confirmations · by value variance "
            f"({len(analytics['top_disputed'])})",
            S["h2"]))
        story.append(_top_disputed_table(analytics["top_disputed"], S))

        # ---------- Page 3 · Unresponsive + Subhead Heatmap -----------------
        story.append(PageBreak())
        story.append(Paragraph(
            f"Stale / Unresponsive · sent 7+ days ago, awaiting reply "
            f"({len(analytics['top_unresponsive'])})",
            S["h2"]))
        story.append(_top_unresponsive_table(analytics["top_unresponsive"], S))

        story.append(Spacer(1, 10))
        story.append(Paragraph("Subhead Coverage Heatmap", S["h2"]))
        story.append(Paragraph(
            "<font color='#6B7280'>Coverage % is computed as (confirmed + reconciled) ₹ "
            "÷ subhead total ₹. Useful for sampling rationale and gap identification.</font>",
            S["body"]))
        story.append(Spacer(1, 4))
        story.append(_subhead_heatmap_table(analytics["subheads"], S))
    else:
        story.append(Paragraph("Status Summary", S["h2"]))
        story.append(_status_banner(kpi, total, S))
        story.append(Spacer(1, 8))
        story.append(_kpi_grid(kpi, S))

    # ---------- Page 4 · Variances (detail, all rows) -----------------------
    disputed = [r for r in responses if r.get("decision") == "disputed"]
    if disputed:
        story.append(PageBreak())
        story.append(Paragraph(
            f"Disputed Confirmations — Detail · {len(disputed)} party(ies)", S["h2"]))
        story.append(_variance_table(disputed, ledger_by_id, S))

    # ---------- Page 5 · Confirmed ------------------------------------------
    confirmed = [r for r in responses if r.get("decision") == "confirmed"]
    if confirmed:
        story.append(PageBreak())
        story.append(Paragraph(
            f"Confirmed Parties · {len(confirmed)} response(s)", S["h2"]))
        story.append(_confirmed_table(confirmed, ledger_by_id, S))

    # ---------- Page 6 · Sign-off -------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Sign-off", S["h2"]))
    story.append(Paragraph(
        "Notes: This summary is auto-generated from the run's ledger data, "
        "send-log and recipient responses captured via the public confirmation "
        "link. The dashboard on page 1 mirrors the live on-screen analytics "
        "and is the primary audit sign-off reference. Disputed responses whose "
        "reconciliation has been logged by the auditor are counted under "
        "'Reconciled' in the coverage metrics. The full per-party send timeline "
        "and raw recipient submissions are available in the companion Excel "
        "workbook (Sent Tracker · Status Timeline · Variances · Confirmed).",
        S["body"]))
    story.append(Spacer(1, 16))
    story.append(_signoff(S))

    drawer = _make_footer_drawer(S, run)
    doc.build(story, onFirstPage=drawer, onLaterPages=drawer)
    return buf.getvalue()
