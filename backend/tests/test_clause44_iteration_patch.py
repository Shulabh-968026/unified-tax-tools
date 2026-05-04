"""Unit tests for the 3-fix iteration (company-name guard, 6-col pivot Excel,
per-cohort sheets).
"""
import asyncio
import io
import openpyxl
from modules.clause44.controller import (
    _company_names_match, _extract_company_name, _normalise_name,
)
from modules.clause44.exports import build_export_response


def _read_export_bytes(run):
    """Consume the StreamingResponse's async body iterator synchronously."""
    resp = build_export_response(run, "Clause44_Test")

    async def _collect():
        buf = b""
        async for chunk in resp.body_iterator:
            buf += chunk
        return buf

    return asyncio.run(_collect())


class TestCompanyNameGuard:
    def test_exact_match(self):
        assert _company_names_match("ABC Textile Mills", "ABC Textile Mills") is True

    def test_suffix_variant(self):
        assert _company_names_match(
            "Velav Garments India P Limited",
            "Velav Garments India Private Limited",
        ) is True

    def test_clear_mismatch_blocks(self):
        assert _company_names_match(
            "Velav Garments India P Limited", "ABC Textile Mills"
        ) is False

    def test_empty_books_name_passes(self):
        """If the books JSON doesn't declare a company, don't block the upload."""
        assert _company_names_match("Velav Garments", "") is True

    def test_empty_client_name_passes(self):
        """Bad data in the client record shouldn't block uploads either."""
        assert _company_names_match("", "Anything") is True

    def test_normalise_strips_suffixes_and_punctuation(self):
        # "P Limited", "&" and punctuation are all dropped / collapsed.
        assert _normalise_name("Velav Garments & Co. P. Limited") == "velav garments"

    def test_extract_company_name_legacy_key(self):
        assert _extract_company_name({"company": {"name": "Foo"}}) == "Foo"

    def test_extract_company_name_new_key(self):
        assert _extract_company_name({"company": {"companyName": "Bar"}}) == "Bar"

    def test_extract_company_name_missing(self):
        assert _extract_company_name({"company": {}}) == ""
        assert _extract_company_name({}) == ""


def _sample_run():
    return {
        "run_id": "run_test",
        "company_name": "Test & Co P Ltd",
        "generated_at": "2026-05-04T00:00:00+00:00",
        "summary": {
            "col2_total": 1000.0, "col3": 100.0, "col4": 200.0,
            "col5": 300.0, "col6": 600.0, "col7": 400.0,
        },
        "by_ledger": {
            "Purchases": {"col3": 100.0, "col4": 200.0, "col5": 300.0, "col7": 400.0, "total": 1000.0},
        },
        "recon": {
            "total_books": 1000.0,
            "excluded_lines": [{"name": "Cash", "amount": 50.0}],
            "excluded_total": 50.0,
            "balance": 950.0,
        },
        "transactions": [
            {"bucket": "col3", "date": "2023-04-01", "voucher_type": "Purchase", "voucher_number": "1",
             "ledger_name": "Purchases", "party_name": "A", "party_gstin": "", "party_reg": "registered",
             "amount": 100.0, "reason": "exempt"},
            {"bucket": "col4", "date": "2023-04-02", "voucher_type": "Purchase", "voucher_number": "2",
             "ledger_name": "Purchases", "party_name": "B", "party_gstin": "", "party_reg": "composition",
             "amount": 200.0, "reason": "composition"},
            {"bucket": "col5", "date": "2023-04-03", "voucher_type": "Purchase", "voucher_number": "3",
             "ledger_name": "Purchases", "party_name": "C", "party_gstin": "07X", "party_reg": "registered",
             "amount": 300.0, "reason": "itc voucher"},
            {"bucket": "col7", "date": "2023-04-04", "voucher_type": "Purchase", "voucher_number": "4",
             "ledger_name": "Purchases", "party_name": "", "party_gstin": "", "party_reg": "",
             "amount": 400.0, "reason": "unregistered"},
        ],
    }


class TestExcelExport:
    def test_workbook_has_6_sheets(self):
        body = _read_export_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        assert wb.sheetnames == [
            "Clause 44 Summary",
            "Reconciliation",
            "Col 3 · Exempt",
            "Col 4 · Composition",
            "Col 5 · Other Reg ITC",
            "Col 7 · Unregistered",
        ]

    def test_summary_has_six_col_pivot(self):
        body = _read_export_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Clause 44 Summary"]
        header_values = [ws.cell(row=5, column=c).value for c in range(1, 8)]
        assert header_values[0] == "Particulars"
        assert "Col 2" in header_values[1]
        assert "Col 3" in header_values[2]
        assert "Col 4" in header_values[3]
        assert "Col 5" in header_values[4]
        assert "Col 6" in header_values[5]
        assert "Col 7" in header_values[6]

    def test_cohort_sheets_contain_only_own_bucket(self):
        body = _read_export_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))

        # Column layout after Release 2 — `Amount` sits at column index 8
        # (Date, Voucher Type, Voucher No, Ledger, Party, Party GSTIN,
        # Party Reg., Country, RCM, **Amount**, Value Eligible for ITC,
        # Reason, Auditor Remarks) — but tests/_sample_run has no
        # division so there's no Division column injected.  Find the
        # column dynamically instead of hard-coding.
        for sheet, amt in [
            ("Col 3 · Exempt", 100.0),
            ("Col 4 · Composition", 200.0),
            ("Col 5 · Other Reg ITC", 300.0),
            ("Col 7 · Unregistered", 400.0),
        ]:
            ws = wb[sheet]
            # Header row is row 3.  Locate the "Amount" column.
            header_row = [c.value for c in ws[3]]
            amount_col = header_row.index("Amount") + 1
            data_amt = ws.cell(row=4, column=amount_col).value
            footer_amt = ws.cell(row=5, column=amount_col).value
            assert data_amt == amt, f"Sheet {sheet!r} first row amount = {data_amt}, expected {amt}"
            assert footer_amt == amt, f"Sheet {sheet!r} footer total = {footer_amt}, expected {amt}"
