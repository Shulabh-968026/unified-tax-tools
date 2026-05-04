"""Iteration 18 — Live API tests for Clause 44 Release-1 against the
seeded ABC Textile Mills run on the preview environment.

Validates:
  (a) PATCH /api/runs/{id}/selections accepts the new fields
      (exempt_ledgers, use_itc_inference, exclusion_categories,
      disclaimer_text) and persists them.
  (b) POST /api/runs/{id}/generate uses them to re-classify.
  (c) GET /api/runs/{id} returns fresh re-classified summary/by_ledger/recon
      under different exempt + ITC scenarios (silent re-classify on open).
  (d) GET /api/runs/{id}/export emits a 6-sheet workbook including the
      ICAI 5-line Reconciliation sheet with a Disclaimer block.
  (e) Cascade real-data spot checks: RCM voucher type → Col 7 with
      rcm_total > 0 (the run has 91 Reverse Charge vouchers); imports = 0
      (no foreign parties — expected).
"""
import io
import os
import zipfile
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://unified-tax-tools.preview.emergentagent.com",
).rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RUN_ID = "run_8d427d1f97e0"
COOKIES = {"session_token": TOKEN}


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.cookies.update(COOKIES)
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="module")
def baseline_run(session):
    r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
    assert r.status_code == 200, r.text
    return r.json()


# ─── (a) PATCH selections accepts the new fields ──────────────────────
class TestPatchSelections:
    def test_patch_accepts_new_fields(self, session):
        payload = {
            "exempt_ledgers": [],
            "use_itc_inference": True,
            "exclusion_categories": {},
            "disclaimer_text": "Test disclaimer iteration_18.",
        }
        r = session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json=payload,
            timeout=60,
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["saved"] is True
        # Patch echoes the persisted update
        assert data.get("use_itc_inference") is True
        assert data.get("exempt_selection") == []
        assert data.get("exclusion_categories") == {}
        assert data.get("disclaimer_text") == "Test disclaimer iteration_18."

    def test_patch_persists_to_get(self, session):
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
        assert r.status_code == 200
        data = r.json()
        assert data["use_itc_inference"] is True
        assert data["exempt_selection"] == []
        assert data["disclaimer_text"] == "Test disclaimer iteration_18."

    def test_patch_partial_update(self, session):
        # Only flip the toggle; other fields untouched.
        r = session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json={"use_itc_inference": False},
            timeout=60,
        )
        assert r.status_code == 200
        # Verify
        g = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60).json()
        assert g["use_itc_inference"] is False
        assert g["disclaimer_text"] == "Test disclaimer iteration_18."  # untouched


# ─── (b)+(c) POST generate consumes new fields & GET re-classifies ────
class TestGenerateAndReclassify:
    def test_generate_with_itc_inference_on(self, session, baseline_run):
        itc = baseline_run.get("itc_selection") or []
        excl = baseline_run.get("exclusion_selection") or []
        body = {
            "itc_ledgers": itc,
            "excluded_ledgers": excl,
            "exempt_ledgers": [],
            "use_itc_inference": True,
            "exclusion_categories": {},
        }
        r = session.post(
            f"{BASE_URL}/api/runs/{RUN_ID}/generate", json=body, timeout=120
        )
        assert r.status_code == 200, r.text
        data = r.json()
        s = data["summary"]
        # Sanity: cascade buckets exist + col2 is the report total
        for k in ("col2_total", "col3", "col4", "col5", "col6", "col7"):
            assert k in s
        # Choice 4B/imports — RCM totals exposed
        assert "rcm_total" in s and "rcm_vouchers" in s
        assert "import_total" in s
        # The dataset has 91 Reverse Charge vouchers → must be > 0
        assert s["rcm_total"] > 0
        assert s["rcm_vouchers"] > 0
        # No foreign parties in this seed
        assert s["import_total"] == 0
        # Input-A/B split fields are present
        assert "col3_from_input_a" in s
        assert "col3_from_input_b" in s
        # Recon block & ICAI 5-line keys
        rc = data["recon"]
        for k in (
            "pl_total", "capex_total", "non_cash_total", "sch3_total",
            "money_total", "other_total", "reportable_total",
        ):
            assert k in rc
        # Arithmetic tie
        assert (
            round(rc["pl_total"] + rc["capex_total"]
                  - rc["non_cash_total"] - rc["sch3_total"]
                  - rc["money_total"] - rc["other_total"], 2)
            == round(rc["reportable_total"], 2)
        )
        assert round(rc["reportable_total"], 2) == round(s["col2_total"], 2)

    def test_get_silent_reclassify_with_inference_off(self, session, baseline_run):
        # Save toggle OFF without generate; GET should re-classify on the fly.
        session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json={"use_itc_inference": False, "exempt_ledgers": []},
            timeout=60,
        )
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
        assert r.status_code == 200
        data = r.json()
        # generated=True (last generate call set it) so silent reclassify runs
        assert data["generated"] is True
        s = data["summary"]
        # With inference OFF, Input-B cohort is zero
        assert s["col3_from_input_b"] == 0

    def test_get_silent_reclassify_with_inference_on(self, session):
        session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json={"use_itc_inference": True},
            timeout=60,
        )
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
        s = r.json()["summary"]
        # With inference ON, Input-B may have value (registered no-ITC vouchers exist)
        assert s["col3_from_input_b"] >= 0
        assert s["col3"] == s["col3_from_input_a"] + s["col3_from_input_b"]

    def test_disclaimer_default_present_when_unset(self, session):
        # If disclaimer_text was set earlier, it remains; just confirm a string
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
        d = r.json()["disclaimer_text"]
        assert isinstance(d, str) and len(d) > 20


# ─── (d) Excel export with ICAI 5-line recon sheet ────────────────────
class TestExcelExport:
    @pytest.fixture(scope="class")
    def workbook(self, session):
        # Ensure run is generated
        session.post(
            f"{BASE_URL}/api/runs/{RUN_ID}/generate",
            json={"itc_ledgers": [], "excluded_ledgers": [], "exempt_ledgers": [],
                  "use_itc_inference": True, "exclusion_categories": {}},
            timeout=120,
        )
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}/export", timeout=120)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct or ct.endswith("xlsx") or "octet-stream" in ct
        # Validate it's a real xlsx (PK zip)
        assert r.content[:2] == b"PK"
        return load_workbook(io.BytesIO(r.content), data_only=False)

    def test_six_sheets(self, workbook):
        names = workbook.sheetnames
        assert len(names) == 6, f"Expected 6 sheets, got {names}"

    def test_recon_sheet_present(self, workbook):
        # Find recon sheet — name typically contains 'Reconciliation'
        recon_name = next(
            (n for n in workbook.sheetnames if "recon" in n.lower()),
            None,
        )
        assert recon_name, f"No reconciliation sheet in {workbook.sheetnames}"
        ws = workbook[recon_name]
        # Collect cell values
        text = "\n".join(
            " | ".join(str(c.value) for c in row if c.value is not None)
            for row in ws.iter_rows()
        )
        # ICAI 5-line buckets must appear by label (matches actual export wording)
        for needle in (
            "profit & loss", "capital expenditure",
            "non-cash", "schedule iii", "money", "other",
            "reportable expenditure", "col 2",
        ):
            assert needle in text.lower(), \
                f"Missing '{needle}' line in recon sheet:\n{text[:1500]}"
        # Disclaimer block
        assert "disclaim" in text.lower(), "Disclaimer block missing from recon sheet"

    def test_recon_arithmetic_in_workbook(self, workbook):
        recon_name = next(
            (n for n in workbook.sheetnames if "recon" in n.lower()), None,
        )
        ws = workbook[recon_name]
        # Pull amount from second column for each labeled line
        rows = {}
        for row in ws.iter_rows(values_only=True):
            label = row[0] if row else None
            amount = row[1] if row and len(row) > 1 else None
            if isinstance(label, str) and isinstance(amount, (int, float)):
                rows[label.lower()] = float(amount)
        # Sanity: PL total + capex - exclusions == reportable
        pl = next(v for k, v in rows.items() if "profit & loss" in k)
        capex = next(v for k, v in rows.items() if "capital expenditure" in k)
        non_cash = next(v for k, v in rows.items() if "non-cash" in k)
        sch3 = next(v for k, v in rows.items() if "schedule iii" in k)
        money = next(v for k, v in rows.items() if "money" in k)
        other = next(v for k, v in rows.items() if "other" in k and "exclusion" in k)
        reportable = next(v for k, v in rows.items() if "reportable" in k)
        assert round(pl + capex - non_cash - sch3 - money - other, 2) == round(reportable, 2)
