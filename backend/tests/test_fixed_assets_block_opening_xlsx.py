"""
Fixed Assets — Opening WDV Excel round-trip + optional 3CD validation.

Validates:
  * GET /runs/{rid}/block-opening/export.xlsx returns a valid workbook
    with one row per active block_label (incl. zero-value rows) and a
    hidden canonical-key column.
  * POST /runs/{rid}/block-opening/import.xlsx parses the (edited)
    workbook and upserts fa_block_opening with source="manual_xlsx".
    Footer informational rows are silently skipped.
  * POST /runs/{rid}/block-opening/validate-3cd compares per-rate sums
    against the 3CD JSON closing WDV; returns ok=True when sub-block
    sums tie back, ok=False otherwise. Read-only.
"""
import io
import json
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://unified-tax-tools.preview.emergentagent.com",
).rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RID = "0e4cc62f-52f9-4668-b598-f60bd0c52803"


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.cookies.set("session_token", TOKEN)
    return s


@pytest.fixture(scope="module")
def baseline_xlsx(session):
    """Snapshot the current openings as an .xlsx — restored at end."""
    r = session.get(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/export.xlsx",
        timeout=30,
    )
    assert r.status_code == 200, r.text
    return r.content


@pytest.fixture(autouse=True)
def _restore(session, baseline_xlsx):
    yield
    # Restore the baseline state after every test
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("baseline.xlsx", io.BytesIO(baseline_xlsx),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )


def test_export_returns_valid_xlsx_with_all_blocks(baseline_xlsx):
    assert len(baseline_xlsx) > 3000
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    assert "Opening WDV" in wb.sheetnames
    ws = wb["Opening WDV"]
    # Hidden canonical-key column
    assert ws.column_dimensions["A"].hidden is True
    # 15 active block_labels + ≥4 chrome rows
    block_rows = [row for row in ws.iter_rows(min_row=5, values_only=True)
                  if row and isinstance(row[0], str) and "Block" in (row[0] or "")]
    assert len(block_rows) >= 14, f"expected 14+ block rows, got {len(block_rows)}"


def test_import_round_trip_persists_with_manual_xlsx_source(session, baseline_xlsx):
    """Edit the Vehicles 15% row to ₹450k and re-import. Source flips to manual_xlsx."""
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    ws = wb["Opening WDV"]
    target_row = None
    for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if r[0].value == "15% Block – Vehicles":
            r[3].value = 450000.0
            r[4].value = "Tata tipper opening WDV"
            target_row = r
            break
    assert target_row is not None
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("edited.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    # The footer informational row must NOT be reported as unknown — bug fix
    assert body["unknown_blocks"] == []
    # Re-fetch
    g = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening", timeout=30).json()
    veh = next(r for r in g["rows"] if r["block_label"] == "15% Block – Vehicles")
    assert veh["opening_wdv"] == 450000.0
    assert veh["source"] == "manual_xlsx"
    assert "Tata" in veh["description"]


def test_import_rejects_non_xlsx(session):
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("foo.txt", b"hello", "text/plain")},
        timeout=30,
    )
    assert r.status_code == 400


def test_import_unknown_blocks_returned(session, baseline_xlsx):
    """A row whose canonical key is not an active block must surface in unknown_blocks."""
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    ws = wb["Opening WDV"]
    # Insert a bogus row at the end
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="99% Block – Bogus")
    ws.cell(row=last, column=2, value="99% Block – Bogus")
    ws.cell(row=last, column=3, value=99)
    ws.cell(row=last, column=4, value=12345)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("bogus.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )
    assert r.status_code == 200
    body = r.json()
    assert "99% Block – Bogus" in body["unknown_blocks"]


def test_validate_3cd_match_when_sums_tie(session, baseline_xlsx):
    """Force Vehicles 15% to ₹450k + P&M 15% to ₹25,783,559 → sums to ₹26,233,559
    which should match a 3CD with closing 15% WDV = ₹26,233,559."""
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    ws = wb["Opening WDV"]
    for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if r[0].value == "15% Block – Vehicles":
            r[3].value = 450000.0
        if r[0].value == "15% Block – Plant & Machinery":
            r[3].value = 25783559.0
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("edited.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    ).raise_for_status()

    cd = {
        "FORM3CA": {"F3CA": {"Form3cdDeprAllw": [
            {"RateOfDep": 40, "OpeningWDV": 0, "WrittenDownVal": 970906,
             "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
             "DescBlockAssets": "Computers + P&M"},
            {"RateOfDep": 15, "OpeningWDV": 0, "WrittenDownVal": 26233559,
             "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
             "DescBlockAssets": "P&M + Vehicles"},
            {"RateOfDep": 10, "OpeningWDV": 0, "WrittenDownVal": 2911192,
             "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
             "DescBlockAssets": "Furniture"},
        ]}}
    }
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/validate-3cd",
        files={"file": ("3cd.json", json.dumps(cd).encode(), "application/json")},
        timeout=30,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["ok"] is True
    rate15 = next(row for row in body["rows"] if row["rate"] == 15.0)
    assert rate15["status"] == "match"
    assert rate15["opening_excel"] == 26233559.0
    # Both sub-blocks contributed — what the user explicitly wants visible
    assert set(rate15["blocks"]) == {"15% Block – Plant & Machinery", "15% Block – Vehicles"}


def test_validate_3cd_mismatch_surfaces_drift(session, baseline_xlsx):
    """Vehicles 15% = ₹999,999 → forces a mismatch against ₹26,233,559 3CD."""
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    ws = wb["Opening WDV"]
    for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if r[0].value == "15% Block – Vehicles":
            r[3].value = 999999.0
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("edited.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    ).raise_for_status()
    cd = {"FORM3CA": {"F3CA": {"Form3cdDeprAllw": [
        {"RateOfDep": 15, "OpeningWDV": 0, "WrittenDownVal": 26233559,
         "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
         "DescBlockAssets": ""},
    ]}}}
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/validate-3cd",
        files={"file": ("3cd.json", json.dumps(cd).encode(), "application/json")},
        timeout=30,
    )
    body = r.json()
    assert body["ok"] is False
    rate15 = next(row for row in body["rows"] if row["rate"] == 15.0)
    assert rate15["status"] == "mismatch"
    assert abs(rate15["diff"]) > 1


def test_validate_3cd_rejects_non_3cd_json(session):
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/validate-3cd",
        files={"file": ("foo.json", b'{"not":"3cd"}', "application/json")},
        timeout=30,
    )
    assert r.status_code == 400
