"""Release-2 tests — Excel cohort columns (ICAI Para 79.20) + disclaimer
block, plus foreign-supplier reason enrichment.
"""
import asyncio
import io
import openpyxl
from modules.clause44.service import classify_vouchers
from modules.clause44.exports import build_export_response


def _read_bytes(run):
    resp = build_export_response(run, "Rel2Test")

    async def _collect():
        buf = b""
        async for chunk in resp.body_iterator:
            buf += chunk
        return buf

    return asyncio.run(_collect())


def _sample_run():
    return {
        "run_id": "run_r2",
        "company_name": "Test Co P Ltd",
        "generated_at": "2026-05-04T00:00:00+00:00",
        "disclaimer_text": "Classification based solely on the books JSON; see ICAI Para 79.21.",
        "summary": {
            "col2_total": 6000.0, "col3": 1000.0, "col4": 500.0,
            "col5": 2500.0, "col6": 4000.0, "col7": 2000.0,
            "rcm_total": 2000.0, "rcm_vouchers": 1, "import_total": 1500.0,
            "col3_from_input_a": 600.0, "col3_from_input_b": 400.0,
        },
        "by_ledger": {"Purchase": {"col3": 1000, "col4": 500, "col5": 2500, "col7": 2000, "total": 6000}},
        "recon": {
            "pl_total": 6000.0, "capex_total": 0.0,
            "non_cash_total": 0.0, "sch3_total": 0.0, "money_total": 0.0, "other_total": 0.0,
            "non_cash_lines": [], "sch3_lines": [], "money_lines": [], "other_lines": [], "capex_addback_lines": [],
            "capex_addback_total": 0.0,
            "reportable_total": 6000.0,
            "total_books": 6000.0,
            "excluded_lines": [],
            "excluded_total": 0.0,
            "balance": 6000.0,
        },
        "transactions": [
            # Col 5 · regular w/ ITC — eligible for ITC column = amount
            {"bucket": "col5", "date": "2023-04-01", "voucher_type": "Purchase", "voucher_number": "P1",
             "ledger_name": "Purchase", "party_name": "Alpha", "party_gstin": "27AAAAA0000A1Z5",
             "party_reg": "regular", "party_country": "India",
             "amount": 2500.0, "reason": "Regular vendor, ITC present",
             "is_rcm": False, "is_import": False, "has_itc_ledger": True, "col3_source": ""},
            # Col 7 · RCM — should NOT show ITC eligible value, RCM = Yes
            {"bucket": "col7", "date": "2023-04-02", "voucher_type": "Reverse Charge", "voucher_number": "R1",
             "ledger_name": "Purchase", "party_name": "Beta Legal", "party_gstin": "07BBBBB0000A1Z5",
             "party_reg": "regular", "party_country": "India",
             "amount": 2000.0, "reason": "RCM voucher — supplier typically unregistered",
             "is_rcm": True, "is_import": False, "has_itc_ledger": True, "col3_source": ""},
            # Col 7 · Foreign import — country must show, reason must mention country
            {"bucket": "col7", "date": "2023-04-03", "voucher_type": "Purchase", "voucher_number": "F1",
             "ledger_name": "Purchase", "party_name": "AWS Inc", "party_gstin": "",
             "party_reg": "regular", "party_country": "USA",
             "amount": 1500.0, "reason": "Foreign supplier 'AWS Inc' (Usa) — import, no Indian GSTIN",
             "is_rcm": False, "is_import": True, "has_itc_ledger": False, "col3_source": ""},
            # Col 3 · Input A — NOT eligible for ITC even if ledger has ITC marker
            {"bucket": "col3", "date": "2023-04-04", "voucher_type": "Purchase", "voucher_number": "E1",
             "ledger_name": "Petrol Purchase", "party_name": "HPCL", "party_gstin": "07CCCCC0000A1Z5",
             "party_reg": "regular", "party_country": "India",
             "amount": 600.0, "reason": "Ledger 'Petrol Purchase' tagged as exempt-supply purchase (Input A)",
             "is_rcm": False, "is_import": False, "has_itc_ledger": True, "col3_source": "input_a"},
            # Col 4 · composition — normal
            {"bucket": "col4", "date": "2023-04-05", "voucher_type": "Purchase", "voucher_number": "C1",
             "ledger_name": "Purchase", "party_name": "Gamma Comp", "party_gstin": "24DDDDD0000A1Z5",
             "party_reg": "composition", "party_country": "India",
             "amount": 500.0, "reason": "Party 'Gamma Comp' is Composition dealer",
             "is_rcm": False, "is_import": False, "has_itc_ledger": False, "col3_source": ""},
            # Col 3 · Input B
            {"bucket": "col3", "date": "2023-04-06", "voucher_type": "Purchase", "voucher_number": "B1",
             "ledger_name": "Purchase", "party_name": "Delta Ltd", "party_gstin": "29EEEEE0000A1Z5",
             "party_reg": "regular", "party_country": "India",
             "amount": 400.0, "reason": "Input B — no ITC",
             "is_rcm": False, "is_import": False, "has_itc_ledger": False, "col3_source": "input_b"},
        ],
    }


class TestCohortSheetColumns:
    def test_all_para_79_20_columns_present(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        for s in ["Col 3 · Exempt", "Col 4 · Composition", "Col 5 · Other Reg ITC", "Col 7 · Unregistered"]:
            headers = [c.value for c in wb[s][3]]
            for required in (
                "Date", "Voucher Type", "Voucher No", "Ledger", "Party",
                "Party GSTIN", "Party Reg.", "Country", "RCM", "Amount",
                "Value Eligible for ITC",
                "Reason for NIL GST / Classification Notes",
                "Auditor Remarks",
            ):
                assert required in headers, f"Sheet {s} missing column {required!r}"

    def test_rcm_column_yes_only_for_rcm_voucher(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Col 7 · Unregistered"]
        headers = [c.value for c in ws[3]]
        rcm_col = headers.index("RCM") + 1
        # Col 7 has 2 vouchers — one RCM, one import.  Search their rows.
        rcm_values = []
        for row in ws.iter_rows(min_row=4, values_only=False):
            voucher_no = row[2].value
            if voucher_no in ("R1", "F1"):
                rcm_values.append((voucher_no, row[rcm_col - 1].value))
        rcm_values = dict(rcm_values)
        assert rcm_values["R1"] == "Yes"
        assert rcm_values["F1"] in (None, "")

    def test_value_eligible_for_itc_excludes_rcm_and_input_a(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))

        # Col 5 · 'Alpha' (Regular + ITC, not RCM, not Input A) → eligible = 2500
        ws = wb["Col 5 · Other Reg ITC"]
        headers = [c.value for c in ws[3]]
        amt_col = headers.index("Amount") + 1
        itc_col = headers.index("Value Eligible for ITC") + 1
        assert ws.cell(row=4, column=itc_col).value == 2500.0
        # Footer totals up correctly
        # Footer is row 5 (only 1 data row + 1 footer)
        assert ws.cell(row=5, column=itc_col).value == 2500.0
        assert ws.cell(row=5, column=amt_col).value == 2500.0

        # Col 7 · RCM voucher — ITC eligible should be 0 despite has_itc_ledger=True
        ws7 = wb["Col 7 · Unregistered"]
        headers7 = [c.value for c in ws7[3]]
        itc_col7 = headers7.index("Value Eligible for ITC") + 1
        # Locate RCM row (voucher no R1)
        for row in ws7.iter_rows(min_row=4, values_only=False):
            if row[2].value == "R1":
                assert row[itc_col7 - 1].value == 0.0
                break
        else:
            raise AssertionError("RCM row not found")

        # Col 3 Input A — ITC eligible = 0 even though has_itc_ledger = True
        ws3 = wb["Col 3 · Exempt"]
        headers3 = [c.value for c in ws3[3]]
        itc_col3 = headers3.index("Value Eligible for ITC") + 1
        for row in ws3.iter_rows(min_row=4, values_only=False):
            if row[2].value == "E1":  # Input A voucher
                assert row[itc_col3 - 1].value == 0.0
                break

    def test_auto_filter_set_on_cohort_sheets(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        for s in ["Col 3 · Exempt", "Col 4 · Composition", "Col 5 · Other Reg ITC", "Col 7 · Unregistered"]:
            assert wb[s].auto_filter.ref is not None, f"{s} missing auto-filter"


class TestDisclaimer:
    def test_disclaimer_stamped_into_recon_sheet(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Reconciliation"]
        cell_texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        assert any("ICAI Para 79.21" in t for t in cell_texts)
        assert any(t.startswith("Classification based solely") for t in cell_texts)

    def test_disclaimer_block_header_present(self):
        body = _read_bytes(_sample_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Reconciliation"]
        cell_values = [c.value for row in ws.iter_rows() for c in row]
        assert "Disclaimer" in cell_values


class TestForeignReason:
    def _expenditure(self, *ledgers):
        return {l: {} for l in ledgers}

    def test_reason_includes_country(self):
        forex = {"name": "AWS Inc", "gstRegistrationType": "Regular", "partyGSTIN": "", "country": "USA"}
        v = {
            "voucherId": "v1", "voucherTypeName": "Purchase",
            "date": "2023-04-03", "voucherNumber": "F1",
            "partyLedgerName": "AWS Inc",
            "ledgerEntries": [
                {"ledger": "Cloud Hosting", "amount": -5000, "isPartyLedger": "No"},
                {"ledger": "AWS Inc", "amount": 5000, "isPartyLedger": "Yes"},
            ],
        }
        res = classify_vouchers(
            [v], self._expenditure("Cloud Hosting"), set(), {"AWS Inc": forex},
            use_itc_inference=True,
        )
        reason = res["transactions"][0]["reason"]
        assert "AWS Inc" in reason
        assert "Usa" in reason  # country.title()
        assert "import" in reason.lower()
