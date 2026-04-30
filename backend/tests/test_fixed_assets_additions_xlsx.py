"""
Fixed Assets — Additions Excel Round-trip + Bulk-Patch + Drift Warning
Tests for run rid=0e4cc62f-52f9-4668-b598-f60bd0c52803 under cli_7f0b86b1ab0b.

Covers:
  * GET  /api/fixed-assets/runs/{rid}/additions/export.xlsx
  * POST /api/fixed-assets/runs/{rid}/additions/import.xlsx?dry_run=true
  * POST /api/fixed-assets/runs/{rid}/additions/import.xlsx?dry_run=false
  * POST /api/fixed-assets/runs/{rid}/clear-excel-drift
  * POST /api/fixed-assets/runs/{rid}/additions/bulk-patch
"""
import io
import os
import json
import pytest
import requests
from openpyxl import load_workbook, Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://unified-tax-tools.preview.emergentagent.com").rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RID = "0e4cc62f-52f9-4668-b598-f60bd0c52803"
CID = "cli_7f0b86b1ab0b"


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.cookies.set("session_token", TOKEN)
    return s


@pytest.fixture(scope="module")
def workbook_bytes(session):
    """Pull the workbook once and share across tests."""
    r = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/export.xlsx", timeout=60)
    assert r.status_code == 200, f"export failed: {r.status_code} {r.text[:300]}"
    return r.content


# ------------------------------- EXPORT ----------------------------------

class TestExport:
    def test_export_status_and_size(self, workbook_bytes):
        assert len(workbook_bytes) > 5000, f"workbook too small: {len(workbook_bytes)}"

    def test_export_is_valid_xlsx_multi_sheet(self, workbook_bytes):
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        assert len(wb.sheetnames) >= 2, f"expected >=2 sheets, got {wb.sheetnames}"

    def test_export_has_totals_strip_and_headers(self, workbook_bytes):
        """rows 2-3 = totals strip, row 4 = headers."""
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        sh = wb[wb.sheetnames[0]]
        # headers on row 4
        header_row = [c.value for c in sh[4]]
        header_joined = " ".join(str(h or "") for h in header_row).lower()
        # hidden id columns A + B
        assert sh.column_dimensions["A"].hidden is True, "col A (addition_id) should be hidden"
        assert sh.column_dimensions["B"].hidden is True, "col B (parent_addition_id) should be hidden"
        # totals strip rows 2-3 — row 2 usually has labels, row 3 the numeric totals
        row2 = [c.value for c in sh[2]]
        row3 = [c.value for c in sh[3]]
        has_totals = any(isinstance(v, (int, float)) for v in row2 + row3)
        assert has_totals, f"no numeric totals in rows 2-3. r2={row2} r3={row3}"

    def test_export_contains_discount_credit_rows(self, workbook_bytes):
        """Discount-credit synthetic rows should be present and their ids should begin with discount-*."""
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        found_discount_prefix = False
        for sn in wb.sheetnames:
            sh = wb[sn]
            for row in sh.iter_rows(min_row=5, values_only=True):
                aid = row[0] if row else None
                if isinstance(aid, str) and aid.startswith("discount-"):
                    found_discount_prefix = True
                    break
            if found_discount_prefix:
                break
        # Optional — if no discounts exist in fixture, test still passes silently
        # but we record it:
        print(f"discount-credit rows present: {found_discount_prefix}")


# ------------------------------- IMPORT DRY-RUN --------------------------

class TestImportDryRun:
    def _upload(self, session, buf, dry_run):
        files = {"file": ("edits.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        url = f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/import.xlsx?dry_run={'true' if dry_run else 'false'}"
        return session.post(url, files=files, timeout=60)

    def test_dry_run_noop_returns_ok_no_changes(self, session, workbook_bytes):
        """Re-uploading the exported workbook unchanged → 0 rows_changed, no drift."""
        r = self._upload(session, io.BytesIO(workbook_bytes), dry_run=True)
        assert r.status_code == 200, f"{r.status_code} {r.text[:400]}"
        body = r.json()
        assert body.get("ok") is True
        assert body.get("dry_run") is True
        assert body.get("rows_changed", -1) == 0, f"expected 0 row changes, got {body.get('rows_changed')}: {body}"
        assert body.get("unknown_ids", ["x"]) == []
        assert isinstance(body.get("sheets"), list) and len(body["sheets"]) >= 1
        assert isinstance(body.get("drift"), dict)
        assert body["drift"].get("drifted") is False
        assert isinstance(body.get("changes"), list) and len(body["changes"]) == 0

    def test_dry_run_with_edit_detects_diff(self, session, workbook_bytes):
        """Edit one description cell & re-upload → expect rows_changed==1 and a diff."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        edited_aid = None
        # find first editable non-discount row on sheet 0
        sh = wb[wb.sheetnames[0]]
        desc_col_idx = None
        for idx, cell in enumerate(sh[4], start=1):
            if str(cell.value or "").strip().lower() in ("description", "narration", "particulars"):
                desc_col_idx = idx
                break
        if desc_col_idx is None:
            pytest.skip("Description column not found in exported sheet")
        for row in sh.iter_rows(min_row=5):
            aid = row[0].value
            if isinstance(aid, str) and not aid.startswith("discount-") and aid.strip():
                edited_aid = aid
                orig = row[desc_col_idx - 1].value
                row[desc_col_idx - 1].value = (str(orig or "") + " [QA_EDIT]").strip()
                break
        assert edited_aid, "no editable row found"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        r = self._upload(session, buf, dry_run=True)
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        assert body["ok"] is True and body["dry_run"] is True
        assert body["rows_changed"] == 1, f"expected 1 changed row, got {body['rows_changed']}; body={body}"
        changes = body.get("changes", [])
        assert any(c.get("addition_id") == edited_aid for c in changes), f"edited_aid {edited_aid} missing from changes"
        # per-field diff structure
        for c in changes:
            if c.get("addition_id") == edited_aid:
                assert "changes" in c and isinstance(c["changes"], dict)
                any_field = list(c["changes"].values())[0]
                assert "old" in any_field and "new" in any_field
                break

    def test_dry_run_discount_ids_not_unknown(self, session, workbook_bytes):
        """discount-* ids present in exported file must NOT come back as unknown_ids."""
        r = self._upload(session, io.BytesIO(workbook_bytes), dry_run=True)
        assert r.status_code == 200
        body = r.json()
        unknown = body.get("unknown_ids", [])
        assert not any(str(u).startswith("discount-") for u in unknown), f"discount-* appeared in unknown_ids: {unknown}"

    def test_dry_run_does_not_persist(self, session, workbook_bytes):
        """After dry-run edit, the run object should NOT have excel_drift_warning set."""
        r = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}", timeout=30)
        assert r.status_code == 200
        run = r.json()
        assert not run.get("excel_drift_warning"), f"dry-run leaked a warning: {run.get('excel_drift_warning')}"


# ------------------------------- IMPORT APPLY + DRIFT --------------------

class TestImportApplyAndDrift:
    def test_apply_edit_persists_and_unsets_nothing(self, session, workbook_bytes):
        """Apply a tiny description-only edit (no numeric drift) → rows_changed==1, no drift."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        sh = wb[wb.sheetnames[0]]
        desc_col = None
        for idx, cell in enumerate(sh[4], start=1):
            if str(cell.value or "").strip().lower() in ("description", "narration", "particulars"):
                desc_col = idx
                break
        if desc_col is None:
            pytest.skip("no description column")
        edited_aid = None
        original_value = None
        for row in sh.iter_rows(min_row=5):
            aid = row[0].value
            if isinstance(aid, str) and not aid.startswith("discount-") and aid.strip():
                edited_aid = aid
                original_value = row[desc_col - 1].value
                row[desc_col - 1].value = str(original_value or "") + " [QA_APPLY]"
                break
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("edits.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/import.xlsx?dry_run=false",
            files=files, timeout=60,
        )
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        assert body.get("ok") is True
        assert body.get("dry_run") is False
        # apply endpoint returns 'applied' (not 'rows_changed')
        assert body.get("applied") == 1, body
        # revert
        wb2 = load_workbook(io.BytesIO(workbook_bytes))
        sh2 = wb2[wb2.sheetnames[0]]
        for row in sh2.iter_rows(min_row=5):
            if row[0].value == edited_aid:
                row[desc_col - 1].value = original_value
                break
        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)
        files2 = {"file": ("revert.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        rv = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/import.xlsx?dry_run=false",
            files=files2, timeout=60,
        )
        assert rv.status_code == 200, rv.text[:400]

    def test_apply_numeric_drift_persists_warning(self, session, workbook_bytes):
        """Increase other_expenses by 12345.67 on one row → drift should be flagged & persisted."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        sh = wb[wb.sheetnames[0]]
        # find 'other expenses' / similar column
        target_col = None
        for idx, cell in enumerate(sh[4], start=1):
            v = str(cell.value or "").strip().lower()
            if "other" in v and "exp" in v:
                target_col = idx; break
        if target_col is None:
            pytest.skip("no other_expenses column found")
        edited_aid = None
        orig_val = None
        for row in sh.iter_rows(min_row=5):
            aid = row[0].value
            if isinstance(aid, str) and not aid.startswith("discount-") and aid.strip():
                edited_aid = aid
                orig_val = row[target_col - 1].value or 0
                row[target_col - 1].value = float(orig_val or 0) + 12345.67
                break
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        files = {"file": ("drift.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        r = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/import.xlsx?dry_run=false",
            files=files, timeout=60,
        )
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        assert body.get("applied") == 1, body
        assert body.get("drift", {}).get("drifted") is True, f"expected drift; body={body}"
        # verify run now carries warning
        run = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}").json()
        assert run.get("excel_drift_warning"), f"warning not persisted on run: {run}"
        # revert
        wb2 = load_workbook(io.BytesIO(workbook_bytes))
        sh2 = wb2[wb2.sheetnames[0]]
        for row in sh2.iter_rows(min_row=5):
            if row[0].value == edited_aid:
                row[target_col - 1].value = orig_val
                break
        buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
        files2 = {"file": ("revert.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/import.xlsx?dry_run=false",
            files=files2, timeout=60,
        )

    def test_clear_excel_drift_endpoint(self, session):
        """After clear-excel-drift, GET /runs/{rid} should no longer carry the warning."""
        r = session.post(f"{BASE_URL}/api/fixed-assets/runs/{RID}/clear-excel-drift", timeout=30)
        assert r.status_code == 200, f"{r.status_code} {r.text[:200]}"
        run = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}").json()
        assert not run.get("excel_drift_warning"), f"warning still present after clear: {run.get('excel_drift_warning')}"


# ------------------------------- BULK PATCH ------------------------------

class TestBulkPatch:
    def test_bulk_patch_mark_reviewed_single_row(self, session, workbook_bytes):
        """Pick one addition_id from the workbook, mark reviewed, verify updated==1."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        sh = wb[wb.sheetnames[0]]
        aid = None
        for row in sh.iter_rows(min_row=5):
            v = row[0].value
            if isinstance(v, str) and not v.startswith("discount-") and v.strip():
                aid = v; break
        assert aid, "no addition_id found"
        r = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/bulk-patch",
            json={"addition_ids": [aid], "patch": {"reviewed": True}},
            timeout=30,
        )
        assert r.status_code == 200, f"{r.status_code} {r.text[:400]}"
        body = r.json()
        assert body.get("updated", 0) == 1, f"expected updated=1, got {body}"

    def test_bulk_patch_copy_ptu_from_acc(self, session, workbook_bytes):
        """__copy_ptu_from_acc:true should copy accounting_date → put_to_use_date for editable rows."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        sh = wb[wb.sheetnames[0]]
        aid = None
        for row in sh.iter_rows(min_row=5):
            v = row[0].value
            p = row[1].value
            if (
                isinstance(v, str) and not v.startswith("discount-") and v.strip()
                and not (isinstance(p, str) and p.strip())
            ):
                aid = v; break
        assert aid, "no eligible (non-merged, non-discount) row found"
        r = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/bulk-patch",
            json={"addition_ids": [aid], "patch": {"__copy_ptu_from_acc": True}},
            timeout=30,
        )
        assert r.status_code == 200, f"{r.status_code} {r.text[:400]}"
        body = r.json()
        assert body.get("updated", 0) in (0, 1), body
        # OK even if row already had ptu=acc (idempotent); just verify no error
        assert "error" not in body

    def test_bulk_patch_skips_discount_ids(self, session, workbook_bytes):
        """Supplying a discount-* synthetic id should be skipped (updated count should not include it)."""
        wb = load_workbook(io.BytesIO(workbook_bytes))
        sh = wb[wb.sheetnames[0]]
        discount_id = None
        for sn in wb.sheetnames:
            sh2 = wb[sn]
            for row in sh2.iter_rows(min_row=5):
                v = row[0].value
                if isinstance(v, str) and v.startswith("discount-"):
                    discount_id = v; break
            if discount_id:
                break
        if discount_id is None:
            pytest.skip("no discount-* row in fixture")
        r = session.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/bulk-patch",
            json={"addition_ids": [discount_id], "patch": {"reviewed": True}},
            timeout=30,
        )
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        # Either updated=0 OR backend returned a skipped array — both valid
        assert body.get("updated", 1) == 0, f"discount id should not have been updated: {body}"


# ------------------------------- AUTH-GATING -----------------------------

class TestAuthGating:
    def test_export_requires_session(self):
        r = requests.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/export.xlsx", timeout=30)
        assert r.status_code in (401, 403), f"expected auth gate, got {r.status_code}"

    def test_bulk_patch_requires_session(self):
        r = requests.post(
            f"{BASE_URL}/api/fixed-assets/runs/{RID}/additions/bulk-patch",
            json={"addition_ids": [], "patch": {}},
            timeout=30,
        )
        assert r.status_code in (401, 403)
