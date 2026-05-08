"""Regression test for the Clause 44 Excel export crashing on Tally
voucher narrations / ledger names that contain ASCII control characters
(e.g. \\x1f, \\v, \\r, \\f) — openpyxl rejects these via
`IllegalCharacterError`.  Release 4.4.10 added a `_clean` sanitiser at
the boundary so the export survives gracefully.
"""
import asyncio
import io
import openpyxl

from modules.clause44.exports import build_export_response, _clean, _clean_row


def test_clean_strips_control_characters():
    # Backslash-x escape sequences cover the three banned ranges.
    s = "Vendor\x1fInvoice\x0b#\x0c123\x07"
    assert _clean(s) == "VendorInvoice#123"


def test_clean_passes_normal_strings_unchanged():
    assert _clean("Plain Insurance Premium - Tata AIG") == \
        "Plain Insurance Premium - Tata AIG"


def test_clean_passes_numbers_unchanged():
    assert _clean(125000) == 125000
    assert _clean(125000.50) == 125000.50
    assert _clean(None) is None


def test_clean_truncates_oversize_strings_safely():
    long_s = "X" * 33000
    out = _clean(long_s)
    assert len(out) <= 32_761   # 32760 cap + ellipsis char


def test_clean_row_handles_mixed_types():
    row = ["Vendor\x1fName", 100.0, None, "OK"]
    assert _clean_row(row) == ["VendorName", 100.0, None, "OK"]


# ─────────────────────────────────────────────────────────────────────────
# End-to-end — build_export_response must not raise on dirty narrations.
# ─────────────────────────────────────────────────────────────────────────
def _drain(resp) -> bytes:
    async def _go():
        return b"".join([c async for c in resp.body_iterator])
    return asyncio.new_event_loop().run_until_complete(_go())


def _dirty_run():
    """Synthetic run with control chars in cohort + col8 narrations and
    ledger names — mirrors what real Tally exports occasionally produce
    when narrations are clipboard-pasted from Word/PDF sources."""
    return {
        "run_id": "dirty-run-1",
        "company_name": "ACME Pvt Ltd",
        "client_name": "ACME Pvt Ltd",
        "period": "2024-25",
        "division_name": None,
        "generated": True,
        "by_ledger": {
            # Ledger name with a stray \x1f from a Tally export.
            "Repairs\x1f& Maintenance": {
                "total": 100000, "col3": 0, "col4": 0,
                "col5": 100000, "col6": 100000, "col7": 0, "col8": 0,
            },
        },
        "summary": {
            "col2_total": 100000, "col3": 0, "col4": 0,
            "col5": 100000, "col7": 0, "col8": 0,
            "reportable_total": 100000,
        },
        "transactions": [
            {
                "bucket": "col5",
                "date": "2024-04-01",
                "voucher_type": "Purchase",
                "voucher_number": "PUR/001\x07",
                "ledger_name": "Repairs\x1f& Maintenance",
                "party_name": "Vendor\x0bA",
                "party_gstin": "24AAAAA0000A1Z5",
                "party_reg": "Regular",
                "party_country": "IN",
                "is_rcm": False,
                "amount": 100000,
                "has_itc_ledger": True,
                "col3_source": None,
                "reason": "Picked from voucher\x1fline\x0c1 — see narration",
            },
        ],
        "recon": {
            "pl_total": 100000,
            "capex_total": 0,
            "reportable_total": 50000,
            "sch3_total": 50000,
            "sch3_lines": [
                {"name": "Salary\x1fJan", "amount": 50000},
            ],
        },
    }


def test_export_response_survives_control_chars_in_voucher_data():
    resp = build_export_response(_dirty_run(), "Clause44_dirty_run")
    raw = _drain(resp)
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # Summary tab — ledger name was sanitised.
    summary = wb["Clause 44 Summary"]
    rows = list(summary.iter_rows(values_only=True))
    flat = [c for r in rows for c in r if isinstance(c, str)]
    # No control chars survive in any cell.
    for s in flat:
        assert "\x1f" not in s and "\x0b" not in s and "\x0c" not in s and "\x07" not in s, \
            f"Control char leaked into cell: {repr(s)}"

    # Cohort sheet (Col 5) — voucher number, party name, reason all clean.
    col5 = wb["Col 5 · Other Reg ITC"]
    rows = list(col5.iter_rows(values_only=True))
    flat = [c for r in rows for c in r if isinstance(c, str)]
    for s in flat:
        assert all(ch not in s for ch in ("\x1f", "\x0b", "\x0c", "\x07"))

    # Reconciliation tab — `less_lines[0].name` was dirty too.
    recon = wb["Reconciliation"]
    rows = list(recon.iter_rows(values_only=True))
    flat = [c for r in rows for c in r if isinstance(c, str)]
    assert any("Salary" in c and "Jan" in c for c in flat), \
        "Sanitised salary line missing from Reconciliation"
    for s in flat:
        assert "\x1f" not in s
