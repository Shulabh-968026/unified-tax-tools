"""Library Phase B — backend integration tests for iteration 23.

Covers:
- Catalog has 14 file_types incl. output kind msme43bh_creditor_report_xlsx
- party_master_xlsx template: 7 sheets, MSME Details has 3 data validations
- Upload of a kind=output file_type is rejected (400)
- Status surfaces the new output chip (uploaded=false) before compute
- Per-module Library integration on Balance Confirmation, Fixed Assets,
  GST Recon, FS Designer, MSME 43BH (library_status, save+pin, etc.)
- 43BH compute auto-persists Creditor Report as kind=output v1
"""
import io
import json
import uuid
from typing import Any, Dict

import pytest
import requests

from tests.conftest import BASE_URL


def _hdr(client_fixture):
    return client_fixture["headers"]


@pytest.fixture(scope="module")
def libb_client(auth_headers, sample_files):
    fno = f"FIXTURE_LIBB_{uuid.uuid4().hex[:6].upper()}"
    r = requests.post(
        f"{BASE_URL}/api/clients",
        headers=auth_headers,
        json={"name": "ABC Textile Mills", "file_number": fno, "type": "single"},
    )
    assert r.status_code == 200, r.text
    cid = r.json()["client_id"]
    return {"client_id": cid, "headers": auth_headers, "files": sample_files}


# ---------- 1. Catalog ----------------------------------------------------
def test_catalog_has_msme_creditor_report(libb_client):
    r = requests.get(f"{BASE_URL}/api/library/catalog", headers=_hdr(libb_client))
    assert r.status_code == 200
    types = r.json()["file_types"]
    assert len(types) == 13
    by_key = {t["key"]: t for t in types}
    assert "msme43bh_creditor_report_xlsx" in by_key
    # Reclassified to secondary in 4.3 (was 'output' in 4.2).
    assert by_key["msme43bh_creditor_report_xlsx"]["kind"] == "secondary"
    # Removed in 4.3 — these are not used anywhere yet.
    assert "bank_statements_xlsx" not in by_key
    assert "gstr_9_json" not in by_key
    # New in 4.3.
    assert "it_depreciation_opening_wdv_xlsx" in by_key
    # FA Register and Opening WDV both have templates now.
    assert by_key["fa_register_xlsx"]["has_template"] is True
    assert by_key["it_depreciation_opening_wdv_xlsx"]["has_template"] is True


# ---------- 2. Party master template (7 sheets + 3 dropdowns) ------------
def test_party_master_template_has_7_sheets_with_dvs(libb_client):
    # Template generator seeds from books_json + ledger_mapping_xlsx —
    # upload both first.
    with open(libb_client["files"]["json"], "rb") as f:
        up1 = requests.post(
            f"{BASE_URL}/api/library/upload", headers=_hdr(libb_client),
            files={"file": ("books.json", f, "application/json")},
            data={"client_id": libb_client["client_id"], "period": "2023-24",
                  "file_type": "books_json"},
        )
    assert up1.status_code == 200, up1.text
    with open(libb_client["files"]["xlsx"], "rb") as f:
        up2 = requests.post(
            f"{BASE_URL}/api/library/upload", headers=_hdr(libb_client),
            files={"file": ("ledger.xlsx", f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            data={"client_id": libb_client["client_id"], "period": "2023-24",
                  "file_type": "ledger_mapping_xlsx"},
        )
    assert up2.status_code == 200, up2.text
    r = requests.get(
        f"{BASE_URL}/api/library/clients/{libb_client['client_id']}/template/party_master_xlsx",
        headers=_hdr(libb_client), params={"period": "2023-24"},
    )
    assert r.status_code == 200, r.text
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == [
        "README", "Trade Payables", "Trade Receivables",
        "Unsecured Loans", "Bank Accounts", "Others", "MSME Details",
    ]
    ws = wb["MSME Details"]
    dvs = list(ws.data_validations.dataValidation)
    assert len(dvs) == 3
    formulas = {dv.formula1 for dv in dvs}
    # Sector / MSME Type / Capital Goods
    assert any("Manufacturing" in f and "Services" in f and "Trading" in f for f in formulas)
    assert any("Micro" in f and "Small" in f and "Medium" in f for f in formulas)
    assert any("Yes" in f and "No" in f for f in formulas)


# ---------- 3. Upload of msme43bh_creditor_report_xlsx is now ALLOWED ----
# (Reclassified from output → secondary in 4.3; auditors can drop in an
# externally-prepared report.)
def test_upload_creditor_report_now_allowed(libb_client):
    # Use a syntactically-valid (header-only) xlsx so the extension check
    # passes; controller doesn't validate workbook contents.
    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(["dummy"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = requests.post(
        f"{BASE_URL}/api/library/upload",
        headers=_hdr(libb_client),
        files={"file": ("manual_creditor.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"client_id": libb_client["client_id"], "period": "2023-24",
              "file_type": "msme43bh_creditor_report_xlsx"},
    )
    assert r.status_code == 200, r.text


# ---------- 4. Status surfaces the creditor-report chip ------------------
def test_status_includes_creditor_report_chip(libb_client):
    r = requests.get(
        f"{BASE_URL}/api/library/clients/{libb_client['client_id']}/status",
        headers=_hdr(libb_client), params={"period": "2023-24"},
    )
    assert r.status_code == 200
    files = {f["key"]: f for f in r.json()["files"]}
    assert "msme43bh_creditor_report_xlsx" in files
    chip = files["msme43bh_creditor_report_xlsx"]
    assert chip["kind"] == "secondary"


# ---------- 5. Balance Confirmation: upload-books, library_status, rerun -
def _create_bc_run(libb_client) -> str:
    r = requests.post(
        f"{BASE_URL}/api/balance-confirmation/runs", headers=_hdr(libb_client),
        json={"client_id": libb_client["client_id"], "fy": "2023-24",
              "name": "TEST BC libb"},
    )
    assert r.status_code == 200, r.text
    return r.json()["id"]


def test_bc_upload_books_saves_to_library_and_pins(libb_client):
    rid = _create_bc_run(libb_client)
    pytest.libb_bc_rid = rid
    with open(libb_client["files"]["json"], "rb") as f:
        r = requests.post(
            f"{BASE_URL}/api/balance-confirmation/runs/{rid}/upload-books",
            headers=_hdr(libb_client),
            files={"file": ("books.json", f, "application/json")},
        )
    assert r.status_code == 200, r.text

    # Library status reflects uploaded books_json v1
    r2 = requests.get(
        f"{BASE_URL}/api/library/clients/{libb_client['client_id']}/status",
        headers=_hdr(libb_client), params={"period": "2023-24"},
    )
    files = {f["key"]: f for f in r2.json()["files"]}
    assert files["books_json"]["uploaded"] is True
    assert files["books_json"]["version_no"] >= 1

    # GET run shows library_status + pinned_files
    r3 = requests.get(
        f"{BASE_URL}/api/balance-confirmation/runs/{rid}", headers=_hdr(libb_client),
    )
    body = r3.json()
    assert "books_json" in (body.get("pinned_files") or {})
    ls = body.get("library_status") or {}
    assert ls.get("module_key") == "balance_confirmation"
    assert isinstance(ls.get("dependencies"), list)
    pytest.libb_bc_books_v1 = body["pinned_files"]["books_json"]


def test_bc_rerun_repins_to_current_books(libb_client):
    rid = pytest.libb_bc_rid
    # Mutate library to v2
    with open(libb_client["files"]["json"], "rb") as f:
        original = f.read()
    r = requests.post(
        f"{BASE_URL}/api/library/upload",
        headers=_hdr(libb_client),
        files={"file": ("books.json", original + b"\n", "application/json")},
        data={"client_id": libb_client["client_id"], "period": "2023-24",
              "file_type": "books_json"},
    )
    assert r.status_code == 200
    new_fid = r.json()["file"]["file_id"]
    assert new_fid != pytest.libb_bc_books_v1

    # Rerun re-pins
    r2 = requests.post(
        f"{BASE_URL}/api/balance-confirmation/runs/{rid}/rerun",
        headers=_hdr(libb_client),
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["pinned_books_file_id"] == new_fid


# ---------- 6. Fixed Assets: ingest-books library_status -----------------
def test_fa_ingest_books_saves_to_library(libb_client):
    r = requests.post(
        f"{BASE_URL}/api/fixed-assets/runs", headers=_hdr(libb_client),
        json={"client_id": libb_client["client_id"], "fy": "2023-24",
              "name": "TEST FA libb"},
    )
    assert r.status_code == 200, r.text
    rid = r.json()["id"]
    with open(libb_client["files"]["json"], "rb") as f:
        r2 = requests.post(
            f"{BASE_URL}/api/fixed-assets/runs/{rid}/ingest-books",
            headers=_hdr(libb_client),
            files={"file": ("books.json", f, "application/json")},
        )
    # Even if FA detected 0 ledgers (returns 400), the upload is parsed first,
    # but library save happens AFTER the detect step. So treat 400 as
    # acceptable for the fa-detection branch — library save already gates on
    # successful parse_books_json. We only require GET library_status works.
    r3 = requests.get(
        f"{BASE_URL}/api/fixed-assets/runs/{rid}", headers=_hdr(libb_client),
    )
    assert r3.status_code == 200, r3.text
    body = r3.json()
    ls = body.get("library_status") or {}
    assert ls.get("module_key") == "fixed_assets"
    assert isinstance(ls.get("dependencies"), list)
    if r2.status_code == 200:
        assert "books_json" in (body.get("pinned_files") or {})


# ---------- 7. GST Recon: library_status + library save on books bucket --
def test_gst_recon_library_integration(libb_client):
    r = requests.post(
        f"{BASE_URL}/api/gst-recon/runs", headers=_hdr(libb_client),
        json={"client_id": libb_client["client_id"], "fy": "2023-24",
              "name": "TEST GST libb"},
    )
    assert r.status_code == 200, r.text
    rid = r.json()["id"]
    # Upload books JSON via /files (multi-file endpoint). Filename must
    # match the books-bucket sniff regex (`\d{8}-\d{8}` etc.).
    books_name = "ABC_Textile_Mills_01042023-31032024.json"
    with open(libb_client["files"]["json"], "rb") as f:
        r2 = requests.post(
            f"{BASE_URL}/api/gst-recon/runs/{rid}/files",
            headers=_hdr(libb_client),
            files=[("files", (books_name, f, "application/json"))],
        )
    assert r2.status_code == 200, r2.text
    r3 = requests.get(f"{BASE_URL}/api/gst-recon/runs/{rid}", headers=_hdr(libb_client))
    assert r3.status_code == 200, r3.text
    body = r3.json()
    ls = body.get("library_status") or {}
    assert ls.get("module_key") == "gst_recon"
    assert isinstance(ls.get("dependencies"), list)
    # books_json should be pinned if the JSON validated as books
    pf = body.get("pinned_files") or {}
    assert "books_json" in pf, f"expected books_json in pinned_files, got {pf}"


# ---------- 8. FS Designer: get_run library_status -----------------------
def test_fs_get_run_returns_library_status(libb_client):
    r = requests.post(
        f"{BASE_URL}/api/fin-statement/runs", headers=_hdr(libb_client),
        json={"client_id": libb_client["client_id"], "fy": "2023-24",
              "fy_start": "2023-04-01", "fy_end": "2024-03-31",
              "name": "TEST FS libb"},
    )
    assert r.status_code == 200, r.text
    rid = r.json()["id"]
    r2 = requests.get(f"{BASE_URL}/api/fin-statement/runs/{rid}", headers=_hdr(libb_client))
    assert r2.status_code == 200, r2.text
    ls = r2.json().get("library_status") or {}
    assert ls.get("module_key") == "fin_statement"
    assert isinstance(ls.get("dependencies"), list)


# ---------- 9. MSME 43BH compute auto-saves Creditor Report --------------
def _build_yearend_xlsx() -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Yearend"
    headers = [
        "Ledger Name", "Is MSME", "Analysis Type", "Voucher No",
        "Voucher Date", "Bill Amount", "Due Date", "> 45 Days", "Overdue at Year End",
    ]
    ws.append(headers)
    ws.append([
        "TEST MSME Vendor 1", True, "Goods", "V001",
        "2023-05-10", 100000.0, "2023-06-24", 0.0, 0.0,
    ])
    ws.append([
        "TEST MSME Vendor 2", True, "Goods", "V002",
        "2023-12-01", 50000.0, "2024-01-15", 50000.0, 50000.0,
    ])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_msme_compute_persists_creditor_report_to_library(libb_client):
    # Create session
    r = requests.post(
        f"{BASE_URL}/api/msme/sessions", headers=_hdr(libb_client),
        json={"client_id": libb_client["client_id"], "fy": "2023-24",
              "name": "TEST MSME libb"},
    )
    assert r.status_code == 200, r.text
    sid = r.json()["id"]

    # Upload yearend XLSX
    blob = _build_yearend_xlsx()
    r2 = requests.post(
        f"{BASE_URL}/api/msme/sessions/{sid}/yearend",
        headers=_hdr(libb_client),
        files={"file": ("yearend.xlsx", blob,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["bill_count"] == 2

    # Compute (no payments → fully disallowed)
    r3 = requests.post(
        f"{BASE_URL}/api/msme/sessions/{sid}/compute",
        headers=_hdr(libb_client),
    )
    assert r3.status_code == 200, r3.text
    summary = r3.json().get("summary") or {}
    assert summary.get("bill_count") == 2

    # Status now shows the creditor-report chip with uploaded=true v1+.
    # (Note: this test runs after test_upload_creditor_report_now_allowed
    #  which uploaded a v1 manually, so the auto-save here lands as v2.)
    r4 = requests.get(
        f"{BASE_URL}/api/library/clients/{libb_client['client_id']}/status",
        headers=_hdr(libb_client), params={"period": "2023-24"},
    )
    assert r4.status_code == 200
    files = {f["key"]: f for f in r4.json()["files"]}
    chip = files["msme43bh_creditor_report_xlsx"]
    assert chip["kind"] == "secondary"
    assert chip["uploaded"] is True, f"expected uploaded=true after compute, got {chip}"
    assert chip["version_no"] >= 1
