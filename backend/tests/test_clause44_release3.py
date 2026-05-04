"""Release-3 tests — Col 8 (Excluded) as a first-class classification
bucket + cascade Step 0 + 7-column summary pivot."""
import asyncio
import io
import openpyxl
from modules.clause44.service import classify_vouchers
from modules.clause44.exports import build_export_response


def _read_bytes(run):
    resp = build_export_response(run, "Rel3Test")

    async def _collect():
        buf = b""
        async for chunk in resp.body_iterator:
            buf += chunk
        return buf

    return asyncio.run(_collect())


def _party(name, reg, gstin="", country="India"):
    return {"name": name, "gstRegistrationType": reg, "partyGSTIN": gstin, "country": country}


def _v(vid, party, entries, vtype="Purchase"):
    return {
        "voucherId": vid, "voucherTypeName": vtype, "date": "2023-04-01",
        "voucherNumber": vid, "partyLedgerName": party["name"] if party else "",
        "ledgerEntries": entries,
    }


def _exp(*l):
    return {x: {} for x in l}


class TestStep0ExclusionWins:
    def test_excluded_ledger_routes_to_col8_over_everything(self):
        acme = _party("Acme Ltd", "Regular", "07AAAAA0000A1Z5")
        # Even though Acme is a registered vendor with ITC ledger, the
        # ledger itself is on the excluded list — must land in Col 8.
        v = _v("v1", acme, [
            {"ledger": "Depreciation", "amount": -5000},
            {"ledger": "Input IGST",   "amount": -900},
            {"ledger": "Acme Ltd",     "amount": 5900},
        ])
        out = classify_vouchers(
            [v], _exp("Depreciation"), {"Input IGST"}, {"Acme Ltd": acme},
            exempt_ledgers=set(), excluded_ledgers={"Depreciation"},
            use_itc_inference=True,
        )
        assert out["summary"]["col8"] == 5000
        assert out["summary"]["col5"] == 0
        assert out["by_ledger"]["Depreciation"]["col8"] == 5000

    def test_excluded_wins_over_input_a(self):
        # If a ledger is in BOTH sets, Step 0 (exclude) must beat Step 2.
        acme = _party("Acme", "Regular", "X")
        v = _v("v1", acme, [
            {"ledger": "Petrol Purchase", "amount": -1000},
            {"ledger": "Acme",            "amount": 1000},
        ])
        out = classify_vouchers(
            [v], _exp("Petrol Purchase"), set(), {"Acme": acme},
            exempt_ledgers={"Petrol Purchase"},
            excluded_ledgers={"Petrol Purchase"},
            use_itc_inference=True,
        )
        assert out["summary"]["col8"] == 1000
        assert out["summary"]["col3"] == 0

    def test_col2_is_gross_sum_of_cols_3_to_8(self):
        acme = _party("Acme", "Regular", "X")
        chai = _party("Chai Stall", "Unregistered", "")
        vs = [
            _v("v1", acme, [
                {"ledger": "Purchase", "amount": -1000},
                {"ledger": "Input IGST", "amount": -180},
                {"ledger": "Acme",       "amount": 1180},
            ]),
            _v("v2", chai, [
                {"ledger": "Refreshment", "amount": -100},
                {"ledger": "Chai Stall",  "amount": 100},
            ]),
            _v("v3", acme, [
                {"ledger": "Depreciation", "amount": -500},
                {"ledger": "Acme",         "amount": 500},
            ]),
        ]
        out = classify_vouchers(
            vs, _exp("Purchase", "Refreshment", "Depreciation"), {"Input IGST"},
            {"Acme": acme, "Chai Stall": chai},
            excluded_ledgers={"Depreciation"},
            use_itc_inference=True,
        )
        s = out["summary"]
        # Col 2 (gross) = 1000 + 100 + 500 = 1600
        assert s["col2_total"] == 1600
        assert s["col2_total"] == s["col3"] + s["col4"] + s["col5"] + s["col7"] + s["col8"]
        assert s["col5"] == 1000  # Acme taxable
        assert s["col7"] == 100   # Chai Stall URD
        assert s["col8"] == 500   # Depreciation excluded
        # Reportable = Col 6 + Col 7 (excludes Col 8)
        assert s["reportable_total"] == s["col6"] + s["col7"]
        assert s["reportable_total"] == 1100


class TestSevenColSummarySheet:
    def _mini_run(self):
        return {
            "run_id": "r", "company_name": "X",
            "generated_at": "", "disclaimer_text": "",
            "summary": {
                "col2_total": 1600, "col3": 0, "col4": 0, "col5": 1000,
                "col6": 1000, "col7": 100, "col8": 500, "reportable_total": 1100,
            },
            "by_ledger": {
                "Purchase":     {"col3": 0, "col4": 0, "col5": 1000, "col7": 0,   "col8": 0,   "total": 1000},
                "Refreshment":  {"col3": 0, "col4": 0, "col5": 0,    "col7": 100, "col8": 0,   "total": 100},
                "Depreciation": {"col3": 0, "col4": 0, "col5": 0,    "col7": 0,   "col8": 500, "total": 500},
            },
            "recon": {
                "pl_total": 1600, "capex_total": 0,
                "non_cash_total": 500, "sch3_total": 0, "money_total": 0, "other_total": 0,
                "non_cash_lines": [{"name": "Depreciation", "amount": 500, "bucket": "non_cash"}],
                "sch3_lines": [], "money_lines": [], "other_lines": [], "capex_addback_lines": [],
                "capex_addback_total": 0, "col8_total": 500,
                "reportable_total": 1100,
                "total_books": 1600,
                "excluded_lines": [{"name": "Depreciation", "amount": 500, "bucket": "non_cash"}],
                "excluded_total": 500, "balance": 1100,
            },
            "transactions": [
                {"bucket": "col5", "date": "2023-04-01", "voucher_type": "Purchase", "voucher_number": "v1",
                 "ledger_name": "Purchase", "party_name": "Acme", "party_gstin": "X", "party_reg": "regular",
                 "party_country": "India", "amount": 1000, "reason": "Regular", "is_rcm": False,
                 "is_import": False, "has_itc_ledger": True, "col3_source": ""},
                {"bucket": "col7", "date": "2023-04-01", "voucher_type": "Purchase", "voucher_number": "v2",
                 "ledger_name": "Refreshment", "party_name": "Chai Stall", "party_gstin": "",
                 "party_reg": "unregistered", "party_country": "India", "amount": 100, "reason": "URD",
                 "is_rcm": False, "is_import": False, "has_itc_ledger": False, "col3_source": ""},
                {"bucket": "col8", "date": "2023-04-01", "voucher_type": "Purchase", "voucher_number": "v3",
                 "ledger_name": "Depreciation", "party_name": "Acme", "party_gstin": "X",
                 "party_reg": "regular", "party_country": "India", "amount": 500,
                 "reason": "Ledger 'Depreciation' excluded from Col 3-7 reporting (Col 8)",
                 "is_rcm": False, "is_import": False, "has_itc_ledger": False, "col3_source": ""},
            ],
        }

    def test_summary_sheet_carries_col8_column(self):
        body = _read_bytes(self._mini_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Clause 44 Summary"]
        # The aggregate header row (row 5) should have 8 columns.
        hdrs = [ws.cell(row=5, column=c).value for c in range(1, 9)]
        assert "Col 8" in hdrs[7]
        # Aggregate row row 6 — Col 8 value = 500
        assert ws.cell(row=6, column=8).value == 500

    def test_col8_sheet_present_with_subbucket_grouping(self):
        body = _read_bytes(self._mini_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        assert "Col 8 · Excluded" in wb.sheetnames
        ws = wb["Col 8 · Excluded"]
        # Find the "Non-cash charges" sub-bucket band header anywhere
        all_vals = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        assert any("Non-cash charges" in v for v in all_vals)
        # Col 8 total should be stamped at the bottom
        assert any("Col 8 Total" in v for v in all_vals)

    def test_col8_sheet_shows_voucher_detail(self):
        body = _read_bytes(self._mini_run())
        wb = openpyxl.load_workbook(io.BytesIO(body))
        ws = wb["Col 8 · Excluded"]
        # Find a row with the Depreciation ledger name.
        found = False
        for row in ws.iter_rows(values_only=True):
            if row and "Depreciation" in (row or ()):
                found = True
                break
        assert found, "Depreciation voucher row not found in Col 8 sheet"


class TestDisclaimerContent:
    def test_default_disclaimer_matches_user_approved_text(self):
        from modules.clause44.controller import DEFAULT_DISCLAIMER
        # Key phrases the user dictated for the working-paper footer.
        for phrase in [
            "classification of expenditure under Clause 44",
            "based solely on the books of account",
            "purchase ledgers internally designated by the entity as exempt-supply ledgers",
            "absence of an ITC-input ledger",
            "both as adopted and affirmed by management",
            "RCM vouchers and purchases from foreign suppliers are reported under Column 7",
            "true and complete extract of its books",
            "(Ref: ICAI Guidance Note on Tax Audit, Para 79.20 / 79.21.)",
        ]:
            assert phrase in DEFAULT_DISCLAIMER, f"Missing phrase: {phrase!r}"
