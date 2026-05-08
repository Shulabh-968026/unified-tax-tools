"""Smoke test for the Clause 44 Mapping Snapshot Excel builder."""
import asyncio
import io
import openpyxl

from modules.clause44.exports import build_mapping_export_response


def _sample_run():
    return {
        "run_id": "run-test-1",
        "client_name": "ACME Pvt Ltd",
        "company_name": "ACME Pvt Ltd",
        "period": "2024-25",
        "division_name": "Mumbai",
        "snapshot_at": "2026-02-01T00:00:00+00:00",
        "exempt_ledgers": [
            {"name": "Petrol Expenses", "subhead": "Misc Exp", "group_parent": "Indirect Exp",
             "head": "Other Expenses", "closing_balance": -125000.0, "suggested": True},
            {"name": "Office Rent",     "subhead": "Rent",     "group_parent": "Indirect Exp",
             "head": "Other Expenses", "closing_balance": -240000.0, "suggested": False},
        ],
        "itc_ledgers_all_bs": [
            {"name": "Input CGST", "subhead": "Balance with Revenue Authorities",
             "group_parent": "Duties & Taxes", "head": "Current Assets",
             "closing_balance": 35000.0, "kind": "input", "kind_source": "name",
             "n_purchase": 12, "n_sales": 0, "usage_conflict": False,
             "in_default_view": True, "suggested": True},
            {"name": "Output IGST", "subhead": "Statutory Dues Payable",
             "group_parent": "Duties & Taxes", "head": "Current Liabilities",
             "closing_balance": -55000.0, "kind": "output", "kind_source": "name",
             "n_purchase": 0, "n_sales": 18, "usage_conflict": False,
             "in_default_view": True, "suggested": False},
        ],
        "exclusion_ledgers": [
            {"name": "Depreciation", "subhead": "Depreciation", "group_parent": "Indirect Exp",
             "head": "Other Expenses", "closing_balance": -800000.0,
             "recon_role": "subtract", "suggested": True},
            {"name": "Plant & Machinery - Additions", "subhead": "Plant",
             "group_parent": "Fixed Assets", "head": "Property, Plant and Equipment",
             "closing_balance": 1500000.0, "recon_role": "addback", "suggested": False},
        ],
        "itc_selection": ["Input CGST"],
        "exempt_selection": ["Petrol Expenses"],
        "exclusion_selection": ["Depreciation"],
        "exclusion_categories": {"Depreciation": "non_cash"},
    }


def _collect_streaming_response(resp) -> bytes:
    async def _drain():
        chunks = []
        async for c in resp.body_iterator:
            chunks.append(c)
        return b"".join(chunks)
    return asyncio.new_event_loop().run_until_complete(_drain())


def _load_workbook_from_response(resp) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(_collect_streaming_response(resp)), data_only=True)


def _body_rows(ws):
    """Return all data rows beneath the 'Ledger Name' header, keyed by name."""
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r and r[0] == "Ledger Name")
    return {r[0]: r for r in rows[hdr_idx + 1:] if r and r[0]}


def test_mapping_snapshot_three_sheets():
    resp = build_mapping_export_response(_sample_run(), "Clause44_Mapping_test")
    wb = _load_workbook_from_response(resp)
    assert wb.sheetnames == ["Exempt Purchases", "ITC Ledgers", "Exclusions"]


def test_mapping_snapshot_exempt_pretick_and_selected():
    resp = build_mapping_export_response(_sample_run(), "Clause44_Mapping_test")
    wb = _load_workbook_from_response(resp)
    by_name = _body_rows(wb["Exempt Purchases"])
    # Column layout (Release 4.4.8): Name, Subhead, Group, Head, CB,
    # Vouchers, ITC-Overlap Vouchers, Demoted by ITC Cross-Check?,
    # Auto-Suggested?, Currently Selected?
    assert by_name["Petrol Expenses"][8] == "Yes"   # Auto-Suggested?
    assert by_name["Petrol Expenses"][9] == "Yes"   # Currently Selected?
    assert by_name["Office Rent"][8] == "No"
    assert by_name["Office Rent"][9] == "No"


def test_mapping_snapshot_itc_kind_and_usage_columns():
    resp = build_mapping_export_response(_sample_run(), "Clause44_Mapping_test")
    wb = _load_workbook_from_response(resp)
    by_name = _body_rows(wb["ITC Ledgers"])
    # Kind chip + Auto-Suggested + Currently Selected
    assert by_name["Input CGST"][5] == "input"
    assert by_name["Input CGST"][7] == 12        # purchase voucher count
    assert by_name["Input CGST"][11] == "Yes"    # Auto-Suggested?
    assert by_name["Input CGST"][12] == "Yes"    # Currently Selected?
    assert by_name["Output IGST"][5] == "output"
    assert by_name["Output IGST"][8] == 18       # sales voucher count
    assert by_name["Output IGST"][11] == "No"
    assert by_name["Output IGST"][12] == "No"


def test_mapping_snapshot_exclusions_recon_role_and_bucket():
    resp = build_mapping_export_response(_sample_run(), "Clause44_Mapping_test")
    wb = _load_workbook_from_response(resp)
    by_name = _body_rows(wb["Exclusions"])
    assert by_name["Depreciation"][5] == "subtract"
    assert by_name["Depreciation"][6] == "Yes"
    assert by_name["Depreciation"][7] == "Yes"
    assert by_name["Depreciation"][8] == "non_cash"
    assert by_name["Plant & Machinery - Additions"][5] == "addback"
    assert by_name["Plant & Machinery - Additions"][6] == "No"
    assert by_name["Plant & Machinery - Additions"][7] == "No"
    # No category override → "—"
    assert by_name["Plant & Machinery - Additions"][8] == "—"


def test_mapping_snapshot_empty_pools_render_placeholder():
    run = _sample_run()
    run["exempt_ledgers"] = []
    run["itc_ledgers_all_bs"] = []
    run["exclusion_ledgers"] = []
    resp = build_mapping_export_response(run, "Clause44_Mapping_empty")
    wb = _load_workbook_from_response(resp)
    for sheet in ("Exempt Purchases", "ITC Ledgers", "Exclusions"):
        rows = list(wb[sheet].iter_rows(values_only=True))
        flat = [c for r in rows for c in r if isinstance(c, str)]
        assert any("No " in c for c in flat), f"{sheet} missing empty-state row"
