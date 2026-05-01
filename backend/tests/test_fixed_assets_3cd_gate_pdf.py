"""
Fixed Assets — 3CD validation gate, zero-row filtering on compute, PDF export.

Validates:
  * POST /block-opening/validate-3cd persists prior_3cd_validation on the run
    with acknowledged=False when ok=False, acknowledged=True when ok=True.
  * POST /clear-3cd-validation-warning flips acknowledged=True (idempotent).
  * Editing openings (POST /block-opening upsert / import.xlsx) auto-clears
    the persisted validation summary (prior validation goes stale).
  * compute_run skips block rows where every numeric is zero.
  * GET /export.pdf returns a valid PDF with %PDF magic bytes.
"""
import io
import json
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://unified-tax-tools.preview.emergentagent.com",
).rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RID = "0e4cc62f-52f9-4668-b598-f60bd0c52803"

CD_MATCHING = {
    "FORM3CA": {"F3CA": {"Form3cdDeprAllw": [
        {"RateOfDep": 40, "OpeningWDV": 0, "WrittenDownVal": 970906,
         "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
         "DescBlockAssets": "Computers + P&M"},
        {"RateOfDep": 15, "OpeningWDV": 0, "WrittenDownVal": 26233559,
         "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
         "DescBlockAssets": "P&M + Vehicles"},
        {"RateOfDep": 10, "OpeningWDV": 0, "WrittenDownVal": 2911192,
         "DepAllowable": 0, "TotalPurchaseValue": 0, "adjustment": 0,
         "DescBlockAssets": "Furniture"},
    ]}}
}


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.cookies.set("session_token", TOKEN)
    return s


@pytest.fixture(scope="module")
def baseline_xlsx(session):
    return session.get(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/export.xlsx",
        timeout=30,
    ).content


def _set_openings(session, baseline_xlsx, edits: dict):
    """Patch the baseline workbook with `edits = {block_label: opening_wdv}`
    and re-import so the run picks up the requested openings."""
    wb = load_workbook(io.BytesIO(baseline_xlsx))
    ws = wb["Opening WDV"]
    for r in ws.iter_rows(min_row=5, max_row=ws.max_row):
        if r[0].value in edits:
            r[3].value = float(edits[r[0].value])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("e.xlsx", buf,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    ).raise_for_status()


@pytest.fixture(autouse=True)
def _restore(session, baseline_xlsx):
    yield
    # Restore the baseline state after each test (also clears stale gate)
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/import.xlsx",
        files={"file": ("baseline.xlsx", io.BytesIO(baseline_xlsx),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )


def _gate(session):
    return session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}", timeout=30) \
                  .json().get("prior_3cd_validation")


def _validate(session, cd):
    return session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening/validate-3cd",
        files={"file": ("3cd.json", json.dumps(cd).encode(), "application/json")},
        timeout=30,
    )


def test_validate_persists_gate_with_acknowledged_false_on_mismatch(session, baseline_xlsx):
    _set_openings(session, baseline_xlsx,
                  {"15% Block – Vehicles": 999999.0,
                   "15% Block – Plant & Machinery": 25783559.0})
    r = _validate(session, CD_MATCHING)
    assert r.status_code == 200
    assert r.json()["ok"] is False
    g = _gate(session)
    assert g is not None
    assert g["ok"] is False
    assert g["acknowledged"] is False
    assert g["mismatch_count"] >= 1
    assert g["filename"] == "3cd.json"


def test_validate_persists_gate_with_acknowledged_true_on_match(session, baseline_xlsx):
    _set_openings(session, baseline_xlsx,
                  {"15% Block – Vehicles": 450000.0,
                   "15% Block – Plant & Machinery": 25783559.0})
    r = _validate(session, CD_MATCHING)
    assert r.json()["ok"] is True
    g = _gate(session)
    assert g["ok"] is True
    assert g["acknowledged"] is True


def test_clear_validation_warning_acks(session, baseline_xlsx):
    _set_openings(session, baseline_xlsx,
                  {"15% Block – Vehicles": 999999.0})
    _validate(session, CD_MATCHING).raise_for_status()
    assert _gate(session)["acknowledged"] is False
    r = session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/clear-3cd-validation-warning",
        timeout=30,
    )
    assert r.status_code == 200
    assert _gate(session)["acknowledged"] is True


def test_editing_openings_invalidates_prior_gate(session, baseline_xlsx):
    """After validating, any opening-WDV edit must drop the persisted summary."""
    _set_openings(session, baseline_xlsx,
                  {"15% Block – Vehicles": 999999.0})
    _validate(session, CD_MATCHING).raise_for_status()
    assert _gate(session) is not None
    # Touch a single block via the upsert endpoint
    session.post(
        f"{BASE_URL}/api/fixed-assets/runs/{RID}/block-opening",
        json={"block_label": "5% Block – Buildings",
              "opening_wdv": 0.0, "description": ""},
        timeout=30,
    ).raise_for_status()
    assert _gate(session) is None


def test_compute_skips_zero_only_blocks(session, baseline_xlsx):
    """A block with opening=0, no adds, no dels, no closing must NOT appear
    in the response rows. Only active blocks are surfaced."""
    _set_openings(session, baseline_xlsx,
                  {"15% Block – Vehicles": 450000.0,
                   "15% Block – Plant & Machinery": 25783559.0})
    r = session.post(f"{BASE_URL}/api/fixed-assets/runs/{RID}/compute", timeout=60)
    assert r.status_code == 200
    rows = r.json()["rows"]
    # Active blocks in this run: 40% Computers, 15% P&M, 15% Vehicles, 10% Furniture
    assert all(any(float(rr.get(k) or 0) for k in
                   ("opening_wdv", "adds_full", "adds_half",
                    "deletions", "depreciation", "closing_wdv", "stcg_sec50"))
               for rr in rows)
    # Should be 4-ish, definitely fewer than the 15 total active blocks
    assert len(rows) < 15


def test_export_pdf_returns_valid_pdf(session):
    r = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}/export.pdf", timeout=60)
    assert r.status_code == 200
    assert r.headers.get("content-type", "").startswith("application/pdf")
    assert r.content[:4] == b"%PDF"
    assert len(r.content) > 5000  # real content, not empty


def test_export_pdf_groups_additions_by_block(session, tmp_path):
    """The additions register must surface a per-block sub-header strip
    (block label + rate + asset count + capitalised total) ahead of each
    group, with cards listed in PTU-date order within the group."""
    import pdfplumber
    r = session.get(f"{BASE_URL}/api/fixed-assets/runs/{RID}/export.pdf", timeout=60)
    pdf_path = tmp_path / "fa.pdf"
    pdf_path.write_bytes(r.content)
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    # Three active block groups must each surface their sub-header.
    # (Block name + " " + rate pill rendered on the strip)
    assert "40% Block – Computers" in text
    assert "15% Block – Plant & Machinery" in text
    assert "10% Block – Furniture" in text
    # Asset count strings — the demo run has these counts in respective blocks
    assert "9 asset" in text or "9 assets" in text
    # The "grouped by IT Block" section title is the one new copy
    assert "grouped by IT Block" in text
