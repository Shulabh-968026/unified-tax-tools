"""Unit + integration tests for Excel audit working-paper export."""
from __future__ import annotations
import io
import openpyxl

from modules.gst_recon.excel_export import build_workbook
from modules.gst_recon.service import build_month_grid, build_summary


def _doc(fy="2024-25"):
    months = build_month_grid(fy, [])
    return {
        "id": "abc-test-id",
        "client_id": "cli_test",
        "fy": fy,
        "name": "Test Run",
        "status": "summarized",
        "created_at": "2026-04-28T12:00:00+00:00",
        "has_books": True,
        "has_mapping": True,
        "mapping_unmapped_ledgers": ["Writeoff ITC Expenses"],
        "mapping_filename": "ledger_mapping.xlsx",
        "mapping_row_count": 50,
        "files": [
            {"bucket": "books", "filename": "Books_24-25.json"},
            {"bucket": "mapping", "filename": "ledger_mapping.xlsx"},
        ],
        "months": months,
    }


def _summary():
    months = build_month_grid("2024-25", [])
    rows = []
    for i, m in enumerate(months):
        rows.append({
            "period": m["period"], "month_label": m["month_label"],
            "books_outward_taxable": 100000.0, "books_outward_tax": 18000.0, "books_itc_total": 9000.0,
            "r1_outward_taxable": 100000.0, "r1_outward_tax": 18000.0,
            "r2b_itc_total": 9000.0,
            "r3b_outward_taxable": 100000.0, "r3b_outward_tax": 18000.0, "r3b_itc_total": 9000.0,
            "var_r1_vs_r3b_outward": 0.0, "var_r2b_vs_r3b_itc": 0.0,
            "var_books_vs_r1_outward": 0.0, "var_books_vs_r2b_itc": 0.0,
        })
    totals_keys = list(rows[0].keys())[2:]
    totals = {k: round(sum(r[k] for r in rows), 2) for k in totals_keys}
    return {"fy": "2024-25", "rows": rows, "totals": totals}


def _matches_one_period():
    return {"Apr 2024": {
        "matched": [{"books": {"voucher_no": "S-1", "party_gstin": "33A", "total": 1180},
                     "portal": {"invoice_no": "S-1", "party_gstin": "33A", "total": 1180},
                     "value_diff": 0, "books_date": "2024-04-15", "portal_date": "2024-04-15"}],
        "value_mismatch": [],
        "date_mismatch": [],
        "missing_in_books": [],
        "missing_in_portal": [],
        "counts": {"matched": 1, "value_mismatch": 0, "date_mismatch": 0, "missing_in_books": 0, "missing_in_portal": 0},
    }}


def test_workbook_has_eight_sheets_in_correct_order():
    """Adds Annual Party-wise sheets when partywise data is provided."""
    pw = {"direction": "outward", "rows": [], "totals": {
        "books_total": 0, "portal_total": 0, "diff_total": 0,
        "books_taxable": 0, "portal_taxable": 0,
        "books_tax": 0, "portal_tax": 0,
        "diff_taxable": 0, "diff_tax": 0,
    }}
    out = build_workbook(_doc(), _summary(), _matches_one_period(), {}, pw, pw)
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb.sheetnames == [
        "Dashboard",
        "Annual Party-wise (Outward)",
        "Annual Party-wise (Inward)",
        "12-Month Summary",
        "Outward Vouchers",
        "Inward Vouchers",
        "Pending Classification",
        "Run Metadata",
    ]


def test_workbook_has_six_sheets_in_correct_order():
    """Default: when no partywise data, still produces 6-sheet workbook."""
    out = build_workbook(_doc(), _summary(), _matches_one_period(), {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    assert wb.sheetnames == [
        "Dashboard", "12-Month Summary", "Outward Vouchers",
        "Inward Vouchers", "Pending Classification", "Run Metadata",
    ]


def test_dashboard_contains_kpi_section_and_status_banner():
    out = build_workbook(_doc(), _summary(), {}, {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Dashboard"]
    title = ws["B2"].value
    assert title == "Reconciliation Health"
    # All variances are zero in our fixture → ALL RECONCILED banner
    body_text = " ".join(str(c.value or "") for r in ws.iter_rows(values_only=False) for c in r)
    assert "ALL RECONCILED" in body_text


def test_summary_sheet_has_two_blocks_with_annual_totals():
    out = build_workbook(_doc(), _summary(), {}, {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["12-Month Summary"]
    cells = [str(c.value or "") for r in ws.iter_rows(values_only=False) for c in r]
    assert any("OUTWARD TURNOVER" in t for t in cells)
    assert any("INPUT TAX CREDIT" in t for t in cells)
    assert cells.count("Annual") == 2  # one per block


def test_outward_voucher_sheet_lists_matched_pair():
    out = build_workbook(_doc(), _summary(), _matches_one_period(), {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Outward Vouchers"]
    # Header at row 1; matched row at row 2
    assert ws["A1"].value == "Month"
    assert ws["A2"].value == "Apr 2024"
    assert ws["B2"].value == "matched"
    # Books # is column E now (after adding Party Name in column D)
    assert ws["E2"].value == "S-1"
    # Party Name column should exist
    assert ws["D1"].value == "Party Name"


def test_pending_classification_lists_unmapped_ledgers():
    out = build_workbook(_doc(), _summary(), {}, {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Pending Classification"]
    cells = [str(c.value or "") for r in ws.iter_rows(values_only=False) for c in r]
    assert any("Writeoff ITC Expenses" in t for t in cells)


def test_metadata_sheet_includes_run_id_and_filenames():
    out = build_workbook(_doc(), _summary(), {}, {})
    wb = openpyxl.load_workbook(io.BytesIO(out))
    ws = wb["Run Metadata"]
    cells = [str(c.value or "") for r in ws.iter_rows(values_only=False) for c in r]
    assert any("abc-test-id" in t for t in cells)
    assert any("ledger_mapping.xlsx" in t for t in cells)
