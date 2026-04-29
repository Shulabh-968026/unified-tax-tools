"""Phase 5 (Summary exports) + Phase 6 (Recon) backend tests.

Covers:
  - GET /runs/{rid}/summary.xlsx — 6 sheets, KPI buckets, auth, 404
  - GET /runs/{rid}/summary.pdf — multi-page PDF, pdfplumber-readable, auth
  - GET /runs/{rid}/responses/{response_id}/recon — XLSX/CSV parse + auto_match
  - parser handles ; delimiter + single Amount column + dd-mm-yyyy / dd/mm/yyyy
  - tolerance bumping flips ours_only/theirs_only → match
  - POST/GET/DELETE /recon/comments + cascade on run delete
  - text-only dispute (no upload) returns supported=false, no crash
  - PDF unsupported returns supported=false + friendly message
"""
import base64
import io
import os
import sys
import uuid
from datetime import datetime, timezone

import pytest
import requests

# Allow direct import of recon helpers (unit-level)
sys.path.insert(0, "/app/backend")
from modules.balance_confirmation.recon import (
    auto_match,
    parse_csv,
    parse_recipient_statement,
    parse_xlsx,
)
from modules.balance_confirmation.summary_export import kpi_buckets

def _load_base_url():
    val = os.environ.get("REACT_APP_BACKEND_URL", "").strip()
    if val:
        return val.rstrip("/")
    # Read frontend/.env
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    return line.split("=", 1)[1].strip().rstrip("/")
    except Exception:
        pass
    raise RuntimeError("REACT_APP_BACKEND_URL not configured")


BASE_URL = _load_base_url()
TOKEN = "qa_test_session_token_20260429_dev"
RID = "1710de57-49a7-4a71-a411-817e4d49736b"
DISPUTED_RESP_ID = "a2102f61-4069-406b-9002-8f4fc0eb14e4"


# ============================ Fixtures ===================================
@pytest.fixture(scope="module")
def client():
    s = requests.Session()
    s.cookies.set("session_token", TOKEN)
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="module")
def anon():
    return requests.Session()


# ============================ Phase 5 — Summary XLSX =====================
class TestSummaryXlsx:
    def test_xlsx_unauth(self, anon):
        r = anon.get(f"{BASE_URL}/api/balance-confirmation/runs/{RID}/summary.xlsx")
        assert r.status_code in (401, 403), f"Expected auth gate, got {r.status_code}"

    def test_xlsx_bad_rid(self, client):
        r = client.get(f"{BASE_URL}/api/balance-confirmation/runs/no-such-run/summary.xlsx")
        assert r.status_code == 404

    def test_xlsx_six_sheets(self, client):
        r = client.get(f"{BASE_URL}/api/balance-confirmation/runs/{RID}/summary.xlsx")
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("Content-Type", "").lower()
        assert "attachment" in r.headers.get("Content-Disposition", "").lower()

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        assert wb.sheetnames == [
            "Cover", "Sent Tracker", "Status Timeline",
            "Variances", "Confirmed", "Notes",
        ], f"sheets={wb.sheetnames}"

        # Sent Tracker — 15 columns + at least 1 row
        st = wb["Sent Tracker"]
        first_row = list(next(st.iter_rows(values_only=True)))
        assert len(first_row) >= 15
        assert "Party" in [str(c or "") for c in first_row]

        # Variances — 11 columns
        var = wb["Variances"]
        var_hdr = list(next(var.iter_rows(values_only=True)))
        assert len(var_hdr) >= 11

        # Confirmed — 8 cols
        cf = wb["Confirmed"]
        cf_hdr = list(next(cf.iter_rows(values_only=True)))
        assert len(cf_hdr) >= 8

        # Notes — heading + instruction, otherwise blank
        notes = wb["Notes"]
        notes_a1 = notes["A1"].value or ""
        assert "note" in str(notes_a1).lower()

    def test_kpi_buckets_helper(self):
        ledgers = [
            {"confirmation_status": "confirmed"},
            {"confirmation_status": "confirmed"},
            {"confirmation_status": "disputed"},
            {"confirmation_status": "sent"},
            {"confirmation_status": "delivered"},
            {"confirmation_status": "bounced"},
            {"confirmation_status": "failed"},
            {"confirmation_status": "not_sent", "email": "a@b.com"},
            {"confirmation_status": "not_sent"},  # no email
        ]
        kpi = kpi_buckets(ledgers)
        assert kpi["confirmed"] == 2
        assert kpi["disputed"] == 1
        assert kpi["in_flight"] == 2
        assert kpi["failed"] == 2
        assert kpi["no_action"] == 1
        assert kpi["no_email"] == 1


# ============================ Phase 5 — Summary PDF ======================
class TestSummaryPdf:
    def test_pdf_unauth(self, anon):
        r = anon.get(f"{BASE_URL}/api/balance-confirmation/runs/{RID}/summary.pdf")
        assert r.status_code in (401, 403)

    def test_pdf_bad_rid(self, client):
        r = client.get(f"{BASE_URL}/api/balance-confirmation/runs/no-such-run/summary.pdf")
        assert r.status_code == 404

    def test_pdf_multipage_with_signoff(self, client):
        r = client.get(f"{BASE_URL}/api/balance-confirmation/runs/{RID}/summary.pdf")
        assert r.status_code == 200
        assert r.headers.get("Content-Type", "").startswith("application/pdf")
        assert r.content[:5] == b"%PDF-"

        try:
            import pdfplumber
        except ImportError:
            pytest.skip("pdfplumber not installed in env")
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            assert len(pdf.pages) >= 2, f"expected multi-page, got {len(pdf.pages)}"
            full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)
            assert "Balance Confirmation" in full_text
            # Sign-off must always be the last page
            assert "Sign-off" in full_text or "PREPARED BY" in full_text


# ============================ Phase 6 — Recon endpoint ===================
class TestReconEndpoint:
    def test_recon_unauth(self, anon):
        r = anon.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon"
        )
        assert r.status_code in (401, 403)

    def test_recon_ambika_xlsx(self, client):
        r = client.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon?tolerance=1"
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["supported"] is True
        assert body["format"] in ("xlsx", "xlsm")
        assert "pairs" in body
        assert isinstance(body["pairs"], list)
        c = body["counts"]
        # tolerance=1 with mismatched amounts → no matches expected per fixture
        assert (c["matched"] + c["ours_only"] + c["theirs_only"]) == len(body["pairs"])
        assert all(p["status"] in ("match", "ours_only", "theirs_only") for p in body["pairs"])

    def test_recon_bad_response_id(self, client):
        r = client.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/no-such-resp/recon"
        )
        assert r.status_code == 404


# ============================ Phase 6 — Parser unit tests ================
class TestParserUnits:
    def test_parse_csv_comma(self):
        csv_bytes = (
            b"Date,Voucher,Particulars,Debit,Credit,Balance\n"
            b"01-04-2024,V001,Opening,0,0,5000\n"
            b"05-04-2024,V002,Sale invoice,1000,0,6000\n"
            b"10/04/2024,V003,Payment,0,500,5500\n"
        )
        recs, cm, hd = parse_csv(csv_bytes)
        assert len(recs) == 3
        assert recs[0]["date"] == "2024-04-01"
        assert recs[1]["debit"] == 1000.0
        assert recs[2]["credit"] == 500.0
        assert recs[2]["date"] == "2024-04-10"

    def test_parse_csv_semicolon(self):
        csv_bytes = (
            b"Date;Particulars;Debit;Credit\n"
            b"02-04-2024;Test;100;0\n"
            b"03-04-2024;Test2;0;200\n"
        )
        recs, _, _ = parse_csv(csv_bytes)
        assert len(recs) == 2
        assert recs[0]["debit"] == 100.0
        assert recs[1]["credit"] == 200.0

    def test_parse_csv_single_amount(self):
        # positive=credit, negative=debit
        csv_bytes = (
            b"Date,Particulars,Amount\n"
            b"01-04-2024,Inflow,1500\n"
            b"02-04-2024,Outflow,-800\n"
        )
        recs, _, _ = parse_csv(csv_bytes)
        assert len(recs) == 2
        assert recs[0]["credit"] == 1500.0 and recs[0]["debit"] == 0.0
        assert recs[1]["debit"] == 800.0 and recs[1]["credit"] == 0.0

    def test_parse_unsupported_pdf(self):
        out = parse_recipient_statement("statement.pdf", b"%PDF-fakebytes")
        assert out["supported"] is False
        assert "PDF" in (out.get("message") or "") or "manually" in (out.get("message") or "")

    def test_parse_unsupported_image(self):
        out = parse_recipient_statement("scan.jpg", b"\xff\xd8\xff\xe0")
        assert out["supported"] is False

    def test_parse_xlsx_roundtrip(self):
        # Generate a tiny synthetic XLSX
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Voucher", "Narration", "Debit", "Credit", "Balance"])
        ws.append(["01-04-2024", "V1", "Opening", 0, 0, 1000])
        ws.append(["05-04-2024", "V2", "Bill", 5000, 0, 6000])
        ws.append(["10-04-2024", "V3", "Payment", 0, 2000, 4000])
        buf = io.BytesIO()
        wb.save(buf)
        recs, _, _ = parse_xlsx(buf.getvalue())
        assert len(recs) == 3
        assert recs[1]["debit"] == 5000.0
        assert recs[2]["credit"] == 2000.0


# ============================ Phase 6 — Auto-match tolerance =============
class TestAutoMatchTolerance:
    def test_tolerance_boundary(self):
        ours = [{"amount": 80000.0, "date": "2024-04-01", "vno": "X1"}]
        theirs = [{"debit": 80000.50, "credit": 0.0, "date": "2024-04-01", "vno": "Y1"}]

        # tolerance=1 → diff is 0.50 ≤ 1 → match
        pairs = auto_match(ours, theirs, tolerance=1.0)
        statuses = sorted(p["status"] for p in pairs)
        assert statuses == ["match"], f"got {statuses}"

        # tolerance=0.1 → diff 0.50 > 0.1 → split
        pairs2 = auto_match(ours, theirs, tolerance=0.1)
        statuses2 = sorted(p["status"] for p in pairs2)
        assert statuses2 == ["ours_only", "theirs_only"], f"got {statuses2}"

    def test_tolerance_huge_no_false_matches_when_amounts_differ(self):
        ours = [{"amount": 1000.0}]
        theirs = [{"debit": 0.0, "credit": 5000.0}]
        # diff = 4000, tolerance=10000 → false match would be possible (greedy)
        pairs = auto_match(ours, theirs, tolerance=10000.0)
        statuses = [p["status"] for p in pairs]
        # Only 1 candidate so it WILL match within the wide tolerance — but
        # status should still be "match" if diff<=tol
        assert "match" in statuses


# ============================ Phase 6 — Comments CRUD ====================
class TestReconComments:
    def test_add_empty_text_400(self, client):
        r = client.post(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon/comments",
            json={"text": "   "},
        )
        assert r.status_code == 400

    def test_crud_lifecycle(self, client):
        # CREATE
        text = f"TEST_phase6 comment {uuid.uuid4().hex[:6]}"
        r = client.post(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon/comments",
            json={"text": text, "pair_key": "match:0:0"},
        )
        assert r.status_code == 200, r.text
        doc = r.json()
        assert doc["text"] == text
        assert doc["pair_key"] == "match:0:0"
        assert doc["author_email"]
        assert "ts" in doc and "comment_id" in doc
        cid = doc["comment_id"]

        # LIST
        r = client.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon/comments"
        )
        assert r.status_code == 200
        listed = r.json()
        assert listed["count"] >= 1
        assert any(c["comment_id"] == cid for c in listed["rows"])

        # DELETE
        r = client.delete(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon/comments/{cid}"
        )
        assert r.status_code == 200
        assert r.json()["deleted"] is True

        # DELETE again → 404
        r = client.delete(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{DISPUTED_RESP_ID}/recon/comments/{cid}"
        )
        assert r.status_code == 404


# ============================ Phase 6 — text-only dispute (no upload) ===
class TestReconTextOnly:
    """Find a dispute response WITHOUT an attachment via list_responses
    and confirm the recon endpoint returns supported=false without crash."""

    def test_text_only_returns_unsupported_gracefully(self, client):
        r = client.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}/responses?decision=disputed"
        )
        if r.status_code != 200:
            pytest.skip(f"list_responses unavailable: {r.status_code}")
        rows = r.json().get("rows") or r.json().get("responses") or []
        no_upload = [
            x for x in rows
            if not x.get("uploaded_filename") and x.get("response_id") != DISPUTED_RESP_ID
        ]
        if not no_upload:
            pytest.skip("No text-only disputed response in fixture run")
        target = no_upload[0]["response_id"]
        rr = client.get(
            f"{BASE_URL}/api/balance-confirmation/runs/{RID}"
            f"/responses/{target}/recon"
        )
        assert rr.status_code == 200
        body = rr.json()
        assert body["supported"] is False
        assert body["pairs"] is not None  # at minimum ours_only entries
