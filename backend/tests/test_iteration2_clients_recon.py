"""Iteration 2 backend tests: clients CRUD, multi-division, recon, consolidated."""
import io
import openpyxl
import pytest
import requests


# ----- Clients CRUD ---------------------------------------------------------

class TestClientsCRUD:
    def test_create_single_client(self, base_url, auth_headers):
        r = requests.post(
            f"{base_url}/api/clients",
            headers=auth_headers,
            json={"file_number": "TEST_FN_001", "name": "TEST_Single_Co", "type": "single"},
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["client_id"].startswith("cli_")
        assert d["file_number"] == "TEST_FN_001"
        assert d["name"] == "TEST_Single_Co"
        assert d["type"] == "single"
        assert d["divisions"] == []
        # GET to verify persisted
        g = requests.get(f"{base_url}/api/clients/{d['client_id']}", headers=auth_headers)
        assert g.status_code == 200
        assert g.json()["client_id"] == d["client_id"]

    def test_create_multi_client(self, base_url, auth_headers):
        r = requests.post(
            f"{base_url}/api/clients",
            headers=auth_headers,
            json={
                "file_number": "TEST_FN_M01",
                "name": "TEST_Multi_Co",
                "type": "multi",
                "divisions": ["Spinning", "Weaving"],
            },
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["type"] == "multi"
        assert len(d["divisions"]) == 2
        names = [div["name"] for div in d["divisions"]]
        assert "Spinning" in names and "Weaving" in names
        for div in d["divisions"]:
            assert div["division_id"].startswith("div_")

    def test_create_multi_without_divisions_returns_400(self, base_url, auth_headers):
        r = requests.post(
            f"{base_url}/api/clients",
            headers=auth_headers,
            json={"file_number": "TEST_FN_BAD", "name": "TEST_Bad_Multi", "type": "multi"},
        )
        assert r.status_code == 400, r.text

    def test_list_clients(self, base_url, auth_headers):
        r = requests.get(f"{base_url}/api/clients", headers=auth_headers)
        assert r.status_code == 200
        items = r.json()["clients"]
        names = [c["name"] for c in items]
        assert "TEST_Single_Co" in names
        assert "TEST_Multi_Co" in names

    def test_archived_filter(self, base_url, auth_headers):
        r = requests.get(f"{base_url}/api/clients?archived=true", headers=auth_headers)
        assert r.status_code == 200
        names = [c["name"] for c in r.json()["clients"]]
        assert "TEST_Single_Co" not in names

    def test_patch_add_division_to_single_rejected(self, base_url, auth_headers):
        # find single
        r = requests.get(f"{base_url}/api/clients", headers=auth_headers)
        single = next(c for c in r.json()["clients"] if c["name"] == "TEST_Single_Co")
        p = requests.patch(
            f"{base_url}/api/clients/{single['client_id']}",
            headers=auth_headers,
            json={"add_divisions": ["Should_Fail"]},
        )
        assert p.status_code == 400

    def test_patch_add_division_to_multi(self, base_url, auth_headers):
        r = requests.get(f"{base_url}/api/clients", headers=auth_headers)
        multi = next(c for c in r.json()["clients"] if c["name"] == "TEST_Multi_Co")
        p = requests.patch(
            f"{base_url}/api/clients/{multi['client_id']}",
            headers=auth_headers,
            json={"add_divisions": ["Dyeing"]},
        )
        assert p.status_code == 200
        assert any(d["name"] == "Dyeing" for d in p.json()["divisions"])


# ----- Helpers --------------------------------------------------------------

@pytest.fixture(scope="module")
def single_client(base_url, auth_headers):
    r = requests.post(
        f"{base_url}/api/clients",
        headers=auth_headers,
        json={"file_number": "TEST_RUN_FN", "name": "TEST_RunSingle", "type": "single"},
    )
    assert r.status_code == 200, r.text
    return r.json()


@pytest.fixture(scope="module")
def multi_client(base_url, auth_headers):
    r = requests.post(
        f"{base_url}/api/clients",
        headers=auth_headers,
        json={
            "file_number": "TEST_MULTI_FN",
            "name": "TEST_RunMulti",
            "type": "multi",
            "divisions": ["DivA", "DivB"],
        },
    )
    assert r.status_code == 200, r.text
    return r.json()


def _post_run(base_url, headers, files_paths, client_id, period, division_id=None):
    with open(files_paths["json"], "rb") as fj, open(files_paths["xlsx"], "rb") as fx:
        files = {
            "accounting_json": ("accounting.json", fj, "application/json"),
            "ledger_xlsx": ("ledger.xlsx", fx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        }
        data = {"client_id": client_id, "period": period}
        if division_id:
            data["division_id"] = division_id
        return requests.post(f"{base_url}/api/runs", headers=headers, files=files, data=data, timeout=180)


# ----- Run requires client_id & period --------------------------------------

class TestRunRequiresClient:
    def test_run_without_client_id_fails(self, base_url, auth_headers, sample_files):
        with open(sample_files["json"], "rb") as fj, open(sample_files["xlsx"], "rb") as fx:
            files = {
                "accounting_json": ("a.json", fj, "application/json"),
                "ledger_xlsx": ("l.xlsx", fx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            }
            r = requests.post(f"{base_url}/api/runs", headers=auth_headers, files=files, timeout=60)
        assert r.status_code in (400, 422)

    def test_run_multi_without_division_fails(self, base_url, auth_headers, sample_files, multi_client):
        r = _post_run(base_url, auth_headers, sample_files, multi_client["client_id"], "FY2023-24")
        assert r.status_code == 400, r.text


# ----- Single client: full run + recon --------------------------------------

@pytest.fixture(scope="module")
def single_generated(base_url, auth_headers, sample_files, single_client):
    period = "FY2023-24"
    r = _post_run(base_url, auth_headers, sample_files, single_client["client_id"], period)
    assert r.status_code == 200, r.text
    run = r.json()
    itc = [c["name"] for c in run["itc_candidates"] if c.get("suggested")]
    excluded = [p["name"] for p in run["pl_ledgers"] if p.get("suggested")]
    g = requests.post(
        f"{base_url}/api/runs/{run['run_id']}/generate",
        headers=auth_headers,
        json={"itc_ledgers": itc, "excluded_ledgers": excluded},
        timeout=180,
    )
    assert g.status_code == 200, g.text
    return {
        "run_id": run["run_id"],
        "client_id": single_client["client_id"],
        "period": period,
        "data": g.json(),
    }


class TestSingleRecon:
    def test_recon_present(self, single_generated):
        d = single_generated["data"]
        assert "recon" in d
        recon = d["recon"]
        for k in ("total_books", "excluded_lines", "excluded_total", "balance"):
            assert k in recon

    def test_recon_balance_equals_col2_total(self, single_generated):
        s = single_generated["data"]["summary"]
        recon = single_generated["data"]["recon"]
        assert abs(recon["balance"] - s["col2_total"]) < 0.01

    def test_recon_arithmetic(self, single_generated):
        recon = single_generated["data"]["recon"]
        assert abs((recon["total_books"] - recon["excluded_total"]) - recon["balance"]) < 0.5
        # excluded_lines sum equals excluded_total
        s = sum(line["amount"] for line in recon["excluded_lines"])
        assert abs(s - recon["excluded_total"]) < 0.5

    def test_get_run_returns_recon_and_client(self, base_url, auth_headers, single_generated):
        r = requests.get(f"{base_url}/api/runs/{single_generated['run_id']}", headers=auth_headers)
        assert r.status_code == 200
        d = r.json()
        assert d["client_id"] == single_generated["client_id"]
        assert d["period"] == single_generated["period"]
        assert "recon" in d
        assert abs(d["recon"]["balance"] - d["summary"]["col2_total"]) < 0.01

    def test_runs_filter_by_client_period(self, base_url, auth_headers, single_generated):
        r = requests.get(
            f"{base_url}/api/runs",
            headers=auth_headers,
            params={"client_id": single_generated["client_id"], "period": single_generated["period"]},
        )
        assert r.status_code == 200
        runs = r.json()["runs"]
        assert len(runs) >= 1
        assert all(run["client_id"] == single_generated["client_id"] for run in runs)
        assert all(run["period"] == single_generated["period"] for run in runs)

    def test_export_xlsx_has_recon_sheet(self, base_url, auth_headers, single_generated):
        r = requests.get(f"{base_url}/api/runs/{single_generated['run_id']}/export",
                         headers=auth_headers, timeout=60)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
        assert "Clause 44" in wb.sheetnames
        assert "Reconciliation" in wb.sheetnames
        assert "Transaction Audit Trail" in wb.sheetnames
        ws = wb["Reconciliation"]
        # Find row 'Expenditure as per Clause 44 Report'
        col2_total = single_generated["data"]["summary"]["col2_total"]
        found = False
        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and "Expenditure as per Clause 44 Report" in cell:
                    # Look for numeric in same row
                    for v in row:
                        if isinstance(v, (int, float)) and abs(v - col2_total) < 0.01:
                            found = True
        assert found, "Reconciliation sheet should have row 'Expenditure as per Clause 44 Report' with col2_total"


# ----- Multi client consolidated --------------------------------------------

@pytest.fixture(scope="module")
def multi_generated(base_url, auth_headers, sample_files, multi_client):
    period = "FY2023-24"
    div_results = []
    for div in multi_client["divisions"][:2]:
        r = _post_run(base_url, auth_headers, sample_files, multi_client["client_id"], period, div["division_id"])
        assert r.status_code == 200, r.text
        run = r.json()
        itc = [c["name"] for c in run["itc_candidates"] if c.get("suggested")]
        excluded = [p["name"] for p in run["pl_ledgers"] if p.get("suggested")]
        g = requests.post(
            f"{base_url}/api/runs/{run['run_id']}/generate",
            headers=auth_headers,
            json={"itc_ledgers": itc, "excluded_ledgers": excluded},
            timeout=180,
        )
        assert g.status_code == 200, g.text
        div_results.append({"division_id": div["division_id"], "summary": g.json()["summary"], "recon": g.json()["recon"]})
    return {"client_id": multi_client["client_id"], "period": period, "divs": div_results}


class TestConsolidated:
    def test_consolidated_endpoint(self, base_url, auth_headers, multi_generated):
        r = requests.get(
            f"{base_url}/api/clients/{multi_generated['client_id']}/consolidated",
            headers=auth_headers,
            params={"period": multi_generated["period"]},
        )
        assert r.status_code == 200, r.text
        d = r.json()
        assert "summary" in d and "division_summaries" in d
        assert "transactions" in d
        assert "recon" in d
        # division_summaries length == 2
        assert len(d["division_summaries"]) == 2
        # summed col2_total
        expected = sum(div["summary"]["col2_total"] for div in multi_generated["divs"])
        assert abs(d["summary"]["col2_total"] - expected) < 0.5
        # recon balance == merged col2_total
        assert abs(d["recon"]["balance"] - d["summary"]["col2_total"]) < 0.5

    def test_consolidated_export(self, base_url, auth_headers, multi_generated):
        r = requests.get(
            f"{base_url}/api/clients/{multi_generated['client_id']}/consolidated/export",
            headers=auth_headers,
            params={"period": multi_generated["period"]},
            timeout=120,
        )
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
        for s in ("Clause 44", "Reconciliation", "Transaction Audit Trail"):
            assert s in wb.sheetnames, f"Missing sheet {s}"

    def test_consolidated_404_for_unknown_period(self, base_url, auth_headers, multi_generated):
        r = requests.get(
            f"{base_url}/api/clients/{multi_generated['client_id']}/consolidated",
            headers=auth_headers,
            params={"period": "FY-NEVER"},
        )
        assert r.status_code == 404
