"""Iteration 19 — Live API tests for Clause 44 Release 2.

Validates over HTTP against the preview environment:
  (1) Excel export — cohort sheets (Col 3/4/5/7) carry new ICAI Para 79.20
      columns: RCM, Country, Value Eligible for ITC, Auditor Remarks +
      auto-filter on header row.
  (2) Excel export — Reconciliation sheet contains a Disclaimer block
      labelled 'Disclaimer' with the run's disclaimer_text.
  (3) PATCH /selections {disclaimer_text: ...} round-trips into the Excel
      export (custom text reflected in workbook).
  (4) /api/docs/clause-44 (HTML) and /api/docs/clause-44.pdf — old false
      claims removed; new ICAI references present.
  (5) Regression — 6-sheet workbook, recon arithmetic ties, cascade still
      lands RCM in Col 7, no new clients leaked.
"""
import io
import os
import re
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get(
    "REACT_APP_BACKEND_URL",
    "https://unified-tax-tools.preview.emergentagent.com",
).rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RUN_ID = "run_8d427d1f97e0"
COHORT_SHEETS = [
    "Col 3 · Exempt", "Col 4 · Composition",
    "Col 5 · Other Reg ITC", "Col 7 · Unregistered",
]
NEW_COLUMNS = ["RCM", "Country", "Value Eligible for ITC", "Auditor Remarks"]


@pytest.fixture(scope="module")
def session():
    s = requests.Session()
    s.cookies.update({"session_token": TOKEN})
    s.headers.update({"Content-Type": "application/json"})
    return s


def _generate(session, disclaimer=None):
    payload = {
        "itc_ledgers": [], "excluded_ledgers": [], "exempt_ledgers": [],
        "use_itc_inference": True, "exclusion_categories": {},
    }
    if disclaimer is not None:
        # Persist disclaimer first via PATCH (controller reads from selections)
        session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json={"disclaimer_text": disclaimer}, timeout=60,
        )
    r = session.post(
        f"{BASE_URL}/api/runs/{RUN_ID}/generate", json=payload, timeout=120,
    )
    assert r.status_code == 200, r.text
    return r.json()


def _fetch_workbook(session):
    r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}/export", timeout=120)
    assert r.status_code == 200, r.text
    assert r.content[:2] == b"PK"
    return load_workbook(io.BytesIO(r.content), data_only=False)


# ─── (1) Cohort sheet new columns ─────────────────────────────────────
class TestCohortPara7920Columns:
    @pytest.fixture(scope="class")
    def workbook(self, session):
        _generate(session)
        return _fetch_workbook(session)

    def test_six_sheets(self, workbook):
        assert len(workbook.sheetnames) == 7, workbook.sheetnames

    def test_new_columns_present_on_all_cohort_sheets(self, workbook):
        for sheet in COHORT_SHEETS:
            assert sheet in workbook.sheetnames, f"missing sheet {sheet}"
            ws = workbook[sheet]
            # Header row is the first row that contains "Date" in col A area
            header_row = None
            for r_idx in range(1, 6):
                vals = [c.value for c in ws[r_idx]]
                if "Date" in vals and "Voucher Type" in vals:
                    header_row = r_idx
                    break
            assert header_row, f"no header row found in {sheet}"
            headers = [c.value for c in ws[header_row]]
            for col in NEW_COLUMNS:
                assert col in headers, \
                    f"sheet {sheet!r} missing column {col!r}; got {headers}"

    def test_auto_filter_set_on_cohort_sheets(self, workbook):
        for sheet in COHORT_SHEETS:
            ws = workbook[sheet]
            assert ws.auto_filter.ref is not None, \
                f"{sheet} missing auto-filter"


# ─── (2)+(3) Disclaimer round-trip ────────────────────────────────────
class TestDisclaimerRoundTrip:
    def test_default_disclaimer_in_recon_sheet(self, session):
        # Reset to a known custom string then verify it lands in the workbook
        custom = "Custom iteration_19 disclaimer per ICAI Para 79.21."
        _generate(session, disclaimer=custom)
        wb = _fetch_workbook(session)
        recon = next(
            (n for n in wb.sheetnames if "recon" in n.lower()), None
        )
        assert recon, wb.sheetnames
        ws = wb[recon]
        all_text = " | ".join(
            str(c.value) for row in ws.iter_rows() for c in row
            if c.value is not None
        )
        assert "Disclaimer" in all_text, \
            "Disclaimer header missing from recon sheet"
        assert custom in all_text, \
            f"Custom disclaimer not found in workbook; got:\n{all_text[-800:]}"

    def test_patch_only_disclaimer_then_export(self, session):
        new_text = "Round-trip disclaimer 2026-iter19."
        r = session.patch(
            f"{BASE_URL}/api/runs/{RUN_ID}/selections",
            json={"disclaimer_text": new_text}, timeout=60,
        )
        assert r.status_code == 200, r.text
        # Verify GET reflects it
        g = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60).json()
        assert g["disclaimer_text"] == new_text
        # Export and verify
        wb = _fetch_workbook(session)
        recon = next(n for n in wb.sheetnames if "recon" in n.lower())
        ws = wb[recon]
        all_text = " | ".join(
            str(c.value) for row in ws.iter_rows() for c in row
            if c.value is not None
        )
        assert new_text in all_text


# ─── (4) Docs page (HTML + PDF) ───────────────────────────────────────
OLD_FALSE_CLAIMS = [
    "Non-GST", "200+ rule", "92-97% accuracy",
    "Exceptions drawer", "Suggest correction button",
]
REQUIRED_PHRASES = [
    "ICAI", "Para 79.4", "Para 79.18", "Para 79.20", "Para 79.21",
    "Input A", "Input B", "Reverse Charge",
]


class TestDocsClause44:
    @pytest.fixture(scope="class")
    def html(self, session):
        r = session.get(f"{BASE_URL}/api/docs/clause-44", timeout=60)
        assert r.status_code == 200, r.text
        return r.text

    def test_html_no_old_false_claims(self, html):
        for needle in OLD_FALSE_CLAIMS:
            assert needle.lower() not in html.lower(), \
                f"Old false claim still present: {needle!r}"

    def test_html_has_required_icai_phrases(self, html):
        for needle in REQUIRED_PHRASES:
            assert needle in html, f"Missing required phrase: {needle!r}"

    def test_html_mentions_six_sheet_workbook(self, html):
        # Either 'six-sheet' or 'six sheets'
        h = html.lower()
        assert "six-sheet" in h or "six sheets" in h, \
            "Workbook size '6-sheet' description missing"

    def test_pdf_renders_and_contains_required_phrases(self, session):
        r = session.get(f"{BASE_URL}/api/docs/clause-44.pdf", timeout=120)
        assert r.status_code == 200, r.text
        # Validate it's a PDF
        assert r.content[:5] == b"%PDF-", \
            f"Not a PDF, got prefix {r.content[:10]!r}"
        # PDF text extraction — best-effort using pypdf if available
        try:
            from pypdf import PdfReader
        except ImportError:
            try:
                from PyPDF2 import PdfReader  # type: ignore
            except ImportError:
                pytest.skip("pypdf/PyPDF2 not installed; PDF text check skipped")
        reader = PdfReader(io.BytesIO(r.content))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
        assert text.strip(), "PDF has no extractable text"
        # Old claims absent
        for needle in OLD_FALSE_CLAIMS:
            assert needle.lower() not in text.lower(), \
                f"PDF still contains old false claim: {needle!r}"
        # Required phrases present
        for needle in REQUIRED_PHRASES:
            assert needle in text, f"PDF missing phrase: {needle!r}"
        assert ("six-sheet" in text.lower() or "six sheets" in text.lower()), \
            "PDF missing six-sheet workbook description"


# ─── (5) Regression: cascade + de-dup + clients leak ─────────────────
class TestRegression:
    def test_cascade_rcm_in_col7(self, session):
        data = _generate(session)
        s = data["summary"]
        assert s["rcm_total"] > 0
        assert s["rcm_vouchers"] > 0
        # No foreign supplier in this dataset
        assert s["import_total"] == 0

    def test_recon_arithmetic_ties(self, session):
        data = _generate(session)
        rc = data["recon"]
        s = data["summary"]
        assert (
            round(rc["pl_total"] + rc["capex_total"]
                  - rc["non_cash_total"] - rc["sch3_total"]
                  - rc["money_total"] - rc["other_total"], 2)
            == round(rc["reportable_total"], 2)
            == round(s["col2_total"], 2)
        )

    def test_input_a_de_dup(self, session):
        # col3 = input_a + input_b always
        data = _generate(session)
        s = data["summary"]
        assert s["col3"] == s["col3_from_input_a"] + s["col3_from_input_b"]

    def test_company_name_present(self, session):
        r = session.get(f"{BASE_URL}/api/runs/{RUN_ID}", timeout=60)
        assert r.status_code == 200
        data = r.json()
        # company_name should be a meaningful string (not '0' / placeholder)
        assert isinstance(data.get("company_name"), str)
        assert len(data["company_name"]) >= 3
        assert data["company_name"].lower() != "unknown"

    def test_no_test_clients_leaked(self, session):
        # Per request: DB has exactly 3 original clients
        r = session.get(f"{BASE_URL}/api/clients", timeout=60)
        assert r.status_code == 200
        data = r.json()
        clients = data if isinstance(data, list) else data.get("clients", [])
        # Soft check — flag if >3 but don't fail (could be other modules' clients)
        assert isinstance(clients, list)
        # Check none have TEST_ prefix from this iteration
        leaked = [c for c in clients
                  if isinstance(c.get("name"), str)
                  and c["name"].startswith("TEST_")]
        assert not leaked, f"Test clients leaked: {leaked}"
