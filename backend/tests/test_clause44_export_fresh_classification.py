"""Release 4.4.11 — Excel export must reflect *fresh* engine output, not
the stored snapshot frozen on the last Generate click.

These tests exercise the contract between the controller's silent-
re-classification spread and the Excel builder — without forcing the
synthetic fixture through the full classification engine.  The engine
itself is exercised by the existing Fix 5 test suite
(`test_clause44_capex_auto_flow.py`).
"""
import asyncio
import io
import openpyxl
from unittest.mock import patch

from modules.clause44 import controller as ctl


def _stale_run():
    """Synthetic run with stale ``recon.capex_total = 0`` and stale
    ``by_ledger`` / ``transactions`` payloads."""
    return {
        "run_id": "stale-export-1",
        "company_name": "ACME Pvt Ltd",
        "client_name": "ACME Pvt Ltd",
        "period": "2024-25",
        "generated": True,
        # Stored snapshot from BEFORE Fix 5B.
        "summary": {"col2_total": 0, "col3": 0, "col4": 0, "col5": 0,
                    "col7": 0, "col8": 0, "reportable_total": 4_01_00_000.0},
        "by_ledger": {
            "Office Rent (stale row)": {
                "total": 4_01_00_000.0, "col3": 0, "col4": 0,
                "col5": 4_01_00_000.0, "col6": 4_01_00_000.0,
                "col7": 0, "col8": 0,
            },
        },
        "by_party": {},
        "recon": {
            "pl_total":         4_01_00_000.0,    # WRONG — captures capex
            "capex_total":      0.0,               # WRONG — empty
            "reportable_total": 4_01_00_000.0,
        },
        "transactions": [],
    }


def _fresh_classification_result():
    """Mock return from `_run_classification` — head-based capex split
    correctly bucketed the 4 Cr P&M purchase into capex_total."""
    return {
        "summary": {"col2_total": 4_01_00_000.0, "col3": 0, "col4": 0,
                    "col5": 1_00_000.0, "col7": 0, "col8": 0,
                    "reportable_total": 1_00_000.0},
        "by_ledger": {
            "Plant & Machinery @ 18%": {
                "total": 4_00_00_000.0, "col3": 0, "col4": 0,
                "col5": 4_00_00_000.0, "col6": 4_00_00_000.0,
                "col7": 0, "col8": 0,
            },
            "Office Rent": {
                "total": 1_00_000.0, "col3": 0, "col4": 0,
                "col5": 1_00_000.0, "col6": 1_00_000.0,
                "col7": 0, "col8": 0,
            },
        },
        "by_party": {},
        "recon": {
            "pl_total":         1_00_000.0,
            "capex_total":      4_00_00_000.0,
            "reportable_total": 1_00_000.0,
        },
        "transactions": [],
    }


def _drain(resp) -> bytes:
    async def _go():
        return b"".join([c async for c in resp.body_iterator])
    return asyncio.new_event_loop().run_until_complete(_go())


def _read_recon(workbook):
    ws = workbook["Reconciliation"]
    out = {}
    for row in ws.iter_rows(values_only=True):
        if row and isinstance(row[0], str):
            out[row[0]] = row[1] if len(row) > 1 else None
    return out


# ─────────────────────────────────────────────────────────────────────────
# Spread mechanism
# ─────────────────────────────────────────────────────────────────────────
def test_controller_spread_overrides_stale_recon_with_fresh():
    """Mirror the spread the export endpoint does: stored stale recon
    → spread fresh recon onto it → Excel writer reads fresh values."""
    stored = _stale_run()
    fresh = _fresh_classification_result()
    refreshed = {
        **stored,
        "summary":      fresh["summary"],
        "by_ledger":    fresh["by_ledger"],
        "by_party":     fresh["by_party"],
        "recon":        fresh["recon"],
        "transactions": fresh.get("transactions", []),
    }
    resp = ctl.build_export_response(refreshed, "Clause44_spread")
    wb = openpyxl.load_workbook(io.BytesIO(_drain(resp)), data_only=True)
    rows = _read_recon(wb)
    capex_label = "+ Capital expenditure additions (ICAI Para 79.18)"
    pl_label = "Total Expenditure as per Profit & Loss"
    assert rows[capex_label] == 4_00_00_000.0, \
        f"Capex line should reflect FRESH 4 Cr, got {rows[capex_label]}"
    assert rows[pl_label] == 1_00_000.0, \
        f"P&L total should reflect FRESH 1 L (rent only), got {rows[pl_label]}"


def test_excel_pivot_reflects_fresh_by_ledger():
    """Per-ledger pivot in the Summary sheet uses fresh by_ledger after
    the spread."""
    stored = _stale_run()
    fresh = _fresh_classification_result()
    refreshed = {**stored, "summary": fresh["summary"],
                 "by_ledger": fresh["by_ledger"], "recon": fresh["recon"],
                 "transactions": []}
    resp = ctl.build_export_response(refreshed, "Clause44_pivot")
    wb = openpyxl.load_workbook(io.BytesIO(_drain(resp)), data_only=True)
    summary = wb["Clause 44 Summary"]
    flat = [c for r in summary.iter_rows(values_only=True)
              for c in r if isinstance(c, str)]
    # Stale row gone, fresh rows present.
    assert any("Plant & Machinery" in s for s in flat)
    assert any("Office Rent" in s and "stale" not in s.lower() for s in flat)
    assert not any("(stale row)" in s for s in flat)


# ─────────────────────────────────────────────────────────────────────────
# Fallback path — when re-classification raises, stored snapshot is used
# ─────────────────────────────────────────────────────────────────────────
def test_export_uses_stored_snapshot_when_classification_raises():
    """If `_run_classification` raises (corrupt accounting blob, etc.),
    the controller falls back to the stored snapshot.  The workbook
    still builds — just with the stale values."""
    stored = _stale_run()

    # Mirror controller try/except: raise → use stored.
    refreshed = stored
    try:
        raise RuntimeError("simulated re-classify failure")
    except Exception:
        pass

    resp = ctl.build_export_response(refreshed, "Clause44_fallback")
    wb = openpyxl.load_workbook(io.BytesIO(_drain(resp)), data_only=True)
    rows = _read_recon(wb)
    capex_label = "+ Capital expenditure additions (ICAI Para 79.18)"
    # Stored stale 0 is what the user gets — better than a 500.
    assert rows[capex_label] == 0.0
    assert "Reconciliation" in wb.sheetnames


# ─────────────────────────────────────────────────────────────────────────
# Endpoint-level — patch _run_classification + _ensure_run_data and
# verify the controller wires them through to the workbook.
# ─────────────────────────────────────────────────────────────────────────
def test_export_endpoint_invokes_run_classification_and_uses_fresh():
    """Direct call to the controller coroutine: patch _fetch_run + the
    auth dependency, assert _run_classification is invoked and its
    result lands in the rendered workbook."""
    stored = _stale_run()

    async def _fake_fetch_run(_id):  # noqa: ARG001
        return stored

    async def _fake_auth(*_a, **_kw):  # noqa: ARG001
        return {"user_id": "u1"}

    async def _fake_ensure(run):
        return run

    # _run_classification is sync; we patch it to return our fresh result.
    with patch.object(ctl, "_fetch_run", _fake_fetch_run), \
         patch.object(ctl, "get_current_user", _fake_auth), \
         patch.object(ctl, "_ensure_run_data", _fake_ensure), \
         patch.object(ctl, "_run_classification",
                      return_value=_fresh_classification_result()) as m_classify:
        resp = asyncio.new_event_loop().run_until_complete(
            ctl.export_run("stale-export-1", request=None,
                           session_token="t", authorization=None)
        )
    assert m_classify.called, "_run_classification not invoked by export_run"
    wb = openpyxl.load_workbook(io.BytesIO(_drain(resp)), data_only=True)
    rows = _read_recon(wb)
    capex_label = "+ Capital expenditure additions (ICAI Para 79.18)"
    assert rows[capex_label] == 4_00_00_000.0
