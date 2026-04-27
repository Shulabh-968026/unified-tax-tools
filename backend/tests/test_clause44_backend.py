"""End-to-end backend tests for Clause 44 Report Builder."""
import io
import os
import pytest
import requests
import openpyxl


# ----- Auth -----------------------------------------------------------------

class TestAuth:
    def test_me_unauthenticated(self, base_url):
        r = requests.get(f"{base_url}/api/auth/me")
        assert r.status_code == 401

    def test_me_authenticated(self, base_url, auth_headers, session_info):
        r = requests.get(f"{base_url}/api/auth/me", headers=auth_headers)
        assert r.status_code == 200
        data = r.json()
        assert data["user_id"] == session_info["user_id"]
        assert "email" in data and "name" in data

    def test_runs_requires_auth(self, base_url):
        r = requests.get(f"{base_url}/api/runs")
        assert r.status_code == 401

    def test_root(self, base_url):
        r = requests.get(f"{base_url}/api/")
        assert r.status_code == 200
        assert r.json().get("ok") is True


# ----- Run upload + state ---------------------------------------------------

@pytest.fixture(scope="module")
def created_run(base_url, auth_headers, sample_files):
    # Bootstrap a single-division client for iter1 regression tests
    cr = requests.post(
        f"{base_url}/api/clients",
        headers=auth_headers,
        json={"file_number": "TEST_ITER1_FN", "name": "TEST_Iter1_Client", "type": "single"},
    )
    assert cr.status_code == 200, cr.text
    client_id = cr.json()["client_id"]
    with open(sample_files["json"], "rb") as fj, open(sample_files["xlsx"], "rb") as fx:
        files = {
            "accounting_json": ("accounting.json", fj, "application/json"),
            "ledger_xlsx": ("ledger.xlsx", fx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        }
        data = {"client_id": client_id, "period": "FY2023-24"}
        r = requests.post(f"{base_url}/api/runs", headers=auth_headers, files=files, data=data, timeout=120)
    assert r.status_code == 200, r.text
    return r.json()


class TestRunUpload:
    def test_upload_returns_expected_structure(self, created_run):
        d = created_run
        assert "run_id" in d and d["run_id"].startswith("run_")
        assert d["vouchers_count"] > 0
        assert d["ledgers_count"] > 0
        assert isinstance(d["itc_candidates"], list) and len(d["itc_candidates"]) > 0
        assert isinstance(d["pl_ledgers"], list) and len(d["pl_ledgers"]) > 0
        # suggested flags exist
        assert any("suggested" in c for c in d["itc_candidates"])
        assert any("suggested" in p for p in d["pl_ledgers"])
        # at least one P&L ledger should be auto-suggested for exclusion
        assert any(p.get("suggested") for p in d["pl_ledgers"])

    def test_list_runs_includes_created(self, base_url, auth_headers, created_run):
        r = requests.get(f"{base_url}/api/runs", headers=auth_headers)
        assert r.status_code == 200
        ids = [run["run_id"] for run in r.json()["runs"]]
        assert created_run["run_id"] in ids

    def test_get_run_full_state(self, base_url, auth_headers, created_run):
        r = requests.get(f"{base_url}/api/runs/{created_run['run_id']}", headers=auth_headers)
        assert r.status_code == 200
        d = r.json()
        assert d["run_id"] == created_run["run_id"]
        assert d["generated"] is False
        assert len(d["itc_candidates"]) == len(created_run["itc_candidates"])
        assert len(d["pl_ledgers"]) == len(created_run["pl_ledgers"])

    def test_runs_archived_filter(self, base_url, auth_headers, created_run):
        r = requests.get(f"{base_url}/api/runs?archived=true", headers=auth_headers)
        assert r.status_code == 200
        ids = [run["run_id"] for run in r.json()["runs"]]
        assert created_run["run_id"] not in ids


# ----- Generate -------------------------------------------------------------

@pytest.fixture(scope="module")
def generated_run(base_url, auth_headers, created_run):
    itc = [c["name"] for c in created_run["itc_candidates"] if c.get("suggested")]
    excluded = [p["name"] for p in created_run["pl_ledgers"] if p.get("suggested")]
    body = {"itc_ledgers": itc, "excluded_ledgers": excluded}
    r = requests.post(f"{base_url}/api/runs/{created_run['run_id']}/generate",
                      headers=auth_headers, json=body, timeout=120)
    assert r.status_code == 200, r.text
    return {"run_id": created_run["run_id"], "data": r.json(), "itc": itc, "excluded": excluded}


class TestGenerate:
    def test_summary_has_nonzero_buckets(self, generated_run):
        s = generated_run["data"]["summary"]
        for k in ("col2_total", "col3", "col4", "col5", "col6", "col7"):
            assert k in s
        # col5 must be non-zero (ITC bucket) and col7 (unregistered) generally non-zero
        assert s["col2_total"] > 0
        assert s["col5"] > 0
        assert s["col7"] > 0

    def test_summary_arithmetic(self, generated_run):
        s = generated_run["data"]["summary"]
        # col6 == col3+col4+col5, col2_total == col3+col4+col5+col7 (within float epsilon)
        assert abs(s["col6"] - (s["col3"] + s["col4"] + s["col5"])) < 0.5
        assert abs(s["col2_total"] - (s["col3"] + s["col4"] + s["col5"] + s["col7"])) < 0.5

    def test_by_ledger_present(self, generated_run):
        bl = generated_run["data"]["by_ledger"]
        assert isinstance(bl, dict)
        assert len(bl) > 0
        any_row = next(iter(bl.values()))
        for k in ("col3", "col4", "col5", "col7", "total"):
            assert k in any_row


# ----- Transactions ---------------------------------------------------------

class TestTransactions:
    def test_col5_filter(self, base_url, auth_headers, generated_run):
        r = requests.get(f"{base_url}/api/runs/{generated_run['run_id']}/transactions?bucket=col5",
                         headers=auth_headers)
        assert r.status_code == 200
        txns = r.json()["transactions"]
        assert len(txns) > 0
        assert all(t["bucket"] == "col5" for t in txns)
        assert all(t.get("reason") for t in txns)

    def test_col6_returns_3_4_5(self, base_url, auth_headers, generated_run):
        r = requests.get(f"{base_url}/api/runs/{generated_run['run_id']}/transactions?bucket=col6",
                         headers=auth_headers)
        assert r.status_code == 200
        txns = r.json()["transactions"]
        assert all(t["bucket"] in ("col3", "col4", "col5") for t in txns)

    def test_all_bucket(self, base_url, auth_headers, generated_run):
        r = requests.get(f"{base_url}/api/runs/{generated_run['run_id']}/transactions?bucket=all",
                         headers=auth_headers)
        assert r.status_code == 200
        all_txns = r.json()["transactions"]
        assert len(all_txns) > 0
        buckets = {t["bucket"] for t in all_txns}
        assert "col5" in buckets and "col7" in buckets

    def test_ledger_filter(self, base_url, auth_headers, generated_run):
        # pick a ledger that exists in by_ledger
        bl = generated_run["data"]["by_ledger"]
        ledger = next(iter(bl.keys()))
        r = requests.get(
            f"{base_url}/api/runs/{generated_run['run_id']}/transactions",
            headers=auth_headers, params={"bucket": "all", "ledger": ledger})
        assert r.status_code == 200
        txns = r.json()["transactions"]
        assert len(txns) > 0
        assert all(t["ledger_name"] == ledger for t in txns)


# ----- Export ---------------------------------------------------------------

class TestExport:
    def test_export_xlsx(self, base_url, auth_headers, generated_run):
        r = requests.get(f"{base_url}/api/runs/{generated_run['run_id']}/export",
                         headers=auth_headers, timeout=60)
        assert r.status_code == 200
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml.sheet" in ct
        assert len(r.content) > 5000
        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=False)
        assert "Clause 44" in wb.sheetnames
        assert "Transaction Audit Trail" in wb.sheetnames
        ws2 = wb["Transaction Audit Trail"]
        # Header + at least 1 row
        assert ws2.max_row > 1
        # 'Reason' is column 10 (J)
        reason_cells = [ws2.cell(row=i, column=10).value for i in range(2, min(ws2.max_row + 1, 20))]
        assert any(reason_cells), "Audit trail rows should have classification reasons"


# ----- Archive --------------------------------------------------------------

class TestArchive:
    def test_archive_toggles(self, base_url, auth_headers, generated_run):
        rid = generated_run["run_id"]
        r1 = requests.post(f"{base_url}/api/runs/{rid}/archive", headers=auth_headers)
        assert r1.status_code == 200
        assert r1.json()["archived"] is True
        # Toggle back
        r2 = requests.post(f"{base_url}/api/runs/{rid}/archive", headers=auth_headers)
        assert r2.status_code == 200
        assert r2.json()["archived"] is False
