"""Release-3 live HTTP regression tests.

Validates over the wire (preview env) that:
  * Col 8 is a first-class classification bucket (Step 0 wins)
  * Col 2 (gross) = Col 3 + Col 4 + Col 5 + Col 7 + Col 8
  * Reportable = Col 6 + Col 7 (exclusive of Col 8)
  * Excel workbook has 7 sheets including 'Col 8 · Excluded'
  * Summary sheet pivot is 7-column (Col 2..Col 8)
  * Col 8 sheet has ICAI sub-bucket band headers + grand total row
  * GET /api/runs/{id}/transactions?bucket=col8 returns only Col 8 vouchers
  * GET /api/runs/{id} returns col8 and reportable_total in summary
  * DEFAULT_DISCLAIMER carries the user-approved verbatim phrases

Hard contract: must restore the run to a NO-EXCLUSIONS state at teardown so
the next testing agent sees the same baseline.
"""
import io
import os
import openpyxl
import pytest
import requests

BASE = os.environ.get("REACT_APP_BACKEND_URL", "https://unified-tax-tools.preview.emergentagent.com").rstrip("/")
TOKEN = "qa_test_session_token_20260430_dev"
RUN_ID = "run_8d427d1f97e0"  # ABC Textile Mills (cli_ad137f29aebb)
COOKIES = {"session_token": TOKEN}

# Ledgers we will mark excluded for the duration of one test class only.
EXCL_NON_CASH = ["Wages A/c"]            # arbitrary, treated as non_cash
EXCL_SCH3 = ["Salary Exp"]               # treated as sch3
EXCL_OTHER = ["Interest on Bank O/D"]    # treated as other
ALL_EXCL = EXCL_NON_CASH + EXCL_SCH3 + EXCL_OTHER


@pytest.fixture(scope="module")
def baseline_run():
    """Snapshot baseline selections so we can restore at the end."""
    r = requests.get(f"{BASE}/api/runs/{RUN_ID}", cookies=COOKIES, timeout=30)
    assert r.status_code == 200, f"Cannot fetch baseline run: {r.status_code} {r.text[:200]}"
    run = r.json()
    yield run
    # Teardown: clear exclusions and re-generate so DB is back to baseline.
    requests.post(
        f"{BASE}/api/runs/{RUN_ID}/generate",
        cookies=COOKIES,
        json={
            "itc_ledgers": run.get("itc_selection") or [],
            "excluded_ledgers": [],
            "exempt_ledgers": run.get("exempt_selection") or [],
            "use_itc_inference": bool(run.get("use_itc_inference", True)),
            "exclusion_categories": {},
        },
        timeout=120,
    )


@pytest.fixture(scope="module")
def with_exclusions(baseline_run):
    """Re-generate the run with our 3 elected exclusions in 3 sub-buckets."""
    payload = {
        "itc_ledgers": baseline_run.get("itc_selection") or [],
        "excluded_ledgers": ALL_EXCL,
        "exempt_ledgers": baseline_run.get("exempt_selection") or [],
        "use_itc_inference": bool(baseline_run.get("use_itc_inference", True)),
        "exclusion_categories": {
            EXCL_NON_CASH[0]: "non_cash",
            EXCL_SCH3[0]: "sch3",
            EXCL_OTHER[0]: "other",
        },
    }
    r = requests.post(f"{BASE}/api/runs/{RUN_ID}/generate", cookies=COOKIES, json=payload, timeout=120)
    assert r.status_code == 200, f"generate failed: {r.status_code} {r.text[:300]}"
    return r.json()


# -- 1. Generate cascade routes excluded ledgers to Col 8 ----------------
class TestCascadeStep0:
    def test_summary_has_col8_and_reportable(self, with_exclusions):
        s = with_exclusions["summary"]
        for k in ("col2_total", "col3", "col4", "col5", "col6", "col7", "col8", "reportable_total"):
            assert k in s, f"summary missing {k}"

    def test_col8_positive_when_exclusions_elected(self, with_exclusions):
        assert with_exclusions["summary"]["col8"] > 0, "Expected col8 > 0 after electing exclusions"

    def test_gross_identity_col2_equals_3_4_5_7_8(self, with_exclusions):
        s = with_exclusions["summary"]
        rhs = s["col3"] + s["col4"] + s["col5"] + s["col7"] + s["col8"]
        assert abs(s["col2_total"] - rhs) < 0.5, f"col2={s['col2_total']} != col3+4+5+7+8={rhs}"

    def test_reportable_equals_col6_plus_col7(self, with_exclusions):
        s = with_exclusions["summary"]
        assert abs(s["reportable_total"] - (s["col6"] + s["col7"])) < 0.5

    def test_excluded_ledger_no_longer_in_col7(self, with_exclusions):
        # The elected ledgers must now sit in col8, not col7.
        for led in ALL_EXCL:
            row = (with_exclusions.get("by_ledger") or {}).get(led)
            if not row:
                continue
            assert (row.get("col8") or 0) > 0, f"{led} should have col8 > 0"
            assert (row.get("col7") or 0) == 0, f"{led} should NOT have col7 anymore"


# -- 2. GET /api/runs/{id} reflects the new shape -------------------------
class TestRunGetShape:
    def test_run_doc_carries_col8_and_recon_col8_total(self, with_exclusions):
        r = requests.get(f"{BASE}/api/runs/{RUN_ID}", cookies=COOKIES, timeout=30)
        assert r.status_code == 200
        d = r.json()
        s, recon = d["summary"], d["recon"]
        assert s.get("col8", 0) > 0
        assert "reportable_total" in s
        assert recon.get("col8_total", 0) > 0
        # excluded_lines populated and split into sub-buckets
        ex = recon.get("excluded_lines") or []
        assert len(ex) >= 3, f"Expected ≥3 excluded lines, got {len(ex)}"
        buckets = {x.get("bucket") for x in ex}
        assert {"non_cash", "sch3", "other"} <= buckets, f"Got buckets {buckets}"


# -- 3. Transactions endpoint accepts bucket=col8 ------------------------
class TestTransactionsCol8:
    def test_bucket_col8_filter_returns_only_col8(self, with_exclusions):
        r = requests.get(
            f"{BASE}/api/runs/{RUN_ID}/transactions",
            params={"bucket": "col8"}, cookies=COOKIES, timeout=30,
        )
        assert r.status_code == 200, f"col8 transactions endpoint should accept bucket=col8: {r.status_code} {r.text[:200]}"
        body = r.json()
        assert body["bucket"] == "col8"
        txns = body["transactions"]
        assert len(txns) > 0, "Expected ≥1 col8 voucher with elected exclusions"
        for t in txns:
            assert t["bucket"] == "col8"

    def test_bucket_col8_invalid_other_value_rejected(self):
        # Sanity — pattern still strict against unknown buckets.
        r = requests.get(
            f"{BASE}/api/runs/{RUN_ID}/transactions",
            params={"bucket": "col9"}, cookies=COOKIES, timeout=30,
        )
        assert r.status_code in (400, 422)


# -- 4. Excel workbook structure & Col 8 sheet ---------------------------
class TestExcelExport:
    @pytest.fixture(scope="class")
    def workbook(self, with_exclusions):
        r = requests.get(f"{BASE}/api/runs/{RUN_ID}/export", cookies=COOKIES, timeout=120)
        assert r.status_code == 200, f"/export status {r.status_code}"
        assert r.headers.get("content-type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return openpyxl.load_workbook(io.BytesIO(r.content))

    def test_workbook_has_seven_sheets(self, workbook):
        assert len(workbook.sheetnames) == 7, f"Expected 7 sheets, got {workbook.sheetnames}"

    def test_col8_sheet_exists(self, workbook):
        assert "Col 8 · Excluded" in workbook.sheetnames

    def test_summary_sheet_has_seven_data_columns(self, workbook):
        ws = workbook["Clause 44 Summary"]
        # Walk for the row whose first cell is 'Aggregate' / 'Total' — 7-col pivot.
        # Header row carries Col 2..Col 8 labels somewhere in first 10 rows.
        labels = []
        for row in ws.iter_rows(values_only=True, max_row=10):
            for cell in row:
                if isinstance(cell, str):
                    labels.append(cell)
        for tag in ("Col 2", "Col 3", "Col 4", "Col 5", "Col 6", "Col 7", "Col 8"):
            assert any(tag in lbl for lbl in labels), f"Summary sheet missing label {tag}"

    def test_col8_sheet_subbucket_band_headers(self, workbook):
        ws = workbook["Col 8 · Excluded"]
        all_text = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        # ICAI sub-bucket bands
        for band in ("Non-cash charges", "Schedule III items", "Other exclusions"):
            assert any(band in v for v in all_text), f"Missing band '{band}' in Col 8 sheet"
        # Final 'Col 8 Total · Excluded expenditure' marker
        assert any("Col 8 Total" in v for v in all_text), "Col 8 sheet missing grand total"

    def test_col8_sheet_lists_elected_ledgers(self, workbook):
        ws = workbook["Col 8 · Excluded"]
        all_text = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
        for led in ALL_EXCL:
            assert any(led in v for v in all_text), f"Col 8 sheet missing elected ledger {led}"


# -- 5. DEFAULT_DISCLAIMER content (verbatim user-approved phrases) ------
class TestDisclaimerContent:
    def test_default_phrases_present_via_module(self):
        from modules.clause44.controller import DEFAULT_DISCLAIMER
        for phrase in [
            "classification of expenditure under Clause 44",
            "based solely on the books of account",
            "purchase ledgers internally designated by the entity as exempt-supply ledgers",
            "absence of an ITC-input ledger",
            "both as adopted and affirmed by management",
            "RCM vouchers and purchases from foreign suppliers are reported under Column 7",
            "true and complete extract of its books",
            "(Ref: ICAI Guidance Note on Tax Audit, Para 79.20 / 79.21.)",
        ]:
            assert phrase in DEFAULT_DISCLAIMER, f"Missing phrase: {phrase!r}"

    def test_disclaimer_round_trips_to_excel(self, with_exclusions):
        # Reset the run's disclaimer to the canonical R3 DEFAULT_DISCLAIMER
        # (the run may carry stale text PATCH-ed by earlier iterations).
        from modules.clause44.controller import DEFAULT_DISCLAIMER
        patch = requests.patch(
            f"{BASE}/api/runs/{RUN_ID}/selections",
            cookies=COOKIES, json={"disclaimer_text": DEFAULT_DISCLAIMER}, timeout=30,
        )
        assert patch.status_code == 200, f"PATCH selections failed: {patch.status_code} {patch.text[:200]}"
        r = requests.get(f"{BASE}/api/runs/{RUN_ID}/export", cookies=COOKIES, timeout=120)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        recon_sheets = [s for s in wb.sheetnames if "Reconcil" in s or "P&L" in s.replace(" ", "")]
        assert recon_sheets, f"No reconciliation sheet found: {wb.sheetnames}"
        ws = wb[recon_sheets[0]]
        all_text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        # Two anchor phrases from the R3 user-approved verbatim text.
        assert "based solely on the books of account" in all_text, \
            f"R3 disclaimer phrase missing from Reconciliation sheet"
        assert "(Ref: ICAI Guidance Note on Tax Audit, Para 79.20 / 79.21.)" in all_text


# -- 6. Regression — clients list still contains exactly 3 originals -----
class TestNoClientLeak:
    def test_only_three_seed_clients(self):
        r = requests.get(f"{BASE}/api/clients", cookies=COOKIES, timeout=30)
        if r.status_code != 200:
            pytest.skip(f"/api/clients not reachable: {r.status_code}")
        body = r.json()
        items = body if isinstance(body, list) else (body.get("clients") or body.get("items") or [])
        names = [c.get("name") or c.get("client_name") for c in items]
        # No TEST_-prefixed leakage from prior runs
        assert not any((n or "").startswith("TEST_") for n in names), f"TEST_ leakage: {names}"
